import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
import plotly.figure_factory as ff
from plotly.subplots import make_subplots
import scipy.stats as stats
from scipy import stats
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan, het_white
from statsmodels.stats.stattools import durbin_watson, jarque_bera
from statsmodels.stats.outliers_influence import variance_inflation_factor
from statsmodels.tsa.stattools import acf
from sklearn.preprocessing import StandardScaler, PolynomialFeatures
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import cross_val_score
from sklearn.linear_model import Ridge, Lasso
import io
import base64
import pickle
import json
from datetime import datetime
import gc  # For garbage collection
import warnings
warnings.filterwarnings('ignore')

# Set page configuration
st.set_page_config(
    page_title="Econometrics Learning Lab",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for complete styling
st.markdown("""
<style>
    /* Landing Page Styles */
    .landing-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 20px;
        margin: 1rem 0;
        box-shadow: 0 10px 30px rgba(0,0,0,0.3);
        color: white;
        text-align: center;
    }
    .app-title {
        font-size: 3.5rem;
        font-weight: bold;
        margin-bottom: 1rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        background: linear-gradient(45deg, #FFD700, #FFA500);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .subtitle {
        font-size: 1.5rem;
        margin-bottom: 2rem;
        opacity: 0.9;
    }
    .creator-badge {
        background: linear-gradient(45deg, #FF6B6B, #4ECDC4);
        padding: 1rem 2rem;
        border-radius: 50px;
        font-size: 1.3rem;
        font-weight: bold;
        margin: 2rem auto;
        display: inline-block;
        box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        text-shadow: 1px 1px 2px rgba(0,0,0,0.3);
    }
    .feature-box {
        background: rgba(255,255,255,0.1);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.2);
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1rem;
        transition: transform 0.3s ease;
    }
    .feature-box:hover {
        transform: translateY(-5px);
        box-shadow: 0 10px 25px rgba(0,0,0,0.2);
    }
    .math-viz {
        background: rgba(255,255,255,0.95);
        border-radius: 15px;
        padding: 1rem;
        margin: 2rem 0;
        box-shadow: 0 5px 15px rgba(0,0,0,0.1);
    }
    .stats-badge {
        background: linear-gradient(45deg, #667eea, #764ba2);
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: bold;
        margin: 0.5rem;
        display: inline-block;
    }
    
    /* Main App Styles */
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
    }
    .section-header {
        font-size: 1.5rem;
        color: #2e8b57;
        margin: 1rem 0;
        border-bottom: 2px solid #2e8b57;
        padding-bottom: 0.5rem;
    }
    .formula-box {
        background-color: #f0f8ff;
        border: 2px solid #4169e1;
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .interpretation-box {
        background-color: #f5f5dc;
        border: 2px solid #daa520;
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .metric-box {
        background-color: #e6f3ff;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid #1f77b4;
    }
    .creator-label {
        position: fixed;
        bottom: 10px;
        right: 10px;
        background-color: rgba(0,0,0,0.7);
        color: white;
        padding: 5px 10px;
        border-radius: 5px;
        font-size: 0.8rem;
        z-index: 999;
    }
    
    /* RED COLOR POLICY FOR SPECIFIED TEXT */
    .red-header {
        color: #dc143c !important;
        font-weight: bold;
        font-size: 1.2rem;
    }
    .red-text {
        color: #dc143c !important;
        font-weight: bold;
    }
    .red-list {
        color: #dc143c !important;
    }
    .red-hypothesis {
        color: #dc143c !important;
        font-weight: bold;
    }
    .red-interpretation {
        color: #dc143c !important;
    }
    
    /* Enhanced Features Styles */
    .data-table-container {
        background: white;
        border-radius: 10px;
        padding: 1rem;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }
    .advanced-metric {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        margin: 0.5rem 0;
    }
    .upload-progress {
        background: linear-gradient(45deg, #4CAF50, #45a049);
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Creator label
st.markdown("""
<div class="creator-label">
    Created by HAMDI Boulanouar
</div>
""", unsafe_allow_html=True)

# Language selection
def get_language():
    return st.sidebar.selectbox(
        "🌐 Language / اللغة",
        ["English", "العربية"],
        help="Select your preferred language"
    )

# Translation dictionary
translations = {
    "English": {
        "title": "📊 Econometrics Learning Laboratory",
        "subtitle": "Complete Linear Regression Analysis Tool for Students",
        "upload_data": "📁 Upload Your Dataset",
        "data_preview": "👀 Data Preview",
        "variable_selection": "🎯 Variable Selection",
        "regression_analysis": "📈 Regression Analysis",
        "diagnostics": "🔍 Model Diagnostics",
        "predictions": "🔮 Predictions",
        "dependent_var": "Dependent Variable (Y)",
        "independent_vars": "Independent Variables (X)",
        "run_regression": "Run Regression Analysis",
        "simple_regression": "Simple Linear Regression",
        "multiple_regression": "Multiple Linear Regression"
    },
    "العربية": {
        "title": "📊 مختبر تعلم الاقتصاد القياسي",
        "subtitle": "أداة تحليل الانحدار الخطي الشاملة للطلاب",
        "upload_data": "📁 تحميل البيانات",
        "data_preview": "👀 معاينة البيانات",
        "variable_selection": "🎯 اختيار المتغيرات",
        "regression_analysis": "📈 تحليل الانحدار",
        "diagnostics": "🔍 تشخيص النموذج",
        "predictions": "🔮 التنبؤات",
        "dependent_var": "المتغير التابع (Y)",
        "independent_vars": "المتغيرات المستقلة (X)",
        "run_regression": "تشغيل تحليل الانحدار",
        "simple_regression": "الانحدار الخطي البسيط",
        "multiple_regression": "الانحدار الخطي المتعدد"
    }
}

# Enhanced session state initialization with memory management
def initialize_session_state():
    """Initialize session state with memory management"""
    
    # Core data storage
    if 'data' not in st.session_state:
        st.session_state.data = None
    
    if 'regression_results' not in st.session_state:
        st.session_state.regression_results = None
    
    # Variable selections
    if 'dependent_var' not in st.session_state:
        st.session_state.dependent_var = None
    
    if 'independent_vars' not in st.session_state:
        st.session_state.independent_vars = []
    
    # Analysis settings
    if 'analysis_settings' not in st.session_state:
        st.session_state.analysis_settings = {
            'confidence_level': 95,
            'regression_type': 'OLS',
            'include_diagnostics': True,
            'language': 'English'
        }
    
    # File upload tracking
    if 'uploaded_file_name' not in st.session_state:
        st.session_state.uploaded_file_name = None
    
    if 'data_hash' not in st.session_state:
        st.session_state.data_hash = None
    
    # Page state tracking
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "Data Upload & Preview"
    
    # Analysis history (limit to 10 to save memory)
    if 'analysis_history' not in st.session_state:
        st.session_state.analysis_history = []
    
    # Saved analyses (limit to 5 to save memory)
    if 'saved_analyses' not in st.session_state:
        st.session_state.saved_analyses = []
    
    # Landing page control
    if 'show_landing' not in st.session_state:
        st.session_state.show_landing = True
    
    # Clean up memory periodically
    if len(st.session_state.get('saved_analyses', [])) > 5:
        st.session_state.saved_analyses = st.session_state.saved_analyses[-5:]
    
    if len(st.session_state.get('analysis_history', [])) > 10:
        st.session_state.analysis_history = st.session_state.analysis_history[-10:]

# Fixed file upload function - NO MORE HANGING!
def cache_uploaded_data(file_content, file_name):
    """Load uploaded data without caching issues - FIXED VERSION"""
    try:
        # Clear any existing large objects from memory
        gc.collect()
        
        # Process file with size limits
        file_size_mb = len(file_content) / (1024 * 1024)
        
        if file_size_mb > 50:  # Limit to 50MB
            st.error(f"⚠️ File too large ({file_size_mb:.1f}MB). Please use files smaller than 50MB.")
            return None, None
        
        # Load data based on file type
        if file_name.lower().endswith('.csv'):
            # Use chunked reading for large CSV files
            try:
                data = pd.read_csv(
                    io.StringIO(file_content.decode('utf-8')),
                    low_memory=False,
                    encoding='utf-8'
                )
            except UnicodeDecodeError:
                # Try different encodings
                try:
                    data = pd.read_csv(
                        io.StringIO(file_content.decode('latin-1')),
                        low_memory=False,
                        encoding='latin-1'
                    )
                except:
                    data = pd.read_csv(
                        io.StringIO(file_content.decode('cp1252')),
                        low_memory=False,
                        encoding='cp1252'
                    )
        else:
            # Excel files
            data = pd.read_excel(
                io.BytesIO(file_content),
                engine='openpyxl'
            )
        
        # Limit number of rows for performance
        if len(data) > 10000:
            st.warning(f"⚠️ Large dataset ({len(data)} rows). Using first 10,000 rows for better performance.")
            data = data.head(10000)
        
        # Simple hash without converting entire dataset
        data_hash = hash(str(data.shape) + str(data.columns.tolist()[:5]))  # Only first 5 columns
        
        # Basic data cleaning
        data = data.dropna(how='all')  # Remove completely empty rows
        data = data.loc[:, ~data.columns.str.contains('^Unnamed')]  # Remove unnamed columns
        
        return data, data_hash
        
    except Exception as e:
        st.error(f"❌ Error loading data: {str(e)}")
        return None, None

def optimize_memory_usage(df):
    """Optimized memory usage function"""
    try:
        original_memory = df.memory_usage(deep=True).sum() / (1024**2)
        
        # Only optimize if dataframe is reasonably sized
        if original_memory < 100:  # Less than 100MB
            for col in df.columns:
                if df[col].dtype == 'float64':
                    df[col] = pd.to_numeric(df[col], downcast='float', errors='ignore')
                elif df[col].dtype == 'int64':
                    df[col] = pd.to_numeric(df[col], downcast='integer', errors='ignore')
        
        return df
    except:
        return df  # Return original if optimization fails

# Professional landing page
def create_landing_page():
    """Create a professional landing page with visual elements"""
    
    # Main landing container
    st.markdown("""
    <div class="landing-container">
        <h1 class="app-title">📊 ECONOMETRICS LEARNING LAB</h1>
        <p class="subtitle">Professional Linear Regression Analysis Tool for Academic Excellence</p>
        <div class="creator-badge">
            👨‍💼 Created by HAMDI Boulanouar 👨‍💼
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Create visual regression demonstration
    create_regression_visualization()
    
    # Feature highlights
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-box">
            <h3>🎯 Complete Analysis</h3>
            <p>• Simple & Multiple Regression<br>
            • All Statistical Tests<br>
            • Diagnostic Testing<br>
            • Professional Reports</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-box">
            <h3>📚 Educational Focus</h3>
            <p>• Step-by-step Explanations<br>
            • Mathematical Foundations<br>
            • Interactive Learning<br>
            • Bilingual Support</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-box">
            <h3>🔬 Advanced Features</h3>
            <p>• Model Comparison<br>
            • Time Series Analysis<br>
            • Export Capabilities<br>
            • Professional Visualizations</p>
        </div>
        """, unsafe_allow_html=True)
    
    # App statistics
    st.markdown("""
    <div style="text-align: center; margin: 2rem 0;">
        <span class="stats-badge">🔢 15+ Statistical Tests</span>
        <span class="stats-badge">📊 Multiple Model Types</span>
        <span class="stats-badge">🌍 2 Languages</span>
        <span class="stats-badge">📈 Real-time Analysis</span>
        <span class="stats-badge">💾 Data Persistence</span>
        <span class="stats-badge">📤 Export Options</span>
    </div>
    """, unsafe_allow_html=True)
    
    # Enter application button
    st.markdown('<div style="text-align: center;">', unsafe_allow_html=True)
    if st.button("🚀 ENTER APPLICATION", key="enter_app", help="Click to start your econometrics analysis"):
        st.session_state.show_landing = False
        st.experimental_rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer with professional info
    st.markdown("""
    <div style="text-align: center; padding: 2rem; opacity: 0.7;">
        <p><strong>🎓 Professional Econometrics Education Platform</strong></p>
        <p>Designed for Students, Researchers, and Professionals</p>
        <p><em>Making Advanced Statistics Accessible to Everyone</em></p>
    </div>
    """, unsafe_allow_html=True)

def create_regression_visualization():
    """Create an attractive regression visualization for the landing page"""
    
    st.markdown('<div class="math-viz">', unsafe_allow_html=True)
    
    # Generate sample data for visualization
    np.random.seed(42)
    x = np.linspace(0, 10, 50)
    y = 2 + 1.5 * x + np.random.normal(0, 1, 50)
    
    # Create the plot
    fig = go.Figure()
    
    # Add scatter points
    fig.add_trace(go.Scatter(
        x=x, y=y,
        mode='markers',
        name='Data Points',
        marker=dict(
            size=10,
            color='rgba(102, 126, 234, 0.8)',
            line=dict(width=2, color='rgba(102, 126, 234, 1)')
        ),
        hovertemplate='X: %{x:.2f}<br>Y: %{y:.2f}<extra></extra>'
    ))
    
    # Add regression line
    x_line = np.linspace(0, 10, 100)
    slope = np.polyfit(x, y, 1)[0]
    intercept = np.polyfit(x, y, 1)[1]
    y_line = intercept + slope * x_line
    
    fig.add_trace(go.Scatter(
        x=x_line, y=y_line,
        mode='lines',
        name='Regression Line',
        line=dict(color='rgba(255, 107, 107, 1)', width=4),
        hovertemplate=f'Regression Line<br>Y = {intercept:.2f} + {slope:.2f}X<extra></extra>'
    ))
    
    # Customize layout
    fig.update_layout(
        title={
            'text': "<b>📈 Linear Regression Demonstration</b><br><sub>Y = β₀ + β₁X + ε</sub>",
            'x': 0.5,
            'font': {'size': 20, 'color': '#2E4053'}
        },
        xaxis=dict(
            title="<b>Independent Variable (X)</b>",
            title_font=dict(size=14),
            showgrid=True,
            gridwidth=1,
            gridcolor='rgba(0,0,0,0.1)'
        ),
        yaxis=dict(
            title="<b>Dependent Variable (Y)</b>",
            title_font=dict(size=14),
            showgrid=True,
            gridwidth=1,
            gridcolor='rgba(0,0,0,0.1)'
        ),
        plot_bgcolor='rgba(248,249,250,0.8)',
        paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=12),
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01,
            bgcolor="rgba(255,255,255,0.8)",
            bordercolor="rgba(0,0,0,0.2)",
            borderwidth=1
        ),
        height=400,
        margin=dict(l=50, r=50, t=80, b=50)
    )
    
    # Add equation annotation
    fig.add_annotation(
        x=0.02, y=0.98,
        xref="paper", yref="paper",
        text=f"<b>Equation: Y = {intercept:.2f} + {slope:.2f}X</b><br>R² = {np.corrcoef(x, y)[0,1]**2:.3f}",
        showarrow=False,
        font=dict(size=12, color="#2E4053"),
        bgcolor="rgba(255,255,255,0.9)",
        bordercolor="rgba(102, 126, 234, 0.8)",
        borderwidth=2
    )
    
    st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

# Enhanced data display functions
def display_full_data_table(data, max_rows=None):
    """Display complete data table with pagination options"""
    
    st.markdown('<div class="data-table-container">', unsafe_allow_html=True)
    st.subheader("📋 Complete Dataset")
    
    # Display options
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        show_all = st.checkbox("📊 Show All Rows", value=False, key="show_all_data")
    
    with col2:
        if not show_all:
            rows_to_show = st.selectbox(
                "Rows to Display", 
                [10, 25, 50, 100, 200], 
                index=0,
                key="rows_display"
            )
        else:
            rows_to_show = len(data)
    
    with col3:
        start_row = st.number_input(
            "Start from Row", 
            min_value=1, 
            max_value=len(data), 
            value=1,
            key="start_row"
        ) - 1
    
    with col4:
        st.metric("Total Rows", len(data))
    
    # Display the data
    if show_all:
        st.info(f"📊 Showing all {len(data)} rows")
        st.dataframe(data, use_container_width=True, height=600)
    else:
        end_row = min(start_row + rows_to_show, len(data))
        st.info(f"📊 Showing rows {start_row + 1} to {end_row} of {len(data)}")
        st.dataframe(data.iloc[start_row:end_row], use_container_width=True, height=400)
    
    st.markdown('</div>', unsafe_allow_html=True)
    return show_all, rows_to_show

def generate_session_id():
    """Generate unique session ID"""
    import hashlib
    timestamp = str(datetime.now().timestamp())
    return hashlib.md5(timestamp.encode()).hexdigest()[:8]

def save_analysis_results():
    """Save complete analysis results to session state with timestamp"""
    
    if st.session_state.regression_results:
        # Create comprehensive results package
        results_package = {
            'model_summary': str(st.session_state.regression_results['model'].summary()),
            'coefficients': get_coefficients_data(),
            'model_stats': get_model_statistics(),
            'data_info': {
                'shape': st.session_state.data.shape,
                'columns': list(st.session_state.data.columns),
                'dependent_var': st.session_state.dependent_var,
                'independent_vars': st.session_state.independent_vars
            },
            'timestamp': datetime.now().isoformat(),
            'session_id': generate_session_id()
        }
        
        # Save to session state with history
        if 'saved_analyses' not in st.session_state:
            st.session_state.saved_analyses = []
        
        st.session_state.saved_analyses.append(results_package)
        
        # Keep only last 5 analyses to manage memory
        if len(st.session_state.saved_analyses) > 5:
            st.session_state.saved_analyses = st.session_state.saved_analyses[-5:]
        
        return True
    return False

def get_coefficients_data():
    """Extract coefficients data in a structured format"""
    model = st.session_state.regression_results['model']
    independent_vars = st.session_state.regression_results['independent_vars']
    
    try:
        if hasattr(model.params, 'index'):
            return {
                'variables': list(model.params.index),
                'coefficients': list(model.params.values),
                'std_errors': list(model.bse.values),
                't_statistics': list(model.tvalues.values),
                'p_values': list(model.pvalues.values),
                'conf_int_lower': list(model.conf_int()[0].values),
                'conf_int_upper': list(model.conf_int()[1].values)
            }
    except:
        return {
            'variables': ['Intercept'] + independent_vars,
            'coefficients': list(np.array(model.params).flatten()),
            'std_errors': list(np.array(model.bse).flatten()),
            't_statistics': list(np.array(model.tvalues).flatten()),
            'p_values': list(np.array(model.pvalues).flatten()),
            'conf_int_lower': list(np.array(model.conf_int()[0]).flatten()),
            'conf_int_upper': list(np.array(model.conf_int()[1]).flatten())
        }

def get_model_statistics():
    """Extract model statistics"""
    model = st.session_state.regression_results['model']
    
    return {
        'r_squared': model.rsquared,
        'adj_r_squared': model.rsquared_adj,
        'f_statistic': model.fvalue,
        'f_pvalue': model.f_pvalue,
        'aic': model.aic,
        'bic': model.bic,
        'observations': int(model.nobs),
        'df_residuals': int(model.df_resid)
    }

def reset_current_data():
    """Enhanced reset function with memory cleanup"""
    # Clear data-related session state
    st.session_state.data = None
    st.session_state.uploaded_file_name = None
    st.session_state.data_hash = None
    
    # Clear analysis results
    st.session_state.dependent_var = None
    st.session_state.independent_vars = []
    st.session_state.regression_results = None
    
    # Force garbage collection
    gc.collect()

def load_sample_data_enhanced(dataset_type, replace_current=True):
    """Enhanced sample data loading with replace option"""
    
    if not replace_current and st.session_state.data is not None:
        st.warning("⚠️ Cannot load sample data: Current data would be replaced. Check 'Replace current data' option.")
        return
    
    np.random.seed(42)
    
    if dataset_type == "wage":
        n = 500
        education = np.random.normal(12, 3, n)
        experience = np.random.exponential(8, n)
        gender = np.random.choice([0, 1], n, p=[0.6, 0.4])
        region = np.random.choice([1, 2, 3, 4], n)
        
        wage = (
            25000 +
            2500 * education +
            800 * experience +
            -100 * experience**2 / 10 +
            -3000 * gender +
            1000 * region +
            np.random.normal(0, 5000, n)
        )
        
        data = pd.DataFrame({
            'wage': wage,
            'education': education,
            'experience': experience,
            'gender': gender,
            'region': region
        })
        dataset_name = "Sample Wage Data (Enhanced)"
    
    elif dataset_type == "housing":
        n = 400
        square_feet = np.random.normal(2000, 500, n)
        bedrooms = np.random.choice([1, 2, 3, 4, 5], n, p=[0.1, 0.2, 0.4, 0.2, 0.1])
        bathrooms = np.random.normal(2, 0.5, n)
        age = np.random.exponential(15, n)
        location_premium = np.random.choice([0, 15000, 30000, 50000], n, p=[0.4, 0.3, 0.2, 0.1])
        garage = np.random.choice([0, 1], n, p=[0.3, 0.7])
        
        price = (
            50000 + 
            100 * square_feet + 
            5000 * bedrooms +
            3000 * bathrooms -
            500 * age + 
            location_premium +
            8000 * garage +
            np.random.normal(0, 15000, n)
        )
        
        data = pd.DataFrame({
            'price': price,
            'square_feet': square_feet,
            'bedrooms': bedrooms,
            'bathrooms': bathrooms,
            'age': age,
            'location_premium': location_premium,
            'garage': garage
        })
        dataset_name = "Sample Housing Data (Enhanced)"
    
    elif dataset_type == "stock":
        n = 300
        market_return = np.random.normal(0.05, 0.15, n)
        company_beta = np.random.normal(1.2, 0.3, n)
        company_return = 0.02 + company_beta * market_return + np.random.normal(0, 0.1, n)
        volatility = np.abs(np.random.normal(0.2, 0.05, n))
        market_cap = np.random.lognormal(15, 2, n)
        
        data = pd.DataFrame({
            'company_return': company_return,
            'market_return': market_return,
            'volatility': volatility,
            'trading_volume': np.random.lognormal(10, 1, n),
            'market_cap': market_cap,
            'beta': company_beta
        })
        dataset_name = "Sample Stock Data (Enhanced)"
    
    else:  # production data
        n = 250
        labor = np.random.exponential(50, n)
        capital = np.random.exponential(100, n)
        technology = np.random.normal(1, 0.1, n)
        energy = np.random.exponential(20, n)
        
        output = technology * (labor ** 0.6) * (capital ** 0.3) * (energy ** 0.1) + np.random.normal(0, 5, n)
        
        data = pd.DataFrame({
            'output': output,
            'labor': labor,
            'capital': capital,
            'technology': technology,
            'energy': energy
        })
        dataset_name = "Sample Production Data"
    
    # Update session state
    st.session_state.data = data
    st.session_state.uploaded_file_name = dataset_name
    st.session_state.data_hash = hash(str(data.values.tobytes()))
    
    # Reset analysis-specific state
    if replace_current:
        st.session_state.dependent_var = None
        st.session_state.independent_vars = []
        st.session_state.regression_results = None
    
    st.success(f"✅ {dataset_name} loaded successfully!")
    st.experimental_rerun()

# FIXED Enhanced data upload page - NO MORE HANGING!
def enhanced_data_upload_page(language):
    """Fixed data upload page without hanging issues"""
    
    t = translations[language]
    
    st.markdown(f'<h2 class="section-header">{t["upload_data"]}</h2>', unsafe_allow_html=True)
    
    # Current data status
    if st.session_state.data is not None:
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.success(f"✅ **Data Loaded**: {st.session_state.data.shape[0]} rows × {st.session_state.data.shape[1]} columns")
            if st.session_state.uploaded_file_name:
                st.info(f"📁 **File**: {st.session_state.uploaded_file_name}")
        
        with col2:
            if st.button("🔄 **Reset Current Data**", type="secondary", key="reset_data_btn"):
                reset_current_data()
                st.success("✅ Data reset successfully!")
                st.experimental_rerun()
        
        with col3:
            if st.session_state.regression_results:
                if st.button("💾 **Save Current Analysis**", type="primary", key="save_analysis_btn"):
                    if save_analysis_results():
                        st.success("✅ Analysis saved successfully!")
    
    # File upload section with better handling
    st.markdown("### 📁 Upload New Data")
    
    # File size warning
    st.info("📝 **File Requirements**: CSV or Excel files, maximum 50MB, up to 10,000 rows for optimal performance")
    
    uploaded_file = st.file_uploader(
        "Choose a CSV or Excel file",
        type=['csv', 'xlsx', 'xls'],
        help="Upload your dataset for analysis - CSV or Excel format",
        key="file_uploader_fixed"
    )
    
    # Upload options
    col1, col2 = st.columns(2)
    with col1:
        replace_current = st.checkbox(
            "🔄 Replace current data", 
            value=True, 
            help="Replace existing data with new upload"
        )
    
    with col2:
        show_progress = st.checkbox(
            "📊 Show upload progress", 
            value=True, 
            help="Display progress during file processing"
        )
    
    # Handle file upload with progress
    if uploaded_file is not None:
        if replace_current or st.session_state.data is None:
            
            # Show file info
            file_size = len(uploaded_file.getvalue()) / (1024 * 1024)
            st.write(f"📄 **File**: {uploaded_file.name} ({file_size:.2f} MB)")
            
            # Process with progress bar
            if st.button("🚀 **Process File**", type="primary", key="process_file_btn"):
                
                progress_bar = st.progress(0) if show_progress else None
                status_text = st.empty() if show_progress else None
                
                try:
                    if show_progress:
                        status_text.text("Reading file...")
                        progress_bar.progress(20)
                    
                    # Get file content
                    file_content = uploaded_file.getvalue()
                    
                    if show_progress:
                        status_text.text("Processing data...")
                        progress_bar.progress(40)
                    
                    # Load and process data
                    data, data_hash = cache_uploaded_data(file_content, uploaded_file.name)
                    
                    if data is not None:
                        if show_progress:
                            status_text.text("Optimizing memory usage...")
                            progress_bar.progress(60)
                        
                        # Optimize memory
                        data = optimize_memory_usage(data)
                        
                        if show_progress:
                            status_text.text("Updating session...")
                            progress_bar.progress(80)
                        
                        # Update session state
                        st.session_state.data = data
                        st.session_state.uploaded_file_name = uploaded_file.name
                        st.session_state.data_hash = data_hash
                        
                        # Reset analysis state if replacing
                        if replace_current:
                            st.session_state.dependent_var = None
                            st.session_state.independent_vars = []
                            st.session_state.regression_results = None
                        
                        if show_progress:
                            status_text.text("Complete!")
                            progress_bar.progress(100)
                        
                        st.success("✅ Data uploaded successfully!")
                        
                        # Clean up
                        gc.collect()
                        
                        # Auto-refresh after short delay
                        import time
                        time.sleep(1)
                        st.experimental_rerun()
                    
                    else:
                        if show_progress:
                            status_text.text("❌ Upload failed")
                            progress_bar.progress(0)
                        st.error("❌ Failed to process file. Please check file format and size.")
                
                except Exception as e:
                    if show_progress:
                        status_text.text("❌ Error occurred")
                        progress_bar.progress(0)
                    st.error(f"❌ Upload error: {str(e)}")
                    st.info("💡 **Try these solutions:**\n- Use a smaller file\n- Save as CSV format\n- Check for special characters in data")
    
    # Display data if available
    if st.session_state.data is not None:
        enhanced_data_preview(st.session_state.data, language)
    
    # Sample data section
    st.markdown("---")
    st.subheader("📊 Or Try Sample Data (No Upload Issues)")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("💰 Wage Dataset", key="wage_data_enhanced"):
            load_sample_data_enhanced("wage", replace_current)
    
    with col2:
        if st.button("🏠 Housing Dataset", key="housing_data_enhanced"):
            load_sample_data_enhanced("housing", replace_current)
    
    with col3:
        if st.button("📈 Stock Dataset", key="stock_data_enhanced"):
            load_sample_data_enhanced("stock", replace_current)
    
    with col4:
        if st.button("🏭 Production Dataset", key="production_data_enhanced"):
            load_sample_data_enhanced("production", replace_current)

def enhanced_data_preview(data, language):
    """Enhanced data preview with full display options"""
    
    st.markdown("### 📊 Data Overview")
    
    # Enhanced metrics
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("📊 Rows", data.shape[0])
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("📈 Columns", data.shape[1])
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        numeric_cols = len(data.select_dtypes(include=[np.number]).columns)
        st.metric("🔢 Numeric", numeric_cols)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col4:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        missing_vals = data.isnull().sum().sum()
        st.metric("❓ Missing", missing_vals)
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col5:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        memory_usage = data.memory_usage(deep=True).sum() / (1024**2)
        st.metric("💾 Memory (MB)", f"{memory_usage:.2f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Full data display with options
    show_all, rows_shown = display_full_data_table(data)
    
    # Data info section
    if st.expander("📋 Column Information", expanded=False):
        column_info = pd.DataFrame({
            'Column': data.columns,
            'Type': data.dtypes,
            'Non-Null': data.count(),
            'Null Count': data.isnull().sum(),
            'Unique Values': data.nunique()
        })
        st.dataframe(column_info, use_container_width=True)
    
    # Enhanced statistics
    if st.checkbox("📊 Show Descriptive Statistics", key="show_stats_enhanced"):
        st.subheader("📊 Descriptive Statistics")
        
        col1, col2 = st.columns(2)
        with col1:
            include_all = st.checkbox("Include all columns", value=False)
        with col2:
            percentiles = st.multiselect(
                "Additional percentiles",
                [0.01, 0.05, 0.10, 0.25, 0.75, 0.90, 0.95, 0.99],
                default=[0.25, 0.75]
            )
        
        if include_all:
            stats_df = data.describe(include='all', percentiles=percentiles)
        else:
            stats_df = data.describe(percentiles=percentiles)
        
        st.dataframe(stats_df, use_container_width=True)
        
        # Download statistics
        csv_stats = stats_df.to_csv()
        st.download_button(
            label="📥 Download Statistics CSV",
            data=csv_stats,
            file_name="descriptive_statistics.csv",
            mime="text/csv"
        )
    
    # Data visualization
    if st.checkbox("📈 Show Data Visualization", key="show_viz_enhanced"):
        st.subheader("📈 Data Visualization")
        
        numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
        
        if len(numeric_cols) >= 2:
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("🔥 Correlation Heatmap")
                fig, ax = plt.subplots(figsize=(10, 8))
                sns.heatmap(data[numeric_cols].corr(), annot=True, cmap='coolwarm', center=0)
                st.pyplot(fig)
            
            with col2:
                st.subheader("📊 Distribution Plots")
                selected_col = st.selectbox("Select column for distribution:", numeric_cols)
                
                fig, ax = plt.subplots(figsize=(10, 6))
                sns.histplot(data[selected_col], kde=True, ax=ax)
                plt.title(f'Distribution of {selected_col}')
                st.pyplot(fig)

def regression_analysis_page(language):
    """Enhanced regression analysis page with persistent settings"""
    
    t = translations[language]
    
    if st.session_state.data is None:
        st.warning("⚠️ Please upload data first!")
        st.info("👈 Go to 'Data Upload & Preview' to load your dataset")
        return
    
    data = st.session_state.data
    numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
    
    if len(numeric_cols) < 2:
        st.error("❌ Need at least 2 numeric columns for regression analysis!")
        return
    
    st.markdown(f'<h2 class="section-header">{t["regression_analysis"]}</h2>', unsafe_allow_html=True)
    
    # Show current data info
    st.info(f"📊 **Current Dataset**: {data.shape[0]} rows × {data.shape[1]} columns | File: {st.session_state.uploaded_file_name}")
    
    # Variable selection with state persistence
    st.markdown(f'<h3 class="section-header">{t["variable_selection"]}</h3>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Get current dependent variable or use default
        current_dep = st.session_state.dependent_var if st.session_state.dependent_var in numeric_cols else numeric_cols[0]
        
        dependent_var = st.selectbox(
            f"🎯 {t['dependent_var']}",
            numeric_cols,
            index=numeric_cols.index(current_dep) if current_dep in numeric_cols else 0,
            help="Choose the variable you want to predict",
            key="dep_var_select"
        )
        
        # Update session state
        if st.session_state.dependent_var != dependent_var:
            st.session_state.dependent_var = dependent_var
    
    with col2:
        # Get available independent variables
        available_indep = [col for col in numeric_cols if col != dependent_var]
        
        # Restore previous selections if valid
        current_indep = [var for var in st.session_state.independent_vars if var in available_indep]
        
        independent_vars = st.multiselect(
            f"📊 {t['independent_vars']}",
            available_indep,
            default=current_indep,
            help="Choose variables to predict the dependent variable",
            key="indep_var_select"
        )
        
        # Update session state
        if st.session_state.independent_vars != independent_vars:
            st.session_state.independent_vars = independent_vars
    
    if not independent_vars:
        st.warning("⚠️ Please select at least one independent variable!")
        return
    
    # Show analysis summary
    regression_type = t["simple_regression"] if len(independent_vars) == 1 else t["multiple_regression"]
    st.info(f"🔍 **Analysis Type**: {regression_type}")
    
    # Analysis preview
    with st.expander("📋 Analysis Preview", expanded=True):
        st.markdown(f"""
        **📊 Model Specification:**
        - **Dependent Variable**: {dependent_var}
        - **Independent Variables**: {', '.join(independent_vars)}
        - **Model Type**: {regression_type}
        - **Sample Size**: {len(data)} observations
        """)
    
    if st.button(f"🚀 {t['run_regression']}", type="primary", key="run_regression_btn"):
        with st.spinner("Running regression analysis..."):
            run_regression_analysis(data, dependent_var, independent_vars)

def run_regression_analysis(data, dependent_var, independent_vars):
    """Run regression analysis with results persistence"""
    
    try:
        # Prepare data
        y = data[dependent_var].values
        X = data[independent_vars].values
        X_with_const = sm.add_constant(X)
        
        # Fit OLS model
        model = sm.OLS(y, X_with_const).fit()
        
        # Store results in session state
        st.session_state.regression_results = {
            'model': model,
            'dependent_var': dependent_var,
            'independent_vars': independent_vars,
            'X': X,
            'y': y,
            'X_with_const': X_with_const,
            'timestamp': datetime.now().isoformat()
        }
        
        # Add to analysis history
        analysis_record = {
            'dependent_var': dependent_var,
            'independent_vars': independent_vars,
            'r_squared': model.rsquared,
            'timestamp': datetime.now().isoformat()
        }
        st.session_state.analysis_history.append(analysis_record)
        
        st.success("✅ Regression analysis completed successfully!")
        
        # Display results
        display_regression_results(model, dependent_var, independent_vars, "English")
        
    except Exception as e:
        st.error(f"❌ Error in regression analysis: {str(e)}")
        st.info("Please check your data and variable selections.")

def display_regression_results(model, dependent_var, independent_vars, language):
    """Display comprehensive regression results with mathematical explanations"""
    
    # Mathematical Foundation - RED TEXT
    st.markdown("""
    <div class="formula-box">
    <h4 class="red-header">📚 Mathematical Foundation</h4>
    """, unsafe_allow_html=True)
    
    if len(independent_vars) == 1:
        st.latex(r"Y_i = \beta_0 + \beta_1 X_i + \varepsilon_i")
        st.markdown("""
        **Simple Linear Regression Model:**
        - Y₍ᵢ₎ = Dependent variable (what we're predicting)
        - β₀ = Intercept (value of Y when X = 0)
        - β₁ = Slope coefficient (change in Y for 1-unit change in X)
        - εᵢ = Error term (unexplained variation)
        """)
    else:
        st.latex(r"Y_i = \beta_0 + \beta_1 X_{1i} + \beta_2 X_{2i} + ... + \beta_k X_{ki} + \varepsilon_i")
        st.markdown("""
        **Multiple Linear Regression Model:**
        - Each β coefficient represents the **partial effect** of that variable
        - "Holding other variables constant" interpretation
        - More complex but more realistic relationships
        """)
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # OLS Estimation Explanation - RED TEXT
    st.markdown("""
    <div class="formula-box">
    <h4 class="red-header">🎯 Ordinary Least Squares (OLS) Method</h4>
    """, unsafe_allow_html=True)
    
    st.latex(r"\min_{\beta} \sum_{i=1}^{n} (Y_i - \hat{Y}_i)^2")
    st.markdown("""
    **The OLS Principle:**
    - Finds the line that **minimizes the sum of squared residuals**
    - Residual = Actual value - Predicted value
    - Why squared? Prevents positive and negative errors from canceling out
    - Results in the "best fitting" line through the data points
    """)
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Regression Results Table - ENHANCED WITH FULL DISPLAY
    st.subheader("📊 Regression Results Summary")
    
    # Handle both pandas Series and numpy array cases properly
    try:
        if hasattr(model.params, 'index') and hasattr(model.params, 'values'):
            # pandas Series case
            param_names = list(model.params.index)
            param_values = list(model.params.values)
            std_errors = list(model.bse.values)
            t_statistics = list(model.tvalues.values)
            p_values = list(model.pvalues.values)
            ci_lower = list(model.conf_int()[0].values)
            ci_upper = list(model.conf_int()[1].values)
        else:
            raise AttributeError("Not a pandas Series")
            
    except (AttributeError, IndexError):
        try:
            # Create variable names
            param_names = ['Intercept'] + independent_vars
            
            # Convert to numpy arrays and flatten
            param_values = np.atleast_1d(np.array(model.params)).flatten()
            std_errors = np.atleast_1d(np.array(model.bse)).flatten()
            t_statistics = np.atleast_1d(np.array(model.tvalues)).flatten()
            p_values = np.atleast_1d(np.array(model.pvalues)).flatten()
            
            # Handle confidence intervals
            try:
                conf_int = model.conf_int()
                if hasattr(conf_int, 'values'):
                    ci_lower = list(conf_int[0].values)
                    ci_upper = list(conf_int[1].values)
                else:
                    ci_lower = np.atleast_1d(np.array(conf_int[0])).flatten()
                    ci_upper = np.atleast_1d(np.array(conf_int[1])).flatten()
            except:
                t_crit = 1.96  # Approximate for large samples
                margin = t_crit * std_errors
                ci_lower = param_values - margin
                ci_upper = param_values + margin
            
            # Convert to lists
            param_values = list(param_values)
            std_errors = list(std_errors)
            t_statistics = list(t_statistics)
            p_values = list(p_values)
            ci_lower = list(ci_lower)
            ci_upper = list(ci_upper)
            
        except Exception as e:
            st.error(f"Error processing regression results: {str(e)}")
            return
    
    # Ensure all lists have the same length
    max_len = max(len(param_names), len(param_values), len(std_errors), 
                  len(t_statistics), len(p_values), len(ci_lower), len(ci_upper))
    
    def pad_list(lst, target_len, default_val=np.nan):
        while len(lst) < target_len:
            lst.append(default_val)
        return lst[:target_len]
    
    param_names = pad_list(param_names, max_len, "Unknown")
    param_values = pad_list(param_values, max_len, np.nan)
    std_errors = pad_list(std_errors, max_len, np.nan)
    t_statistics = pad_list(t_statistics, max_len, np.nan)
    p_values = pad_list(p_values, max_len, np.nan)
    ci_lower = pad_list(ci_lower, max_len, np.nan)
    ci_upper = pad_list(ci_upper, max_len, np.nan)
    
    # Create results dataframe
    results_df = pd.DataFrame({
        'Variable': param_names,
        'Coefficient': param_values,
        'Std Error': std_errors,
        't-statistic': t_statistics,
        'p-value': p_values,
        '95% CI Lower': ci_lower,
        '95% CI Upper': ci_upper
    })
    
    # Format the results table
    results_df['Coefficient'] = pd.to_numeric(results_df['Coefficient'], errors='coerce').round(4)
    results_df['Std Error'] = pd.to_numeric(results_df['Std Error'], errors='coerce').round(4)
    results_df['t-statistic'] = pd.to_numeric(results_df['t-statistic'], errors='coerce').round(3)
    results_df['p-value'] = pd.to_numeric(results_df['p-value'], errors='coerce').round(4)
    results_df['95% CI Lower'] = pd.to_numeric(results_df['95% CI Lower'], errors='coerce').round(4)
    results_df['95% CI Upper'] = pd.to_numeric(results_df['95% CI Upper'], errors='coerce').round(4)
    
    # Always show full results table
    st.dataframe(results_df, use_container_width=True, height=400)
    
    # Download coefficients button
    csv_coeffs = results_df.to_csv(index=False)
    st.download_button(
        label="📥 Download Coefficients Table",
        data=csv_coeffs,
        file_name=f"regression_coefficients_{dependent_var}.csv",
        mime="text/csv",
        key="download_coeffs"
    )
    
    # Statistical Significance Explanation
    st.markdown("""
    <div class="interpretation-box">
    <h4 class="red-header">📖 How to Interpret These Results:</h4>
    """, unsafe_allow_html=True)
    
    for i, var in enumerate(param_names):
        if i >= len(results_df):
            break
            
        coef = results_df.iloc[i]['Coefficient']
        pval = results_df.iloc[i]['p-value']
        
        # Handle NaN values
        if pd.isna(coef) or pd.isna(pval):
            continue
            
        if var == 'Intercept' or 'intercept' in var.lower():
            interpretation = f"**Intercept (β₀ = {coef:.4f})**: When all X variables = 0, {dependent_var} = {coef:.4f}"
        else:
            interpretation = f"**{var} (β = {coef:.4f})**: A 1-unit increase in {var} is associated with a {coef:.4f} change in {dependent_var}"
            if len(independent_vars) > 1:
                interpretation += " (holding other variables constant)"
        
        significance = "statistically significant" if pval < 0.05 else "not statistically significant"
        
        st.markdown(f"• {interpretation}")
        st.markdown(f"  - This effect is **{significance}** (p-value = {pval:.4f})")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Model Fit Statistics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("📊 R-squared", f"{model.rsquared:.4f}")
        st.markdown("% of variation explained")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("📈 Adjusted R²", f"{model.rsquared_adj:.4f}")
        st.markdown("Penalizes extra variables")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("🎯 F-statistic", f"{model.fvalue:.2f}")
        st.markdown(f"p-value: {model.f_pvalue:.4f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col4:
        st.markdown('<div class="metric-box">', unsafe_allow_html=True)
        st.metric("📏 Observations", f"{int(model.nobs)}")
        st.metric("📊 DoF", f"{int(model.df_resid)}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # F-test Explanation
    st.markdown("""
    <div class="formula-box">
    <h4 class="red-header">🧪 F-Test for Overall Model Significance</h4>
    """, unsafe_allow_html=True)
    
    st.latex(r"H_0: \beta_1 = \beta_2 = ... = \beta_k = 0")
    st.latex(r"H_A: \text{At least one } \beta_j \neq 0")
    
    st.markdown(f"""
    **F-statistic = {model.fvalue:.2f}** with p-value = **{model.f_pvalue:.4f}**
    
    **Interpretation:**
    - Tests whether ALL independent variables together have no effect on Y
    - If p-value < 0.05: Reject H₀ - the model is statistically significant
    - **Conclusion**: {"The model is statistically significant" if model.f_pvalue < 0.05 else "The model is not statistically significant"}
    """)
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Visualization
    st.subheader("📈 Regression Visualization")
    
    if len(independent_vars) == 1:
        # Simple regression plot
        X_data = st.session_state.regression_results['X'].flatten()
        y_data = st.session_state.regression_results['y']
        
        fig = px.scatter(
            x=X_data,
            y=y_data,
            title=f"Simple Linear Regression: {dependent_var} vs {independent_vars[0]}"
        )
        
        # Add regression line
        x_range = np.linspace(X_data.min(), X_data.max(), 100)
        y_pred_line = param_values[0] + param_values[1] * x_range
        
        fig.add_trace(go.Scatter(
            x=x_range,
            y=y_pred_line,
            mode='lines',
            name='Regression Line',
            line=dict(color='red', width=3)
        ))
        
        fig.update_layout(
            xaxis_title=independent_vars[0],
            yaxis_title=dependent_var,
            showlegend=True
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    else:
        # Multiple regression - show actual vs predicted
        y_pred = model.predict()
        y_actual = st.session_state.regression_results['y']
        
        fig = px.scatter(
            x=y_pred,
            y=y_actual,
            title="Multiple Regression: Actual vs Predicted Values"
        )
        
        # Add perfect prediction line
        min_val = min(min(y_pred), min(y_actual))
        max_val = max(max(y_pred), max(y_actual))
        
        fig.add_trace(go.Scatter(
            x=[min_val, max_val],
            y=[min_val, max_val],
            mode='lines',
            name='Perfect Prediction',
            line=dict(color='red', dash='dash', width=2)
        ))
        
        fig.update_layout(
            xaxis_title="Predicted Values",
            yaxis_title="Actual Values"
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Residual Analysis Preview
    st.subheader("🔍 Quick Residual Check")
    residuals = model.resid
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 5))
    
    # Residuals vs Fitted
    ax1.scatter(model.fittedvalues, residuals, alpha=0.6)
    ax1.axhline(y=0, color='red', linestyle='--')
    ax1.set_xlabel('Fitted Values')
    ax1.set_ylabel('Residuals')
    ax1.set_title('Residuals vs Fitted Values')
    
    # QQ plot for normality
    stats.probplot(residuals, dist="norm", plot=ax2)
    ax2.set_title('Q-Q Plot (Normality Check)')
    
    st.pyplot(fig)
    
    # Save Analysis Results Section
    st.markdown("---")
    st.markdown("### 💾 Save Analysis Results")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("💾 Save to Session", key="save_to_session"):
            if save_analysis_results():
                st.success("✅ Analysis saved to session!")
            else:
                st.error("❌ Failed to save analysis")
    
    with col2:
        # Export model statistics
        model_stats = get_model_statistics()
        stats_df = pd.DataFrame([model_stats])
        csv_stats = stats_df.to_csv(index=False)
        
        st.download_button(
            label="📈 Download Model Stats",
            data=csv_stats,
            file_name=f"model_statistics_{dependent_var}.csv",
            mime="text/csv",
            key="download_stats"
        )
    
    with col3:
        # Export full summary
        summary_text = str(model.summary())
        st.download_button(
            label="📋 Download Full Summary",
            data=summary_text,
            file_name=f"regression_summary_{dependent_var}.txt",
            mime="text/plain",
            key="download_summary"
        )
    
    st.info("💡 **Next Steps**: Check the 'Model Diagnostics' tab for comprehensive assumption testing!")

def diagnostics_page(language):
    """Comprehensive diagnostic tests for regression assumptions"""
    
    if st.session_state.regression_results is None:
        st.warning("⚠️ Please run regression analysis first!")
        return
    
    st.markdown('<h2 class="section-header">🔍 Model Diagnostics & Assumption Testing</h2>', 
                unsafe_allow_html=True)
    
    model = st.session_state.regression_results['model']
    X = st.session_state.regression_results['X']
    y = st.session_state.regression_results['y']
    X_with_const = st.session_state.regression_results['X_with_const']
    independent_vars = st.session_state.regression_results['independent_vars']
    
    # Overview of assumptions - RED TEXT
    st.markdown("""
    <div class="formula-box">
    <h4 class="red-header">📋 The Five Key Assumptions We're Testing:</h4>
    <div class="red-list">
    1. <strong>Linearity</strong>: The relationship is actually linear<br>
    2. <strong>Homoskedasticity</strong>: Constant error variance<br>
    3. <strong>No Autocorrelation</strong>: Errors are independent<br>
    4. <strong>Normality</strong>: Errors are normally distributed<br>
    5. <strong>No Multicollinearity</strong>: Independent variables aren't perfectly correlated
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Test 1: Heteroskedasticity Tests
    st.subheader("🔥 Test 1: Heteroskedasticity (Constant Variance)")
    
    # Breusch-Pagan Test
    bp_stat, bp_pvalue, _, _ = het_breuschpagan(model.resid, X_with_const)
    
    st.markdown("""
    <div class="formula-box">
    <h5 class="red-header">📚 Breusch-Pagan Test</h5>
    <div class="red-hypothesis">
    <strong>Null Hypothesis (H₀): Homoskedasticity (constant variance)</strong><br>
    <strong>Alternative Hypothesis (H₁): Heteroskedasticity (non-constant variance)</strong>
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("🎯 Breusch-Pagan Statistic", f"{bp_stat:.4f}")
        st.metric("📊 p-value", f"{bp_pvalue:.4f}")
    
    with col2:
        if bp_pvalue < 0.05:
            st.error("❌ **Reject H₀**: Heteroskedasticity detected!")
            st.markdown("**Solution**: Use robust standard errors or transform variables")
        else:
            st.success("✅ **Fail to reject H₀**: No evidence of heteroskedasticity")
    
    # Visual test for heteroskedasticity
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.scatter(model.fittedvalues, model.resid, alpha=0.6)
    ax.axhline(y=0, color='red', linestyle='--')
    ax.set_xlabel('Fitted Values')
    ax.set_ylabel('Residuals')
    ax.set_title('Residuals vs Fitted Values (Heteroskedasticity Check)')
    st.pyplot(fig)
    
    st.markdown("""
    <div class="interpretation-box">
    <div class="red-interpretation">
    <strong>How to interpret this plot:</strong><br>
    • <strong>Good: Random scatter around zero line</strong><br>
    • <strong>Bad: Funnel shape (variance increases) or systematic patterns</strong>
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Test 2: Autocorrelation (Durbin-Watson Test)
    st.subheader("🔄 Test 2: Autocorrelation (Independence of Errors)")
    
    dw_stat = durbin_watson(model.resid)
    
    st.markdown("""
    <div class="formula-box">
    <h5 class="red-header">📚 Durbin-Watson Test</h5>
    """, unsafe_allow_html=True)
    
    st.latex(r"DW = \frac{\sum_{t=2}^{n}(e_t - e_{t-1})^2}{\sum_{t=1}^{n}e_t^2}")
    
    st.markdown("""
    <strong>Interpretation Rules:</strong><br>
    • DW ≈ 2.0: No autocorrelation<br>
    • DW < 1.5: Positive autocorrelation<br>
    • DW > 2.5: Negative autocorrelation
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("🎯 Durbin-Watson Statistic", f"{dw_stat:.4f}")
    
    with col2:
        if 1.5 <= dw_stat <= 2.5:
            st.success("✅ **No significant autocorrelation**")
        elif dw_stat < 1.5:
            st.error("❌ **Positive autocorrelation detected**")
            st.markdown("**Solution**: Add lagged variables or use robust standard errors")
        else:
            st.error("❌ **Negative autocorrelation detected**")
            st.markdown("**Solution**: Check for over-differencing")
    
    # Test 3: Normality (Jarque-Bera Test)
    st.subheader("📊 Test 3: Normality of Residuals")
    
    try:
        jb_result = jarque_bera(model.resid)
        jb_stat = jb_result[0]
        jb_pvalue = jb_result[1]
    except:
        from scipy.stats import jarque_bera as scipy_jb
        jb_stat, jb_pvalue = scipy_jb(model.resid)
    
    st.markdown("""
    <div class="formula-box">
    <h5 class="red-header">📚 Jarque-Bera Test</h5>
    """, unsafe_allow_html=True)
    
    st.latex(r"JB = \frac{n}{6}\left[S^2 + \frac{1}{4}(K-3)^2\right]")
    
    st.markdown("""
    Where:<br>
    • S = Skewness of residuals<br>
    • K = Kurtosis of residuals<br>
    • Normal distribution: S=0, K=3<br><br>
    <strong>H₀:</strong> Residuals are normally distributed<br>
    <strong>H₁:</strong> Residuals are not normally distributed
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("🎯 Jarque-Bera Statistic", f"{jb_stat:.4f}")
        st.metric("📊 p-value", f"{jb_pvalue:.4f}")
        
        # Calculate skewness and kurtosis
        from scipy.stats import skew, kurtosis
        resid_skew = skew(model.resid)
        resid_kurt = kurtosis(model.resid) + 3  # Adding 3 for normal kurtosis
        
        st.metric("📈 Skewness", f"{resid_skew:.4f}")
        st.metric("📊 Kurtosis", f"{resid_kurt:.4f}")
    
    with col2:
        if jb_pvalue < 0.05:
            st.error("❌ **Reject H₀**: Residuals are not normally distributed!")
            st.markdown("**Solutions**: Transform variables, use robust regression, or larger sample")
        else:
            st.success("✅ **Fail to reject H₀**: Residuals appear normally distributed")
    
    # Visual normality tests
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    
    # Histogram of residuals
    ax1.hist(model.resid, bins=20, density=True, alpha=0.7, color='skyblue')
    ax1.set_xlabel('Residuals')
    ax1.set_ylabel('Density')
    ax1.set_title('Histogram of Residuals')
    
    # Add normal curve
    x_norm = np.linspace(model.resid.min(), model.resid.max(), 100)
    y_norm = stats.norm.pdf(x_norm, loc=np.mean(model.resid), scale=np.std(model.resid))
    ax1.plot(x_norm, y_norm, 'r-', linewidth=2, label='Normal Distribution')
    ax1.legend()
    
    # Q-Q plot
    stats.probplot(model.resid, dist="norm", plot=ax2)
    ax2.set_title('Q-Q Plot: Normal Distribution Check')
    
    st.pyplot(fig)
    
    # Test 4: Multicollinearity (VIF)
    if len(independent_vars) > 1:
        st.subheader("🔗 Test 4: Multicollinearity (VIF Analysis)")
        
        st.markdown("""
        <div class="formula-box">
        <h5>📚 Variance Inflation Factor (VIF)</h5>
        """, unsafe_allow_html=True)
        
        st.latex(r"VIF_j = \frac{1}{1-R_j^2}")
        
        st.markdown("""
        Where R²ⱼ comes from regressing Xⱼ on all other X variables<br><br>
        <strong>Interpretation:</strong><br>
        • VIF = 1: No correlation with other variables<br>
        • VIF < 5: Acceptable multicollinearity<br>
        • VIF > 10: Problematic multicollinearity
        </div>
        """, unsafe_allow_html=True)
        
        # Calculate VIF for each variable
        vif_data = pd.DataFrame()
        vif_data["Variable"] = independent_vars
        
        vif_values = []
        for i in range(len(independent_vars)):
            vif = variance_inflation_factor(X, i)
            vif_values.append(vif)
        
        vif_data["VIF"] = vif_values
        vif_data["VIF"] = vif_data["VIF"].round(3)
        
        # Add interpretation
        vif_data["Interpretation"] = vif_data["VIF"].apply(
            lambda x: "✅ Low" if x < 5 else ("⚠️ Moderate" if x < 10 else "❌ High")
        )
        
        # Always show full VIF table
        st.dataframe(vif_data, use_container_width=True, height=300)
        
        # Download VIF results
        vif_csv = vif_data.to_csv(index=False)
        st.download_button(
            label="📥 Download VIF Results",
            data=vif_csv,
            file_name="vif_analysis.csv",
            mime="text/csv",
            key="download_vif"
        )
        
        # Check for problematic multicollinearity
        max_vif = vif_data["VIF"].max()
        if max_vif > 10:
            st.error(f"❌ **High multicollinearity detected!** Maximum VIF = {max_vif:.3f}")
            st.markdown("""
            **Solutions:**
            - Remove highly correlated variables
            - Combine correlated variables into indices
            - Use ridge regression
            - Collect more data
            """)
        elif max_vif > 5:
            st.warning(f"⚠️ **Moderate multicollinearity present.** Maximum VIF = {max_vif:.3f}")
        else:
            st.success(f"✅ **Low multicollinearity.** Maximum VIF = {max_vif:.3f}")

# Enhanced sidebar with saved analyses  
def enhanced_sidebar_with_history():
    """Enhanced sidebar with complete analysis history"""
    
    st.sidebar.markdown("---")
    
    # Current session info
    with st.sidebar.expander("📊 Current Session", expanded=True):
        if st.session_state.data is not None:
            st.write(f"**📁 Dataset**: {st.session_state.uploaded_file_name}")
            st.write(f"**📏 Dimensions**: {st.session_state.data.shape[0]} × {st.session_state.data.shape[1]}")
            
            if st.session_state.dependent_var:
                st.write(f"**🎯 Dependent**: {st.session_state.dependent_var}")
            
            if st.session_state.independent_vars:
                indep_display = ', '.join(st.session_state.independent_vars[:3])
                if len(st.session_state.independent_vars) > 3:
                    indep_display += f" (+{len(st.session_state.independent_vars)-3} more)"
                st.write(f"**📊 Independent**: {indep_display}")
            
            # Analysis status
            if st.session_state.regression_results:
                r_sq = st.session_state.regression_results['model'].rsquared
                st.write(f"**📈 R²**: {r_sq:.3f}")
                st.success("✅ Analysis Complete")
            else:
                st.info("⏳ Ready for Analysis")
        else:
            st.error("❌ No data loaded")
            st.write("👈 Go to Data Upload & Preview")
    
    # Quick data actions
    if st.session_state.data is not None:
        with st.sidebar.expander("⚡ Quick Actions"):
            col1, col2 = st.sidebar.columns(2)
            
            with col1:
                if st.button("🔄 Reset Data", key="sidebar_reset", help="Clear current data"):
                    reset_current_data()
                    st.success("✅ Data reset!")
                    st.experimental_rerun()
            
            with col2:
                if st.session_state.regression_results:
                    if st.button("💾 Save Analysis", key="sidebar_save", help="Save current analysis"):
                        if save_analysis_results():
                            st.success("✅ Saved!")
                        else:
                            st.error("❌ Save failed")
    
    # Analysis history
    if st.session_state.get('saved_analyses'):
        with st.sidebar.expander(f"📈 Analysis History ({len(st.session_state.saved_analyses)})", expanded=False):
            for i, analysis in enumerate(reversed(st.session_state.saved_analyses)):
                with st.container():
                    st.markdown(f"**📊 Analysis {i+1}** `{analysis['session_id']}`")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("R²", f"{analysis['model_stats']['r_squared']:.3f}", delta=None)
                    with col2:
                        st.metric("N", analysis['model_stats']['observations'])
                    
                    st.write(f"**Y**: {analysis['data_info']['dependent_var']}")
                    indep_vars = analysis['data_info']['independent_vars']
                    if len(indep_vars) <= 2:
                        st.write(f"**X**: {', '.join(indep_vars)}")
                    else:
                        st.write(f"**X**: {indep_vars[0]}, {indep_vars[1]} (+{len(indep_vars)-2} more)")
                    
                    timestamp = datetime.fromisoformat(analysis['timestamp'])
                    st.write(f"🕒 {timestamp.strftime('%Y-%m-%d %H:%M')}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("📥 Export", key=f"export_hist_{i}", help="Export this analysis"):
                            analysis_json = json.dumps(analysis, indent=2, default=str)
                            st.download_button(
                                label="Download JSON",
                                data=analysis_json,
                                file_name=f"analysis_{analysis['session_id']}.json",
                                mime="application/json",
                                key=f"download_hist_{i}"
                            )
                    
                    with col2:
                        if st.button("🗑️ Delete", key=f"delete_hist_{i}", help="Remove this analysis"):
                            st.session_state.saved_analyses.pop(-(i+1))
                            st.experimental_rerun()
                    
                    st.markdown("---")
    
    # Session management
    with st.sidebar.expander("🔧 Session Management"):
        if st.session_state.data is not None:
            memory_mb = st.session_state.data.memory_usage(deep=True).sum() / (1024**2)
            st.info(f"💾 Memory Usage: {memory_mb:.1f} MB")
        
        analyses_count = len(st.session_state.get('saved_analyses', []))
        st.info(f"📊 Saved Analyses: {analyses_count}")
        
        if st.button("💾 Save Current Session", key="save_full_session", help="Save current analysis"):
            if st.session_state.regression_results:
                if save_analysis_results():
                    st.success("✅ Current analysis saved to history!")
                else:
                    st.error("❌ Failed to save analysis")
            else:
                st.warning("⚠️ No analysis to save")
        
        if st.button("🧹 Clear All History", key="clear_all_history", help="Remove all saved analyses"):
            st.session_state.saved_analyses = []
            st.success("✅ Analysis history cleared!")
            st.experimental_rerun()
        
        if st.button("🔄 Complete Reset", key="complete_reset", help="Reset everything"):
            keys_to_clear = ['data', 'regression_results', 'dependent_var', 
                           'independent_vars', 'saved_analyses', 'analysis_history']
            for key in keys_to_clear:
                if key in st.session_state:
                    del st.session_state[key]
            
            initialize_session_state()
            st.success("✅ Complete reset performed!")
            st.experimental_rerun()
    
    # Help section
    with st.sidebar.expander("❓ Help & Shortcuts"):
        st.markdown("""
        **Quick Start Guide:**
        1. 📁 Upload data or use samples
        2. 🎯 Select variables  
        3. 🚀 Run regression analysis
        4. 🔍 Check diagnostics
        5. 📤 Export results
        
        **Upload Tips:**
        - 💾 Use CSV for best compatibility
        - 📊 Max 50MB file size
        - 🔍 Check data before uploading
        """)
    
    # App information
    with st.sidebar.expander("ℹ️ About This App"):
        st.markdown("""
        ## 📊 **Econometrics Learning Lab**
        
        **🎓 Professional Statistical Analysis Platform**
        
        **Created by HAMDI Boulanouar**
        
        ### Features:
        - ✅ Complete OLS Analysis
        - ✅ Comprehensive Diagnostics  
        - ✅ Fixed Upload System
        - ✅ Professional Export Options
        - ✅ Educational Materials
        - ✅ Bilingual Support (EN/AR)
        
        ---
        *Making Advanced Econometrics Accessible to Everyone*
        """)

# Simple stub functions for additional pages
def advanced_diagnostics_page():
    """Complete advanced diagnostic features"""
    
    if st.session_state.regression_results is None:
        st.warning("⚠️ Please run regression analysis first!")
        return
    
    st.markdown('<h2 class="section-header">🔬 Advanced Diagnostics</h2>', unsafe_allow_html=True)
    
    model = st.session_state.regression_results['model']
    X = st.session_state.regression_results['X']
    y = st.session_state.regression_results['y']
    independent_vars = st.session_state.regression_results['independent_vars']
    
    # Leverage and Influence Analysis
    st.subheader("📊 Leverage and Influence Analysis")
    
    # Calculate diagnostic measures
    influence = model.get_influence()
    leverage = influence.hat_matrix_diag
    cooks_distance = influence.cooks_distance[0]
    standardized_residuals = influence.resid_studentized_internal
    dffits = influence.dffits[0]
    dfbetas = influence.dfbetas
    
    # Create influence plot
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=leverage,
        y=standardized_residuals,
        mode='markers',
        marker=dict(
            size=cooks_distance * 100 + 5,
            color=cooks_distance,
            colorscale='Reds',
            showscale=True,
            colorbar=dict(title="Cook's Distance")
        ),
        text=[f"Obs {i+1}<br>Cook's D: {d:.3f}<br>Leverage: {l:.3f}<br>Std Resid: {r:.3f}" 
              for i, (d, l, r) in enumerate(zip(cooks_distance, leverage, standardized_residuals))],
        hovertemplate='%{text}<extra></extra>',
        name="Observations"
    ))
    
    fig.update_layout(
        title="Influence Plot: Leverage vs Standardized Residuals",
        xaxis_title="Leverage",
        yaxis_title="Standardized Residuals",
        showlegend=False,
        height=500
    )
    
    # Add reference lines
    fig.add_hline(y=2, line_dash="dash", line_color="red", annotation_text="High Residual (+2)")
    fig.add_hline(y=-2, line_dash="dash", line_color="red", annotation_text="High Residual (-2)")
    
    # Average leverage line
    avg_leverage = len(model.params) / len(y)
    fig.add_vline(x=avg_leverage, line_dash="dash", line_color="blue", 
                  annotation_text=f"Average Leverage ({avg_leverage:.3f})")
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Identify influential observations
    high_leverage_threshold = 2 * len(model.params) / len(y)
    high_cooks_threshold = 4 / len(y)
    high_dffits_threshold = 2 * np.sqrt(len(model.params) / len(y))
    
    high_leverage = leverage > high_leverage_threshold
    high_cooks = cooks_distance > high_cooks_threshold
    high_dffits = np.abs(dffits) > high_dffits_threshold
    high_resid = np.abs(standardized_residuals) > 2
    
    # Create comprehensive influence dataframe
    influence_df = pd.DataFrame({
        'Observation': range(1, len(y) + 1),
        'Leverage': leverage.round(4),
        'Cooks_Distance': cooks_distance.round(4),
        'DFFITS': dffits.round(4),
        'Std_Residual': standardized_residuals.round(4),
        'High_Leverage': high_leverage,
        'High_Cooks': high_cooks,
        'High_DFFITS': high_dffits,
        'High_Residual': high_resid,
        'Problematic': high_leverage | high_cooks | high_dffits | high_resid
    })
    
    # Display influence summary
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        st.metric("🎯 High Leverage", int(sum(high_leverage)))
        st.markdown(f"Threshold: {high_leverage_threshold:.3f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        st.metric("🔥 High Cook's D", int(sum(high_cooks)))
        st.markdown(f"Threshold: {high_cooks_threshold:.3f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        st.metric("📊 High DFFITS", int(sum(high_dffits)))
        st.markdown(f"Threshold: {high_dffits_threshold:.3f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col4:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        problematic_count = int(sum(influence_df['Problematic']))
        st.metric("⚠️ Total Problematic", problematic_count)
        st.markdown(f"Percent: {problematic_count/len(y)*100:.1f}%")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Show influence data table
    show_all_influence = st.checkbox("Show all observations", value=False, key="show_all_influence")
    
    if show_all_influence:
        st.subheader("📋 Complete Influence Analysis")
        st.dataframe(influence_df, use_container_width=True, height=400)
    else:
        problematic = influence_df[influence_df['Problematic']]
        if len(problematic) > 0:
            st.warning(f"⚠️ **{len(problematic)} Potentially Influential Observations Detected:**")
            st.dataframe(problematic, use_container_width=True)
        else:
            st.success("✅ No highly influential observations detected!")
    
    # DFBETAS Analysis
    if len(independent_vars) > 0:
        st.subheader("📈 DFBETAS Analysis (Parameter-Specific Influence)")
        
        st.markdown("""
        <div class="formula-box">
        <h5 class="red-header">📚 DFBETAS Interpretation</h5>
        <div class="red-text">
        DFBETAS measures how much each coefficient changes when observation i is removed<br>
        <strong>Rule of thumb:</strong> |DFBETAS| > 2/√n suggests influential observation
        </div>
        </div>
        """, unsafe_allow_html=True)
        
        # DFBETAS threshold
        dfbetas_threshold = 2 / np.sqrt(len(y))
        
        # Create DFBETAS plot
        var_names = ['Intercept'] + independent_vars
                # DFBETAS Analysis
        if len(independent_vars) > 0:
            st.subheader("📈 DFBETAS Analysis (Parameter-Specific Influence)")
            
            st.markdown("""
            <div class="formula-box">
            <h5 class="red-header">📚 DFBETAS Interpretation</h5>
            <div class="red-text">
            DFBETAS measures how much each coefficient changes when observation i is removed<br>
            <strong>Rule of thumb:</strong> |DFBETAS| > 2/√n suggests influential observation
            </div>
            </div>
            """, unsafe_allow_html=True)
            
            # DFBETAS threshold
            dfbetas_threshold = 2 / np.sqrt(len(y))
            
            # Create DFBETAS plot - CORRECTED VERSION
            var_names = ['Intercept'] + independent_vars
            fig_dfbetas = make_subplots(
                rows=len(var_names), cols=1,
                subplot_titles=[f"DFBETAS for {var}" for var in var_names],
                shared_xaxes=True  # ✅ CORRECT: shared_xaxes (plural)
            )
            
            for i, var_name in enumerate(var_names):
                if i < dfbetas.shape[1]:
                    fig_dfbetas.add_trace(
                        go.Scatter(
                            x=list(range(1, len(y) + 1)),
                            y=dfbetas[:, i],
                            mode='markers',
                            name=var_name,
                            marker=dict(
                                color=['red' if abs(val) > dfbetas_threshold else 'blue' 
                                       for val in dfbetas[:, i]]
                            )
                        ),
                        row=i+1, col=1
                    )
                    
                    # Add threshold lines
                    fig_dfbetas.add_hline(
                        y=dfbetas_threshold, line_dash="dash", line_color="red",
                        row=i+1, col=1
                    )
                    fig_dfbetas.add_hline(
                        y=-dfbetas_threshold, line_dash="dash", line_color="red",
                        row=i+1, col=1
                    )
            
            fig_dfbetas.update_layout(
                height=200 * len(var_names),
                title_text="DFBETAS for Each Parameter",
                showlegend=False
            )
            
            st.plotly_chart(fig_dfbetas, use_container_width=True)

    
    # Outlier Detection Summary
    st.subheader("🎯 Outlier Detection Summary")
    
    outlier_summary = pd.DataFrame({
        'Diagnostic': ['High Leverage', "High Cook's Distance", 'High DFFITS', 'High Residuals'],
        'Count': [int(sum(high_leverage)), int(sum(high_cooks)), 
                  int(sum(high_dffits)), int(sum(high_resid))],
        'Percentage': [sum(high_leverage)/len(y)*100, sum(high_cooks)/len(y)*100,
                       sum(high_dffits)/len(y)*100, sum(high_resid)/len(y)*100],
        'Threshold': [f"{high_leverage_threshold:.3f}", f"{high_cooks_threshold:.3f}",
                      f"{high_dffits_threshold:.3f}", "±2.0"],
        'Interpretation': ['Unusual X values', 'High overall influence', 
                          'High fitted value influence', 'Unusual Y values']
    })
    
    st.dataframe(outlier_summary, use_container_width=True)
    
    # Download influence analysis
    influence_csv = influence_df.to_csv(index=False)
    st.download_button(
        label="📥 Download Complete Influence Analysis",
        data=influence_csv,
        file_name="advanced_diagnostics_influence.csv",
        mime="text/csv",
        key="download_influence_advanced"
    )
    
    # Recommendations
    if problematic_count > 0:
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">📖 Recommendations for Influential Observations:</h5>
        <div class="red-list">
        1. <strong>Investigate Data Quality:</strong> Check for data entry errors<br>
        2. <strong>Substantive Review:</strong> Determine if observations are valid<br>
        3. <strong>Sensitivity Analysis:</strong> Re-run analysis without influential points<br>
        4. <strong>Robust Methods:</strong> Consider robust regression techniques<br>
        5. <strong>Additional Data:</strong> Collect more observations if possible
        </div>
        </div>
        """, unsafe_allow_html=True)

def model_comparison_page():
    """Professional model comparison with multiple techniques"""
    
    if st.session_state.data is None:
        st.warning("⚠️ Please upload data first!")
        return
    
    st.markdown('<h2 class="section-header">⚖️ Advanced Model Comparison</h2>', unsafe_allow_html=True)
    
    data = st.session_state.data
    numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
    
    # Variable selection
    st.subheader("🎯 Model Specification")
    col1, col2 = st.columns(2)
    
    with col1:
        y_var = st.selectbox("📊 Dependent Variable", numeric_cols, key="comp_y_var")
    
    with col2:
        x_vars = st.multiselect(
            "📈 Independent Variables", 
            [col for col in numeric_cols if col != y_var],
            key="comp_x_vars"
        )
    
    if len(x_vars) < 1:
        st.warning("⚠️ Please select at least one independent variable!")
        return
    
    # Model selection options
    st.subheader("🔧 Model Selection")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        include_ols = st.checkbox("📊 OLS Regression", value=True)
        include_ridge = st.checkbox("🏔️ Ridge Regression", value=True)
        include_lasso = st.checkbox("🎯 Lasso Regression", value=True)
    
    with col2:
        include_polynomial = st.checkbox("📈 Polynomial Regression", value=True)
        include_interaction = st.checkbox("🔗 Interaction Model", value=len(x_vars) >= 2)
        include_log = st.checkbox("📊 Log-Linear Model", value=True)
    
    with col3:
        # Cross-validation settings
        cv_folds = st.selectbox("🔄 CV Folds", [3, 5, 10], index=1)
        test_size = st.selectbox("📊 Test Split", [0.2, 0.3, 0.4], index=0)
    
    if st.button("🚀 **Run Model Comparison**", type="primary", key="run_model_comparison"):
        with st.spinner("🔄 Running comprehensive model comparison..."):
            run_model_comparison_analysis(data, y_var, x_vars, {
                'include_ols': include_ols,
                'include_ridge': include_ridge,
                'include_lasso': include_lasso,
                'include_polynomial': include_polynomial,
                'include_interaction': include_interaction,
                'include_log': include_log,
                'cv_folds': cv_folds,
                'test_size': test_size
            })

def run_model_comparison_analysis(data, y_var, x_vars, options):
    """Run comprehensive model comparison analysis"""
    
    try:
        from sklearn.model_selection import train_test_split, cross_val_score
        from sklearn.preprocessing import StandardScaler, PolynomialFeatures
        from sklearn.linear_model import Ridge, Lasso, ElasticNet
        from sklearn.ensemble import RandomForestRegressor
        from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
        import warnings
        warnings.filterwarnings('ignore')
        
        # Prepare data
        y = data[y_var].values
        X = data[x_vars].values
        
        # Train-test split
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=options['test_size'], random_state=42
        )
        
        # Scale data for regularized models
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)
        
        results_comparison = []
        model_objects = {}
        
        # 1. OLS Regression
        if options['include_ols']:
            X_train_const = sm.add_constant(X_train)
            X_test_const = sm.add_constant(X_test)
            
            ols_model = sm.OLS(y_train, X_train_const).fit()
            y_pred_ols = ols_model.predict(X_test_const)
            
            # Cross-validation for OLS
            cv_scores = []
            from sklearn.model_selection import KFold
            kf = KFold(n_splits=options['cv_folds'], shuffle=True, random_state=42)
            
            for train_idx, val_idx in kf.split(X_train):
                X_cv_train, X_cv_val = X_train[train_idx], X_train[val_idx]
                y_cv_train, y_cv_val = y_train[train_idx], y_train[val_idx]
                
                X_cv_train_const = sm.add_constant(X_cv_train)
                X_cv_val_const = sm.add_constant(X_cv_val)
                
                cv_model = sm.OLS(y_cv_train, X_cv_train_const).fit()
                cv_pred = cv_model.predict(X_cv_val_const)
                cv_scores.append(r2_score(y_cv_val, cv_pred))
            
            results_comparison.append({
                'Model': 'OLS Regression',
                'Train_R²': ols_model.rsquared,
                'Test_R²': r2_score(y_test, y_pred_ols),
                'CV_R²_Mean': np.mean(cv_scores),
                'CV_R²_Std': np.std(cv_scores),
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_ols)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_ols),
                'AIC': ols_model.aic,
                'BIC': ols_model.bic,
                'Parameters': len(ols_model.params)
            })
            
            model_objects['OLS'] = ols_model
        
        # 2. Ridge Regression
        if options['include_ridge']:
            from sklearn.linear_model import RidgeCV
            
            ridge = RidgeCV(alphas=[0.1, 1.0, 10.0, 100.0], cv=options['cv_folds'])
            ridge.fit(X_train_scaled, y_train)
            y_pred_ridge = ridge.predict(X_test_scaled)
            
            # Cross-validation scores
            cv_scores_ridge = cross_val_score(ridge, X_train_scaled, y_train, 
                                            cv=options['cv_folds'], scoring='r2')
            
            results_comparison.append({
                'Model': f'Ridge (α={ridge.alpha_:.3f})',
                'Train_R²': ridge.score(X_train_scaled, y_train),
                'Test_R²': r2_score(y_test, y_pred_ridge),
                'CV_R²_Mean': np.mean(cv_scores_ridge),
                'CV_R²_Std': np.std(cv_scores_ridge),
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_ridge)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_ridge),
                'AIC': 'N/A',
                'BIC': 'N/A',
                'Parameters': f'{len(x_vars)+1} (regularized)'
            })
            
            model_objects['Ridge'] = ridge
        
        # 3. Lasso Regression
        if options['include_lasso']:
            from sklearn.linear_model import LassoCV
            
            lasso = LassoCV(alphas=[0.01, 0.1, 1.0, 10.0], cv=options['cv_folds'], max_iter=2000)
            lasso.fit(X_train_scaled, y_train)
            y_pred_lasso = lasso.predict(X_test_scaled)
            
            # Cross-validation scores
            cv_scores_lasso = cross_val_score(lasso, X_train_scaled, y_train, 
                                            cv=options['cv_folds'], scoring='r2')
            
            # Count non-zero coefficients
            non_zero_coefs = np.sum(np.abs(lasso.coef_) > 1e-6)
            
            results_comparison.append({
                'Model': f'Lasso (α={lasso.alpha_:.3f})',
                'Train_R²': lasso.score(X_train_scaled, y_train),
                'Test_R²': r2_score(y_test, y_pred_lasso),
                'CV_R²_Mean': np.mean(cv_scores_lasso),
                'CV_R²_Std': np.std(cv_scores_lasso),
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_lasso)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_lasso),
                'AIC': 'N/A',
                'BIC': 'N/A',
                'Parameters': f'{non_zero_coefs}/{len(x_vars)+1} (selected)'
            })
            
            model_objects['Lasso'] = lasso
        
        # 4. Polynomial Regression
        if options['include_polynomial'] and len(x_vars) <= 3:
            poly_features = PolynomialFeatures(degree=2, include_bias=True)
            X_train_poly = poly_features.fit_transform(X_train)
            X_test_poly = poly_features.transform(X_test)
            
            poly_model = sm.OLS(y_train, X_train_poly).fit()
            y_pred_poly = poly_model.predict(X_test_poly)
            
            results_comparison.append({
                'Model': 'Polynomial (degree=2)',
                'Train_R²': poly_model.rsquared,
                'Test_R²': r2_score(y_test, y_pred_poly),
                'CV_R²_Mean': 'N/A',
                'CV_R²_Std': 'N/A',
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_poly)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_poly),
                'AIC': poly_model.aic,
                'BIC': poly_model.bic,
                'Parameters': len(poly_model.params)
            })
            
            model_objects['Polynomial'] = poly_model
        
        # 5. Interaction Model
        if options['include_interaction'] and len(x_vars) >= 2:
            from itertools import combinations
            
            # Create interaction terms
            X_interaction = X_train.copy()
            X_test_interaction = X_test.copy()
            
            # Add interaction terms (only first two variables to keep manageable)
            interaction_term = X_train[:, 0] * X_train[:, 1]
            X_interaction = np.column_stack([X_interaction, interaction_term])
            
            interaction_term_test = X_test[:, 0] * X_test[:, 1]
            X_test_interaction = np.column_stack([X_test_interaction, interaction_term_test])
            
            X_interaction_const = sm.add_constant(X_interaction)
            X_test_interaction_const = sm.add_constant(X_test_interaction)
            
            interaction_model = sm.OLS(y_train, X_interaction_const).fit()
            y_pred_interaction = interaction_model.predict(X_test_interaction_const)
            
            results_comparison.append({
                'Model': f'Interaction ({x_vars[0]}×{x_vars[1]})',
                'Train_R²': interaction_model.rsquared,
                'Test_R²': r2_score(y_test, y_pred_interaction),
                'CV_R²_Mean': 'N/A',
                'CV_R²_Std': 'N/A',
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_interaction)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_interaction),
                'AIC': interaction_model.aic,
                'BIC': interaction_model.bic,
                'Parameters': len(interaction_model.params)
            })
            
            model_objects['Interaction'] = interaction_model
        
        # 6. Log-Linear Model (if all values are positive)
        if options['include_log'] and np.all(y > 0) and np.all(X > 0):
            log_y = np.log(y_train)
            log_X_train = np.log(X_train)
            log_X_test = np.log(X_test)
            
            log_X_train_const = sm.add_constant(log_X_train)
            log_X_test_const = sm.add_constant(log_X_test)
            
            log_model = sm.OLS(log_y, log_X_train_const).fit()
            log_y_pred = log_model.predict(log_X_test_const)
            y_pred_log = np.exp(log_y_pred)  # Transform back
            
            results_comparison.append({
                'Model': 'Log-Linear',
                'Train_R²': log_model.rsquared,
                'Test_R²': r2_score(y_test, y_pred_log),
                'CV_R²_Mean': 'N/A',
                'CV_R²_Std': 'N/A',
                'Test_RMSE': np.sqrt(mean_squared_error(y_test, y_pred_log)),
                'Test_MAE': mean_absolute_error(y_test, y_pred_log),
                'AIC': log_model.aic,
                'BIC': log_model.bic,
                'Parameters': len(log_model.params)
            })
            
            model_objects['Log-Linear'] = log_model
        
        # Display results
        display_model_comparison_results(results_comparison, model_objects, y_var, x_vars)
        
    except Exception as e:
        st.error(f"❌ Error in model comparison: {str(e)}")
        st.info("Please check your data and variable selections.")

def display_model_comparison_results(results_comparison, model_objects, y_var, x_vars):
    """Display comprehensive model comparison results"""
    
    # Create comparison dataframe
    comparison_df = pd.DataFrame(results_comparison)
    
    # Round numerical values
    numeric_cols = ['Train_R²', 'Test_R²', 'CV_R²_Mean', 'CV_R²_Std', 'Test_RMSE', 'Test_MAE', 'AIC', 'BIC']
    for col in numeric_cols:
        if col in comparison_df.columns:
            comparison_df[col] = pd.to_numeric(comparison_df[col], errors='coerce').round(4)
    
    st.subheader("📊 Model Comparison Results")
    st.dataframe(comparison_df, use_container_width=True, height=400)
    
    # Model performance visualization
    fig = go.Figure()
    
    # Test R² comparison
    models = comparison_df['Model'].tolist()
    test_r2 = comparison_df['Test_R²'].tolist()
    
    fig.add_trace(go.Bar(
        x=models,
        y=test_r2,
        name='Test R²',
        marker_color='lightblue',
        text=[f'{r:.3f}' for r in test_r2],
        textposition='auto'
    ))
    
    fig.update_layout(
        title='Model Performance Comparison (Test R²)',
        xaxis_title='Models',
        yaxis_title='Test R²',
        showlegend=False,
        height=500
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Best model identification
    st.subheader("🏆 Model Selection Results")
    
    # Find best models by different criteria
    best_test_r2_idx = comparison_df['Test_R²'].idxmax()
    best_cv_r2_idx = comparison_df['CV_R²_Mean'].idxmax() if 'CV_R²_Mean' in comparison_df.columns else None
    best_rmse_idx = comparison_df['Test_RMSE'].idxmin()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        st.markdown("🎯 **Best Test R²**")
        st.markdown(f"**{comparison_df.iloc[best_test_r2_idx]['Model']}**")
        st.markdown(f"R² = {comparison_df.iloc[best_test_r2_idx]['Test_R²']:.4f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        if best_cv_r2_idx is not None:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("🔄 **Best CV R²**")
            st.markdown(f"**{comparison_df.iloc[best_cv_r2_idx]['Model']}**")
            st.markdown(f"CV R² = {comparison_df.iloc[best_cv_r2_idx]['CV_R²_Mean']:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
    
    with col3:
        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
        st.markdown("📊 **Lowest RMSE**")
        st.markdown(f"**{comparison_df.iloc[best_rmse_idx]['Model']}**")
        st.markdown(f"RMSE = {comparison_df.iloc[best_rmse_idx]['Test_RMSE']:.4f}")
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Model selection criteria explanation
    st.markdown("""
    <div class="interpretation-box">
    <h5 class="red-text">🏆 Model Selection Criteria:</h5>
    <div class="red-list">
    • <strong>Highest Test R²:</strong> Best fit on unseen data (avoid overfitting)<br>
    • <strong>Highest CV R²:</strong> Most consistent performance across folds<br>
    • <strong>Lowest RMSE:</strong> Best prediction accuracy in original units<br>
    • <strong>Lowest AIC/BIC:</strong> Best balance of fit and model complexity<br>
    • <strong>Cross-validation:</strong> Most reliable estimate of out-of-sample performance
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Feature importance for regularized models
    if 'Lasso' in model_objects:
        st.subheader("🎯 Feature Importance (Lasso)")
        
        lasso_model = model_objects['Lasso']
        feature_importance = pd.DataFrame({
            'Feature': x_vars,
            'Coefficient': lasso_model.coef_,
            'Abs_Coefficient': np.abs(lasso_model.coef_),
            'Selected': np.abs(lasso_model.coef_) > 1e-6
        }).sort_values('Abs_Coefficient', ascending=False)
        
        # Plot feature importance
        fig = px.bar(
            feature_importance[feature_importance['Selected']], 
            x='Abs_Coefficient', 
            y='Feature',
            title='Lasso Feature Selection (Non-zero Coefficients)',
            orientation='h'
        )
        
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(feature_importance, use_container_width=True)
    
    # Download comparison results
    csv_comparison = comparison_df.to_csv(index=False)
    st.download_button(
        label="📥 Download Model Comparison Results",
        data=csv_comparison,
        file_name=f"model_comparison_{y_var}.csv",
        mime="text/csv",
        key="download_model_comparison"
    )
    
    # Recommendations
    st.markdown("""
    <div class="interpretation-box">
    <h5 class="red-text">💡 Model Selection Recommendations:</h5>
    <div class="red-list">
    1. <strong>Prefer simpler models</strong> if performance is similar (Occam's Razor)<br>
    2. <strong>Check cross-validation results</strong> for model stability<br>
    3. <strong>Consider domain knowledge</strong> when selecting features<br>
    4. <strong>Validate on truly unseen data</strong> if possible<br>
    5. <strong>Regularized models</strong> often generalize better with limited data
    </div>
    </div>
    """, unsafe_allow_html=True)

def time_series_features():
    """Complete time series regression analysis"""
    
    if st.session_state.data is None:
        st.warning("⚠️ Please upload data first!")
        return
    
    st.markdown('<h2 class="section-header">📈 Time Series Regression Analysis</h2>', 
                unsafe_allow_html=True)
    
    data = st.session_state.data
    
    # Check if data has date/time columns
    date_cols = []
    for col in data.columns:
        if data[col].dtype == 'object':
            try:
                pd.to_datetime(data[col].head())
                date_cols.append(col)
            except:
                pass
    
    numeric_cols = data.select_dtypes(include=[np.number]).columns.tolist()
    
    st.subheader("📅 Time Series Setup")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Date column selection
        if date_cols:
            date_col = st.selectbox("📅 Date/Time Column", date_cols)
            has_date = True
        else:
            st.warning("⚠️ No date columns detected. Using row index as time.")
            has_date = False
            date_col = None
    
    with col2:
        # Variable selection
        y_var = st.selectbox("📊 Dependent Variable (Y)", numeric_cols)
        x_vars = st.multiselect(
            "📈 Independent Variables (X)", 
            [col for col in numeric_cols if col != y_var],
            default=[col for col in numeric_cols if col != y_var][:2]
        )
    
    if not x_vars:
        st.warning("⚠️ Please select at least one independent variable!")
        return
    
    # Time series options
    st.subheader("⚙️ Time Series Options")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        include_trend = st.checkbox("📈 Include Linear Trend", value=True)
        include_seasonal = st.checkbox("🔄 Include Seasonal Dummies", value=False)
    
    with col2:
        include_lags = st.checkbox("⏱️ Include Lagged Variables", value=True)
        if include_lags:
            max_lags = st.selectbox("Max Lags", [1, 2, 3, 4], index=0)
        else:
            max_lags = 0
    
    with col3:
        test_autocorr = st.checkbox("🔍 Test for Autocorrelation", value=True)
        test_stationarity = st.checkbox("📊 Test for Stationarity", value=True)
    
    if st.button("🚀 **Run Time Series Analysis**", type="primary", key="run_ts_analysis"):
        with st.spinner("🔄 Running time series analysis..."):
            run_time_series_analysis(data, y_var, x_vars, {
                'date_col': date_col,
                'has_date': has_date,
                'include_trend': include_trend,
                'include_seasonal': include_seasonal,
                'include_lags': include_lags,
                'max_lags': max_lags,
                'test_autocorr': test_autocorr,
                'test_stationarity': test_stationarity
            })

def run_time_series_analysis(data, y_var, x_vars, options):
    """Run comprehensive time series regression analysis"""
    
    try:
        # Prepare time series data
        ts_data = data.copy()
        
        # Create time index
        if options['has_date'] and options['date_col']:
            ts_data['date'] = pd.to_datetime(ts_data[options['date_col']])
            ts_data = ts_data.sort_values('date')
            ts_data.set_index('date', inplace=True)
        else:
            ts_data.index = range(len(ts_data))
        
        # Create trend variable
        if options['include_trend']:
            ts_data['trend'] = range(1, len(ts_data) + 1)
            x_vars_extended = x_vars + ['trend']
        else:
            x_vars_extended = x_vars.copy()
        
        # Create seasonal dummies (if monthly/quarterly data)
        if options['include_seasonal'] and options['has_date']:
            ts_data['month'] = ts_data.index.month
            for month in range(2, 13):  # Exclude January as reference
                ts_data[f'month_{month}'] = (ts_data['month'] == month).astype(int)
                x_vars_extended.append(f'month_{month}')
        
        # Create lagged variables
        if options['include_lags'] and options['max_lags'] > 0:
            for var in [y_var] + x_vars:
                for lag in range(1, options['max_lags'] + 1):
                    lag_name = f'{var}_lag{lag}'
                    ts_data[lag_name] = ts_data[var].shift(lag)
                    if var != y_var:  # Don't include lagged dependent variable in x_vars yet
                        x_vars_extended.append(lag_name)
            
            # Remove rows with NaN due to lagging
            ts_data = ts_data.dropna()
        
        # Prepare regression data
        y = ts_data[y_var].values
        X = ts_data[x_vars_extended].values
        X_with_const = sm.add_constant(X)
        
        # Fit time series regression
        ts_model = sm.OLS(y, X_with_const).fit()
        
        # Display results
        st.subheader("📊 Time Series Regression Results")
        
        # Model summary
        st.text(str(ts_model.summary()))
        
        # Time series specific tests
        st.subheader("🔍 Time Series Diagnostic Tests")
        
        # 1. Durbin-Watson test for serial correlation
        dw_stat = durbin_watson(ts_model.resid)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("🔄 **Durbin-Watson Test**")
            st.markdown(f"Statistic: {dw_stat:.4f}")
            if 1.5 <= dw_stat <= 2.5:
                st.markdown("✅ No autocorrelation")
            else:
                st.markdown("❌ Autocorrelation detected")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # 2. Breusch-Godfrey test for higher-order serial correlation
        if options['test_autocorr']:
            from statsmodels.stats.diagnostic import acorr_breusch_godfrey
            
            try:
                bg_test = acorr_breusch_godfrey(ts_model, nlags=4)
                with col2:
                    st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                    st.markdown("🔍 **Breusch-Godfrey Test**")
                    st.markdown(f"LM Statistic: {bg_test[0]:.4f}")
                    st.markdown(f"p-value: {bg_test[1]:.4f}")
                    if bg_test[1] < 0.05:
                        st.markdown("❌ Serial correlation")
                    else:
                        st.markdown("✅ No serial correlation")
                    st.markdown('</div>', unsafe_allow_html=True)
            except:
                with col2:
                    st.info("Breusch-Godfrey test unavailable")
        
        # 3. Stationarity tests
        if options['test_stationarity']:
            st.subheader("📊 Stationarity Tests")
            
            from statsmodels.tsa.stattools import adfuller, kpss
            
            # Augmented Dickey-Fuller test
            try:
                adf_result = adfuller(y)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                    st.markdown("📈 **Augmented Dickey-Fuller Test**")
                    st.markdown(f"ADF Statistic: {adf_result[0]:.4f}")
                    st.markdown(f"p-value: {adf_result[1]:.4f}")
                    if adf_result[1] < 0.05:
                        st.markdown("✅ Series is stationary")
                    else:
                        st.markdown("❌ Series may be non-stationary")
                    st.markdown('</div>', unsafe_allow_html=True)
                
                # KPSS test
                try:
                    kpss_result = kpss(y)
                    with col2:
                        st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                        st.markdown("🔍 **KPSS Test**")
                        st.markdown(f"KPSS Statistic: {kpss_result[0]:.4f}")
                        st.markdown(f"p-value: {kpss_result[1]:.4f}")
                        if kpss_result[1] > 0.05:
                            st.markdown("✅ Series is stationary")
                        else:
                            st.markdown("❌ Series may be non-stationary")
                        st.markdown('</div>', unsafe_allow_html=True)
                except:
                    with col2:
                        st.info("KPSS test unavailable")
            except:
                st.info("Stationarity tests unavailable for this data")
        
        # Time series plots
        st.subheader("📈 Time Series Visualization")
        
        # Actual vs Fitted plot
        fig = go.Figure()
        
        if options['has_date']:
            x_axis = ts_data.index
        else:
            x_axis = range(len(y))
        
        fig.add_trace(go.Scatter(
            x=x_axis,
            y=y,
            mode='lines',
            name='Actual',
            line=dict(color='blue')
        ))
        
        fig.add_trace(go.Scatter(
            x=x_axis,
            y=ts_model.fittedvalues,
            mode='lines',
            name='Fitted',
            line=dict(color='red', dash='dash')
        ))
        
        fig.update_layout(
            title=f'Time Series: Actual vs Fitted Values for {y_var}',
            xaxis_title='Time',
            yaxis_title=y_var,
            hovermode='x'
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Residuals over time
        fig_resid = go.Figure()
        
        fig_resid.add_trace(go.Scatter(
            x=x_axis,
            y=ts_model.resid,
            mode='lines+markers',
            name='Residuals',
            line=dict(color='red')
        ))
        
        fig_resid.add_hline(y=0, line_dash="dash", line_color="black")
        
        fig_resid.update_layout(
            title='Residuals Over Time',
            xaxis_title='Time',
            yaxis_title='Residuals'
        )
        
        st.plotly_chart(fig_resid, use_container_width=True)
        
        # ACF/PACF plots for residuals
        st.subheader("🔄 Autocorrelation Analysis")
        
        from statsmodels.tsa.stattools import acf, pacf
        
        try:
            # Calculate ACF and PACF
            acf_values = acf(ts_model.resid, nlags=min(20, len(ts_model.resid)//4))
            pacf_values = pacf(ts_model.resid, nlags=min(20, len(ts_model.resid)//4))
            
            fig_acf, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 5))
            
            # ACF plot
            ax1.bar(range(len(acf_values)), acf_values)
            ax1.axhline(y=0, color='black')
            ax1.axhline(y=1.96/np.sqrt(len(ts_model.resid)), color='red', linestyle='--')
            ax1.axhline(y=-1.96/np.sqrt(len(ts_model.resid)), color='red', linestyle='--')
            ax1.set_title('Autocorrelation Function (ACF) of Residuals')
            ax1.set_xlabel('Lags')
            ax1.set_ylabel('ACF')
            
            # PACF plot
            ax2.bar(range(len(pacf_values)), pacf_values)
            ax2.axhline(y=0, color='black')
            ax2.axhline(y=1.96/np.sqrt(len(ts_model.resid)), color='red', linestyle='--')
            ax2.axhline(y=-1.96/np.sqrt(len(ts_model.resid)), color='red', linestyle='--')
            ax2.set_title('Partial Autocorrelation Function (PACF) of Residuals')
            ax2.set_xlabel('Lags')
            ax2.set_ylabel('PACF')
            
            st.pyplot(fig_acf)
            
        except Exception as e:
            st.info(f"ACF/PACF plots unavailable: {str(e)}")
        
        # Forecasting (simple one-step ahead)
        st.subheader("🔮 Simple Forecasting")
        
        if len(ts_data) > 10:
            # Use last 80% for training, 20% for testing
            split_point = int(0.8 * len(ts_data))
            
            # Re-fit model on training data
            y_train = ts_data[y_var].iloc[:split_point].values
            X_train = ts_data[x_vars_extended].iloc[:split_point].values
            X_train_const = sm.add_constant(X_train)
            
            forecast_model = sm.OLS(y_train, X_train_const).fit()
            
            # Forecast on test data
            y_test = ts_data[y_var].iloc[split_point:].values
            X_test = ts_data[x_vars_extended].iloc[split_point:].values
            X_test_const = sm.add_constant(X_test)
            
            y_forecast = forecast_model.predict(X_test_const)
            
            # Forecast accuracy
            mse_forecast = mean_squared_error(y_test, y_forecast)
            mae_forecast = mean_absolute_error(y_test, y_forecast)
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                st.markdown("🎯 **Forecast MSE**")
                st.markdown(f"{mse_forecast:.4f}")
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                st.markdown("📊 **Forecast MAE**")
                st.markdown(f"{mae_forecast:.4f}")
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col3:
                st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
                st.markdown("📈 **Forecast R²**")
                forecast_r2 = r2_score(y_test, y_forecast)
                st.markdown(f"{forecast_r2:.4f}")
                st.markdown('</div>', unsafe_allow_html=True)
        
        # Download time series results
        ts_results = pd.DataFrame({
            'Time': x_axis,
            'Actual': y,
            'Fitted': ts_model.fittedvalues,
            'Residuals': ts_model.resid
        })
        
        csv_ts = ts_results.to_csv(index=False)
        st.download_button(
            label="📥 Download Time Series Results",
            data=csv_ts,
            file_name=f"time_series_analysis_{y_var}.csv",
            mime="text/csv",
            key="download_ts_results"
        )
        
        # Time series recommendations
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">📈 Time Series Analysis Recommendations:</h5>
        <div class="red-list">
        1. <strong>Check for autocorrelation</strong> - use robust standard errors if present<br>
        2. <strong>Test for stationarity</strong> - difference the series if non-stationary<br>
        3. <strong>Include appropriate lags</strong> based on ACF/PACF analysis<br>
        4. <strong>Consider seasonality</strong> for monthly/quarterly data<br>
        5. <strong>Use specialized time series models</strong> (ARIMA, VAR) for complex patterns
        </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"❌ Error in time series analysis: {str(e)}")
        st.info("Please check your data format and variable selections.")


def predictions_page(language):
    """Complete prediction interface with confidence intervals"""
    
    if st.session_state.regression_results is None:
        st.warning("⚠️ Please run regression analysis first!")
        return
    
    st.markdown('<h2 class="section-header">🔮 Predictions & Confidence Intervals</h2>', 
                unsafe_allow_html=True)
    
    model = st.session_state.regression_results['model']
    independent_vars = st.session_state.regression_results['independent_vars']
    dependent_var = st.session_state.regression_results['dependent_var']
    data = st.session_state.data
    
    # Mathematical explanation
    st.markdown("""
    <div class="formula-box">
    <h4 class="red-header">📚 Types of Predictions</h4>
    <div class="red-text">
    <strong>1. Point Prediction:</strong> Single best estimate: Ŷ = β₀ + β₁X₁ + ... + βₖXₖ<br>
    <strong>2. Confidence Interval:</strong> Range for the MEAN of Y at given X values<br>
    <strong>3. Prediction Interval:</strong> Range for an INDIVIDUAL Y value (wider)<br><br>
    <strong>Key Difference:</strong> Prediction intervals account for both estimation uncertainty AND individual variation
    </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mathematical formulas
    st.markdown("""
    <div class="formula-box">
    <h5 class="red-header">🧮 Mathematical Formulas</h5>
    """, unsafe_allow_html=True)
    
    st.latex(r"\hat{Y} = \hat{\beta}_0 + \hat{\beta}_1 X_1 + ... + \hat{\beta}_k X_k")
    
    st.markdown("**Confidence Interval for Mean Response:**")
    st.latex(r"\hat{Y} \pm t_{\alpha/2} \cdot SE(\hat{Y})")
    
    st.markdown("**Prediction Interval for Individual Response:**")
    st.latex(r"\hat{Y} \pm t_{\alpha/2} \cdot \sqrt{MSE(1 + \mathbf{x}_0'(\mathbf{X}'\mathbf{X})^{-1}\mathbf{x}_0)}")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Prediction interface
    st.subheader("📊 Make New Predictions")
    
    # Choose prediction method
    prediction_method = st.selectbox(
        "🎯 Prediction Method",
        ["Single Prediction", "Batch Predictions", "Scenario Analysis"],
        help="Choose how you want to make predictions"
    )
    
    if prediction_method == "Single Prediction":
        make_single_prediction(model, independent_vars, dependent_var, data)
    elif prediction_method == "Batch Predictions":
        make_batch_predictions(model, independent_vars, dependent_var, data)
    else:
        make_scenario_analysis(model, independent_vars, dependent_var, data)

def make_single_prediction(model, independent_vars, dependent_var, data):
    """Make single prediction with intervals"""
    
    st.markdown("### 🎯 Single Prediction")
    
    col1, col2 = st.columns([2, 1])
    
    # Get input values for prediction
    new_values = {}
    
    with col1:
        st.markdown("**Enter values for prediction:**")
        
        # Create input fields for each variable
        for i, var in enumerate(independent_vars):
            col_data = data[var]
            min_val = float(col_data.min())
            max_val = float(col_data.max())
            mean_val = float(col_data.mean())
            std_val = float(col_data.std())
            
            # Show variable statistics
            with st.expander(f"📊 {var} Statistics", expanded=False):
                st.write(f"**Min**: {min_val:.3f}")
                st.write(f"**Max**: {max_val:.3f}")
                st.write(f"**Mean**: {mean_val:.3f}")
                st.write(f"**Std**: {std_val:.3f}")
            
            new_values[var] = st.number_input(
                f"**{var}**",
                min_value=min_val * 0.1,  # Allow some extrapolation
                max_value=max_val * 2.0,
                value=mean_val,
                step=(max_val - min_val) / 100,
                help=f"Range in data: {min_val:.2f} to {max_val:.2f}",
                key=f"pred_input_{var}_{i}"
            )
    
    with col2:
        st.markdown("**Prediction Settings:**")
        
        confidence_level = st.slider(
            "Confidence Level (%)",
            min_value=80,
            max_value=99,
            value=95,
            step=1,
            help="Higher confidence = wider intervals",
            key="conf_level_pred"
        )
        
        show_diagnostics = st.checkbox(
            "Show prediction diagnostics",
            value=True,
            help="Display leverage and influence for prediction point"
        )
        
        alpha = (100 - confidence_level) / 100
        
        if st.button("🚀 **Calculate Prediction**", type="primary", key="calc_pred_btn"):
            calculate_single_prediction(model, independent_vars, dependent_var, 
                                      new_values, alpha, confidence_level, show_diagnostics)

def calculate_single_prediction(model, independent_vars, dependent_var, new_values, 
                               alpha, confidence_level, show_diagnostics):
    """Calculate single prediction with full diagnostics"""
    
    try:
        # Prepare prediction data
        X_new = [1] + [new_values.get(var, 0) for var in independent_vars]  # Add intercept
        X_new_array = np.array(X_new).reshape(1, -1)
        
        # Point prediction
        point_pred = model.predict(X_new_array)[0]
        
        # Get design matrix for interval calculations
        X_original = st.session_state.regression_results['X_with_const']
        
        # Calculate prediction standard error
        x_new_centered = X_new_array
        
        # Variance-covariance matrix
        try:
            cov_matrix = model.cov_params()
            pred_var = np.dot(np.dot(x_new_centered, cov_matrix), x_new_centered.T)[0, 0]
        except:
            # Fallback calculation
            pred_var = model.mse_resid * (1 + np.sum(x_new_centered ** 2) / len(model.resid))
        
        pred_se = np.sqrt(pred_var)
        conf_se = np.sqrt(pred_var - model.mse_resid)
        
        # Critical t-value
        t_crit = stats.t.ppf(1 - alpha/2, df=model.df_resid)
        
        # Intervals
        conf_lower = point_pred - t_crit * conf_se
        conf_upper = point_pred + t_crit * conf_se
        
        pred_lower = point_pred - t_crit * pred_se
        pred_upper = point_pred + t_crit * pred_se
        
        # Display results
        st.markdown("---")
        st.subheader("🎯 Prediction Results")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("📍 **Point Prediction**")
            st.markdown(f"**{point_pred:.4f}**")
            st.markdown("Best single estimate")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("📊 **Confidence Interval**")
            st.markdown(f"**[{conf_lower:.4f}, {conf_upper:.4f}]**")
            st.markdown("For the AVERAGE response")
            st.markdown(f"Width: {conf_upper - conf_lower:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col3:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("🎯 **Prediction Interval**")
            st.markdown(f"**[{pred_lower:.4f}, {pred_upper:.4f}]**")
            st.markdown("For an INDIVIDUAL response")
            st.markdown(f"Width: {pred_upper - pred_lower:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Interpretation
        st.markdown(f"""
        <div class="interpretation-box">
        <h5 class="red-header">📖 Detailed Interpretation:</h5>
        <div class="red-text">
        <strong>Point Prediction:</strong> Our best estimate is that {dependent_var} = {point_pred:.4f}<br><br>
        
        <strong>Confidence Interval ({confidence_level}%):</strong><br>
        We are {confidence_level}% confident that the AVERAGE {dependent_var} for all individuals 
        with these characteristics is between {conf_lower:.4f} and {conf_upper:.4f}<br><br>
        
        <strong>Prediction Interval ({confidence_level}%):</strong><br>
        We are {confidence_level}% confident that a SPECIFIC individual with these 
        characteristics will have {dependent_var} between {pred_lower:.4f} and {pred_upper:.4f}<br><br>
        
        <strong>Why is the prediction interval wider?</strong><br>
        It includes both the uncertainty in our estimate of the mean (confidence interval) 
        AND the natural variation of individual observations around that mean.
        </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Prediction diagnostics
        if show_diagnostics:
            st.subheader("🔍 Prediction Diagnostics")
            
            # Calculate leverage for prediction point
            try:
                hat_matrix = np.dot(np.dot(X_original, np.linalg.inv(np.dot(X_original.T, X_original))), X_original.T)
                leverage_pred = np.dot(np.dot(x_new_centered, np.linalg.inv(np.dot(X_original.T, X_original))), x_new_centered.T)[0, 0]
                
                average_leverage = len(model.params) / len(model.resid)
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown('<div class="metric-box">', unsafe_allow_html=True)
                    st.metric("📊 Leverage", f"{leverage_pred:.4f}")
                    st.markdown(f"Average: {average_leverage:.4f}")
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col2:
                    st.markdown('<div class="metric-box">', unsafe_allow_html=True)
                    extrapolation = "Yes" if leverage_pred > 2 * average_leverage else "No"
                    st.metric("🔍 Extrapolation?", extrapolation)
                    st.markdown("High leverage indicates extrapolation")
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col3:
                    st.markdown('<div class="metric-box">', unsafe_allow_html=True)
                    reliability = "High" if leverage_pred <= average_leverage else ("Medium" if leverage_pred <= 2 * average_leverage else "Low")
                    st.metric("🎯 Reliability", reliability)
                    st.markdown("Based on leverage analysis")
                    st.markdown('</div>', unsafe_allow_html=True)
                
                if leverage_pred > 2 * average_leverage:
                    st.warning("⚠️ **High Leverage Detected**: This prediction involves extrapolation beyond the range of your data. Use with caution!")
                
            except Exception as e:
                st.info("Leverage diagnostics unavailable")
        
        # Create prediction results table for download
        pred_results = pd.DataFrame({
            'Variable': list(new_values.keys()),
            'Input_Value': list(new_values.values())
        })
        
        pred_summary = pd.DataFrame({
            'Prediction_Type': [
                'Point Prediction', 
                f'Confidence Interval Lower ({confidence_level}%)', 
                f'Confidence Interval Upper ({confidence_level}%)',
                f'Prediction Interval Lower ({confidence_level}%)', 
                f'Prediction Interval Upper ({confidence_level}%)'
            ],
            'Value': [point_pred, conf_lower, conf_upper, pred_lower, pred_upper],
            'Standard_Error': [np.nan, conf_se, conf_se, pred_se, pred_se]
        })
        
        # Display summary tables
        st.subheader("📋 Prediction Summary")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Input Values:**")
            st.dataframe(pred_results, use_container_width=True)
        
        with col2:
            st.markdown("**Prediction Results:**")
            st.dataframe(pred_summary, use_container_width=True)
        
        # Download prediction results
        pred_results_csv = pred_summary.to_csv(index=False)
        st.download_button(
            label="📥 Download Prediction Results",
            data=pred_results_csv,
            file_name=f"prediction_results_{dependent_var}.csv",
            mime="text/csv",
            key="download_single_pred"
        )
        
    except Exception as e:
        st.error(f"❌ Error calculating prediction: {str(e)}")
        st.info("Please check your input values and try again.")

def make_batch_predictions(model, independent_vars, dependent_var, data):
    """Make batch predictions from uploaded data or manual input"""
    
    st.markdown("### 📊 Batch Predictions")
    
    batch_method = st.selectbox(
        "📈 Batch Method",
        ["Upload CSV File", "Use Sample from Current Data", "Manual Grid Input"]
    )
    
    if batch_method == "Upload CSV File":
        st.markdown("**Upload a CSV file with the same variable names for batch predictions:**")
        
        required_cols = independent_vars
        st.info(f"📋 **Required columns**: {', '.join(required_cols)}")
        
        uploaded_pred_file = st.file_uploader(
            "Choose CSV file for predictions",
            type=['csv'],
            key="batch_pred_file"
        )
        
        if uploaded_pred_file is not None:
            try:
                pred_data = pd.read_csv(uploaded_pred_file)
                
                # Check if required columns exist
                missing_cols = [col for col in required_cols if col not in pred_data.columns]
                
                if missing_cols:
                    st.error(f"❌ Missing required columns: {', '.join(missing_cols)}")
                else:
                    st.success(f"✅ File loaded successfully! {len(pred_data)} rows for prediction.")
                    
                    if st.button("🚀 **Run Batch Predictions**", key="run_batch_pred"):
                        run_batch_predictions(model, pred_data, independent_vars, dependent_var)
                        
            except Exception as e:
                st.error(f"❌ Error loading file: {str(e)}")
    
    elif batch_method == "Use Sample from Current Data":
        st.markdown("**Use a random sample from your current dataset:**")
        
        sample_size = st.slider("Sample size", min_value=5, max_value=min(50, len(data)), value=10)
        
        if st.button("🎲 **Generate Sample & Predict**", key="sample_pred"):
            sample_data = data[independent_vars].sample(n=sample_size, random_state=42)
            run_batch_predictions(model, sample_data, independent_vars, dependent_var)
    
    else:  # Manual Grid Input
        st.markdown("**Create a grid of values for predictions:**")
        
        if len(independent_vars) <= 2:
            create_manual_grid_predictions(model, independent_vars, dependent_var, data)
        else:
            st.warning("⚠️ Manual grid input is only available for 1-2 variables. Please use file upload for more variables.")

def run_batch_predictions(model, pred_data, independent_vars, dependent_var):
    """Run batch predictions on provided data"""
    
    try:
        # Prepare prediction data
        X_pred = pred_data[independent_vars].values
        X_pred_const = sm.add_constant(X_pred)
        
        # Make predictions
        predictions = model.predict(X_pred_const)
        
        # Calculate prediction intervals (simplified for batch)
        pred_se = np.sqrt(model.mse_resid * (1 + np.diag(np.dot(np.dot(X_pred_const, model.cov_params()), X_pred_const.T))))
        t_crit = stats.t.ppf(0.975, df=model.df_resid)  # 95% confidence
        
        pred_lower = predictions - t_crit * pred_se
        pred_upper = predictions + t_crit * pred_se
        
        # Create results dataframe
        results_df = pred_data[independent_vars].copy()
        results_df['Predicted_' + dependent_var] = predictions
        results_df['Lower_95%'] = pred_lower
        results_df['Upper_95%'] = pred_upper
        results_df['Interval_Width'] = pred_upper - pred_lower
        
        # Display results
        st.subheader("📊 Batch Prediction Results")
        st.dataframe(results_df, use_container_width=True, height=400)
        
        # Summary statistics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown('<div class="metric-box">', unsafe_allow_html=True)
            st.metric("📊 Predictions Made", len(predictions))
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown('<div class="metric-box">', unsafe_allow_html=True)
            st.metric("📈 Mean Prediction", f"{np.mean(predictions):.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col3:
            st.markdown('<div class="metric-box">', unsafe_allow_html=True)
            st.metric("📊 Std Dev", f"{np.std(predictions):.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col4:
            st.markdown('<div class="metric-box">', unsafe_allow_html=True)
            st.metric("📏 Avg Interval Width", f"{np.mean(pred_upper - pred_lower):.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Visualization
        if len(independent_vars) == 1:
            # Scatter plot for single variable
            fig = px.scatter(
                x=pred_data[independent_vars[0]], 
                y=predictions,
                title=f'Batch Predictions: {dependent_var} vs {independent_vars[0]}'
            )
            
            # Add prediction intervals
            fig.add_trace(go.Scatter(
                x=pred_data[independent_vars[0]],
                y=pred_lower,
                mode='markers',
                marker=dict(color='red', symbol='triangle-down'),
                name='Lower 95%'
            ))
            
            fig.add_trace(go.Scatter(
                x=pred_data[independent_vars[0]],
                y=pred_upper,
                mode='markers',
                marker=dict(color='red', symbol='triangle-up'),
                name='Upper 95%'
            ))
            
            st.plotly_chart(fig, use_container_width=True)
        
        # Download results
        batch_csv = results_df.to_csv(index=False)
        st.download_button(
            label="📥 Download Batch Predictions",
            data=batch_csv,
            file_name=f"batch_predictions_{dependent_var}.csv",
            mime="text/csv",
            key="download_batch_pred"
        )
        
    except Exception as e:
        st.error(f"❌ Error in batch predictions: {str(e)}")

def make_scenario_analysis(model, independent_vars, dependent_var, data):
    """Scenario analysis for predictions"""
    
    st.markdown("### 🎭 Scenario Analysis")
    
    st.markdown("""
    Create different scenarios by setting optimistic, pessimistic, and baseline values for each variable.
    """)
    
    scenarios = {}
    
    for var in independent_vars:
        col_data = data[var]
        min_val = float(col_data.min())
        max_val = float(col_data.max())
        mean_val = float(col_data.mean())
        
        st.subheader(f"📊 Scenarios for {var}")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            pessimistic = st.number_input(
                f"😟 Pessimistic {var}",
                value=min_val,
                key=f"pess_{var}"
            )
        
        with col2:
            baseline = st.number_input(
                f"😐 Baseline {var}",
                value=mean_val,
                key=f"base_{var}"
            )
        
        with col3:
            optimistic = st.number_input(
                f"😊 Optimistic {var}",
                value=max_val,
                key=f"opt_{var}"
            )
        
        scenarios[var] = {
            'Pessimistic': pessimistic,
            'Baseline': baseline,
            'Optimistic': optimistic
        }
    
    if st.button("🚀 **Run Scenario Analysis**", key="run_scenario"):
        run_scenario_analysis(model, scenarios, independent_vars, dependent_var)

def run_scenario_analysis(model, scenarios, independent_vars, dependent_var):
    """Run scenario analysis with different combinations"""
    
    try:
        from itertools import product
        
        # Create all combinations of scenarios
        scenario_names = ['Pessimistic', 'Baseline', 'Optimistic']
        scenario_combinations = list(product(scenario_names, repeat=len(independent_vars)))
        
        results = []
        
        for combination in scenario_combinations:
            scenario_values = {}
            scenario_label = []
            
            for i, var in enumerate(independent_vars):
                scenario_type = combination[i]
                scenario_values[var] = scenarios[var][scenario_type]
                scenario_label.append(f"{var}:{scenario_type[0]}")  # First letter
            
            # Make prediction
            X_new = [1] + [scenario_values[var] for var in independent_vars]
            X_new_array = np.array(X_new).reshape(1, -1)
            prediction = model.predict(X_new_array)[0]
            
            result = {
                'Scenario': ' | '.join(scenario_label),
                'Prediction': prediction
            }
            
            # Add individual variable values
            for var in independent_vars:
                result[var] = scenario_values[var]
            
            results.append(result)
        
        # Create results dataframe
        results_df = pd.DataFrame(results)
        
        st.subheader("🎭 Scenario Analysis Results")
        st.dataframe(results_df, use_container_width=True)
        
        # Find best and worst scenarios
        best_scenario = results_df.loc[results_df['Prediction'].idxmax()]
        worst_scenario = results_df.loc[results_df['Prediction'].idxmin()]
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("📈 **Best Scenario**")
            st.markdown(f"**{best_scenario['Scenario']}**")
            st.markdown(f"Prediction: {best_scenario['Prediction']:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("📊 **Baseline**")
            baseline_row = results_df[results_df['Scenario'].str.contains('B')].iloc[0] if any('B' in s for s in results_df['Scenario']) else results_df.iloc[len(results_df)//2]
            st.markdown(f"**{baseline_row['Scenario']}**")
            st.markdown(f"Prediction: {baseline_row['Prediction']:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col3:
            st.markdown('<div class="advanced-metric">', unsafe_allow_html=True)
            st.markdown("📉 **Worst Scenario**")
            st.markdown(f"**{worst_scenario['Scenario']}**")
            st.markdown(f"Prediction: {worst_scenario['Prediction']:.4f}")
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Scenario comparison chart
        fig = px.bar(
            results_df,
            x='Scenario',
            y='Prediction',
            title='Scenario Analysis: Predicted Values',
            color='Prediction',
            color_continuous_scale='RdYlGn'
        )
        
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)
        
        # Sensitivity analysis
        st.subheader("📈 Sensitivity Analysis")
        
        # Calculate the range of predictions
        prediction_range = results_df['Prediction'].max() - results_df['Prediction'].min()
        
        st.info(f"📊 **Total Prediction Range**: {prediction_range:.4f} ({dependent_var} units)")
        st.info(f"📈 **Best vs Worst**: {((best_scenario['Prediction'] - worst_scenario['Prediction']) / worst_scenario['Prediction'] * 100):.1f}% difference")
        
        # Download scenario results
        scenario_csv = results_df.to_csv(index=False)
        st.download_button(
            label="📥 Download Scenario Analysis",
            data=scenario_csv,
            file_name=f"scenario_analysis_{dependent_var}.csv",
            mime="text/csv",
            key="download_scenario"
        )
        
    except Exception as e:
        st.error(f"❌ Error in scenario analysis: {str(e)}")


def export_results_feature():
    """Complete export regression results to various formats"""
    
    if st.session_state.regression_results is None:
        st.warning("⚠️ Please run regression analysis first!")
        return
    
    st.markdown('<h2 class="section-header">📤 Professional Export Results</h2>', unsafe_allow_html=True)
    
    model = st.session_state.regression_results['model']
    independent_vars = st.session_state.regression_results['independent_vars']
    dependent_var = st.session_state.regression_results['dependent_var']
    
    # Export format selection
    st.subheader("📋 Choose Export Format")
    
    col1, col2 = st.columns(2)
    
    with col1:
        export_format = st.selectbox(
            "🎯 Export Format:",
            [
                "📄 Summary Report (TXT)", 
                "📊 Excel Workbook", 
                "📝 LaTeX Table", 
                "📈 CSV Data", 
                "🐍 Python Code",
                "📊 PowerPoint Summary",
                "📋 Word Document"
            ]
        )
    
    with col2:
        include_diagnostics = st.checkbox("🔍 Include Diagnostics", value=True)
        include_plots = st.checkbox("📈 Include Plots", value=True)
        include_interpretation = st.checkbox("📖 Include Interpretation", value=True)
    
    # Format-specific options
    if "Excel" in export_format:
        st.subheader("📊 Excel Options")
        col1, col2 = st.columns(2)
        with col1:
            excel_sheets = st.multiselect(
                "📋 Sheets to Include:",
                ["Summary", "Coefficients", "Diagnostics", "Residuals", "Data Sample"],
                default=["Summary", "Coefficients", "Diagnostics"]
            )
        with col2:
            include_charts = st.checkbox("📊 Include Charts in Excel", value=True)
    
    elif "LaTeX" in export_format:
        st.subheader("📝 LaTeX Options")
        col1, col2 = st.columns(2)
        with col1:
            latex_style = st.selectbox("🎨 Table Style", ["Standard", "Booktabs", "Professional"])
        with col2:
            include_stars = st.checkbox("⭐ Significance Stars", value=True)
    
    # Generate export based on selection
    if st.button("🚀 **Generate Export**", type="primary", key="generate_export"):
        with st.spinner(f"🔄 Generating {export_format.split(' ')[1]} export..."):
            
            if "Summary Report" in export_format:
                generate_summary_report(model, independent_vars, dependent_var, 
                                      include_diagnostics, include_interpretation)
            
            elif "Excel Workbook" in export_format:
                generate_excel_workbook(model, independent_vars, dependent_var, 
                                       excel_sheets, include_charts, include_diagnostics)
            
            elif "LaTeX Table" in export_format:
                generate_latex_table(model, independent_vars, dependent_var, 
                                    latex_style, include_stars)
            
            elif "CSV Data" in export_format:
                generate_csv_export(model, independent_vars, dependent_var, include_diagnostics)
            
            elif "Python Code" in export_format:
                generate_python_code(model, independent_vars, dependent_var, include_diagnostics)
            
            elif "PowerPoint" in export_format:
                generate_powerpoint_summary(model, independent_vars, dependent_var)
            
            elif "Word Document" in export_format:
                generate_word_document(model, independent_vars, dependent_var, include_diagnostics)

def generate_summary_report(model, independent_vars, dependent_var, include_diagnostics, include_interpretation):
    """Generate comprehensive summary report"""
    
    try:
        # Create comprehensive text summary
        report_text = f"""
ECONOMETRICS REGRESSION ANALYSIS REPORT
=====================================

Generated by: Econometrics Learning Lab
Created by: HAMDI Boulanouar - Professional Statistical Analysis Platform
Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Dataset: {st.session_state.uploaded_file_name}

EXECUTIVE SUMMARY
================
Analysis Type: {'Simple' if len(independent_vars) == 1 else 'Multiple'} Linear Regression
Dependent Variable: {dependent_var}
Independent Variables: {', '.join(independent_vars)}
Sample Size: {int(model.nobs)} observations
Model R-squared: {model.rsquared:.4f} ({model.rsquared*100:.1f}% of variation explained)

VARIABLES ANALYZED
=================
Dependent Variable (Y): {dependent_var}
Independent Variables (X): {', '.join(independent_vars)}

DATA SUMMARY
============
Total Observations: {int(model.nobs)}
Degrees of Freedom: {int(model.df_resid)}
Missing Values Handling: Complete cases only

STATISTICAL MODEL
================
Model Equation: {dependent_var} = β₀ + β₁{independent_vars[0]}"""

        if len(independent_vars) > 1:
            for i, var in enumerate(independent_vars[1:], 2):
                report_text += f" + β{i}{var}"
        
        report_text += " + ε\n\nWhere:\n"
        report_text += f"- {dependent_var} is the dependent variable\n"
        for i, var in enumerate(independent_vars):
            report_text += f"- {var} is independent variable {i+1}\n"
        report_text += "- ε is the error term\n"

        report_text += f"""

REGRESSION RESULTS
=================
{model.summary()}

KEY FINDINGS
============
Model Performance:
- R-squared: {model.rsquared:.4f} (explains {model.rsquared*100:.1f}% of variation)
- Adjusted R-squared: {model.rsquared_adj:.4f}
- F-statistic: {model.fvalue:.4f} (p-value: {model.f_pvalue:.6f})
- Root Mean Square Error: {np.sqrt(model.mse_resid):.4f}

Model Significance:
"""
        if model.f_pvalue < 0.001:
            report_text += "- The model is HIGHLY statistically significant (p < 0.001)\n"
        elif model.f_pvalue < 0.05:
            report_text += "- The model is statistically significant (p < 0.05)\n"
        else:
            report_text += "- The model is NOT statistically significant (p ≥ 0.05)\n"
        
        report_text += "\nCoefficient Analysis:\n"
        
        # Handle parameter access
        try:
            if hasattr(model.params, 'index'):
                param_names = list(model.params.index)
                param_values = list(model.params.values)
                pvalue_values = list(model.pvalues.values)
            else:
                param_names = ['Intercept'] + independent_vars
                param_values = list(np.array(model.params).flatten())
                pvalue_values = list(np.array(model.pvalues).flatten())
        except:
            param_names = ['Intercept'] + independent_vars
            param_values = [0] * len(param_names)
            pvalue_values = [1] * len(param_names)
        
        for i, (name, coef, pval) in enumerate(zip(param_names, param_values, pvalue_values)):
            significance = "***" if pval < 0.01 else "**" if pval < 0.05 else "*" if pval < 0.10 else ""
            sig_text = "highly significant" if pval < 0.01 else ("significant" if pval < 0.05 else ("marginally significant" if pval < 0.10 else "not significant"))
            
            if name == 'Intercept':
                interpretation = f"When all variables are zero, {dependent_var} = {coef:.4f}"
            else:
                interpretation = f"A 1-unit increase in {name} changes {dependent_var} by {coef:.4f}"
                if len(independent_vars) > 1:
                    interpretation += " (holding other variables constant)"
            
            report_text += f"\n{name}: {coef:.4f} {significance}\n"
            report_text += f"  - {interpretation}\n"
            report_text += f"  - This effect is {sig_text} (p = {pval:.4f})\n"
        
        # Add diagnostics if requested
        if include_diagnostics:
            report_text += f"""

DIAGNOSTIC TESTS
===============
Durbin-Watson Test for Autocorrelation:
- Statistic: {durbin_watson(model.resid):.4f}
- Interpretation: {'No significant autocorrelation' if 1.5 <= durbin_watson(model.resid) <= 2.5 else 'Possible autocorrelation detected'}

Residual Analysis:
- Mean of residuals: {np.mean(model.resid):.6f} (should be close to 0)
- Standard deviation of residuals: {np.std(model.resid):.4f}
- Minimum residual: {np.min(model.resid):.4f}
- Maximum residual: {np.max(model.resid):.4f}

Assumption Checks:
- Linearity: Check residuals vs fitted plot
- Homoscedasticity: Check for constant variance in residuals
- Normality: Check Q-Q plot of residuals
- Independence: Durbin-Watson statistic provided above
"""

        # Add interpretation section
        if include_interpretation:
            report_text += f"""

PRACTICAL INTERPRETATION
=======================
Business/Research Implications:

1. MODEL QUALITY:
   {'This model explains a substantial portion' if model.rsquared > 0.7 else ('This model explains a moderate portion' if model.rsquared > 0.5 else 'This model explains a limited portion')} of the variation in {dependent_var} ({model.rsquared*100:.1f}%).

2. KEY RELATIONSHIPS:
"""
            for i, (name, coef, pval) in enumerate(zip(param_names, param_values, pvalue_values)):
                if name != 'Intercept' and pval < 0.10:
                    direction = "positively" if coef > 0 else "negatively"
                    strength = "strongly" if abs(coef) > np.std(st.session_state.data[name]) else "moderately"
                    report_text += f"   - {name} is {direction} and {strength} related to {dependent_var}\n"
            
            report_text += f"""

3. PREDICTION ACCURACY:
   The model's predictions have a typical error of ±{np.sqrt(model.mse_resid):.4f} {dependent_var} units.

4. STATISTICAL RELIABILITY:
   {'The relationships found are statistically reliable' if model.f_pvalue < 0.05 else 'The relationships may not be statistically reliable'} (F-test p-value: {model.f_pvalue:.4f}).

RECOMMENDATIONS
==============
1. Model Use: {'Suitable for prediction and inference' if model.f_pvalue < 0.05 and model.rsquared > 0.5 else 'Use with caution for prediction'}
2. Data Quality: {'Good model fit suggests data quality is adequate' if model.rsquared > 0.6 else 'Consider collecting additional data or variables'}
3. Further Analysis: {'Consider diagnostic tests for assumption violations' if include_diagnostics else 'Run diagnostic tests to validate assumptions'}
4. Practical Application: Apply insights to {'business decisions' if 'price' in dependent_var.lower() or 'sales' in dependent_var.lower() else 'research conclusions'}
"""

        report_text += f"""

TECHNICAL NOTES
==============
- Analysis performed using Ordinary Least Squares (OLS) regression
- All statistical tests conducted at conventional significance levels
- Model assumes linear relationships between variables
- Residual analysis recommended for assumption validation

METHODOLOGY
===========
Statistical Software: Python with statsmodels library
Estimation Method: Ordinary Least Squares (OLS)
Confidence Level: 95% (unless otherwise specified)
Missing Data: Listwise deletion (complete cases only)

APPENDIX
========
Model Summary Statistics:
- Log-Likelihood: {model.llf:.2f}
- AIC: {model.aic:.2f}
- BIC: {model.bic:.2f}
- Condition Number: {model.condition_number:.2f}

=====================================
End of Report

Generated by Econometrics Learning Lab
Created by HAMDI Boulanouar
Professional Statistical Analysis Platform
For more advanced features, visit the main application
=====================================
        """
        
        # Create download button
        st.success("✅ Summary report generated successfully!")
        
        st.download_button(
            label="📥 **Download Complete Report**",
            data=report_text.encode('utf-8'),
            file_name=f"econometrics_report_{dependent_var}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            key="download_summary_report",
            help="Download comprehensive analysis report"
        )
        
        # Show preview
        with st.expander("📄 **Report Preview**", expanded=False):
            st.text(report_text[:3000] + "\n\n[...Report continues...]" if len(report_text) > 3000 else report_text)
            
    except Exception as e:
        st.error(f"❌ Error generating summary report: {str(e)}")

def generate_excel_workbook(model, independent_vars, dependent_var, excel_sheets, include_charts, include_diagnostics):
    """Generate comprehensive Excel workbook"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import LineChart, ScatterChart, Reference
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Sheet 1: Model Summary
            if "Summary" in excel_sheets:
                summary_data = {
                    'Metric': [
                        'Analysis Date', 'Created By', 'Dataset', 'Model Type',
                        'Dependent Variable', 'Independent Variables', 'Sample Size',
                        'R-squared', 'Adjusted R-squared', 'F-statistic', 'F p-value',
                        'AIC', 'BIC', 'Root MSE', 'Log-Likelihood'
                    ],
                    'Value': [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'HAMDI Boulanouar - Econometrics Learning Lab',
                        st.session_state.uploaded_file_name,
                        'Simple Linear Regression' if len(independent_vars) == 1 else 'Multiple Linear Regression',
                        dependent_var,
                        ', '.join(independent_vars),
                        int(model.nobs),
                        f'{model.rsquared:.6f}',
                        f'{model.rsquared_adj:.6f}',
                        f'{model.fvalue:.4f}',
                        f'{model.f_pvalue:.6f}',
                        f'{model.aic:.2f}',
                        f'{model.bic:.2f}',
                        f'{np.sqrt(model.mse_resid):.4f}',
                        f'{model.llf:.2f}'
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Model_Summary', index=False)
            
            # Sheet 2: Coefficients
            if "Coefficients" in excel_sheets:
                coeffs_data = get_coefficients_data()
                coeffs_df = pd.DataFrame(coeffs_data)
                
                # Add interpretation column
                interpretation = []
                for i, var in enumerate(coeffs_data['variables']):
                    coef = coeffs_data['coefficients'][i]
                    pval = coeffs_data['p_values'][i]
                    
                    if var == 'Intercept':
                        interp = f"Baseline value when all X variables = 0"
                    else:
                        direction = "increases" if coef > 0 else "decreases"
                        significance = "significant" if pval < 0.05 else "not significant"
                        interp = f"1-unit increase in {var} {direction} {dependent_var} by {abs(coef):.4f} ({significance})"
                    
                    interpretation.append(interp)
                
                coeffs_df['Interpretation'] = interpretation
                coeffs_df.to_excel(writer, sheet_name='Coefficients', index=False)
            
            # Sheet 3: Diagnostics
            if "Diagnostics" in excel_sheets and include_diagnostics:
                try:
                    from statsmodels.stats.diagnostic import het_breuschpagan
                    
                    X_with_const = st.session_state.regression_results['X_with_const']
                    bp_stat, bp_pvalue, _, _ = het_breuschpagan(model.resid, X_with_const)
                    dw_stat = durbin_watson(model.resid)
                    
                    try:
                        jb_result = jarque_bera(model.resid)
                        jb_stat, jb_pvalue = jb_result[0], jb_result[1]
                    except:
                        jb_stat, jb_pvalue = 'N/A', 'N/A'
                    
                    diagnostics_data = {
                        'Test': [
                            'Breusch-Pagan (Heteroscedasticity)',
                            'Durbin-Watson (Autocorrelation)',
                            'Jarque-Bera (Normality)'
                        ],
                        'Statistic': [f'{bp_stat:.4f}', f'{dw_stat:.4f}', f'{jb_stat}'],
                        'p-value': [f'{bp_pvalue:.4f}', 'N/A', f'{jb_pvalue}'],
                        'Interpretation': [
                            'Homoscedasticity' if bp_pvalue >= 0.05 else 'Heteroscedasticity detected',
                            'No autocorrelation' if 1.5 <= dw_stat <= 2.5 else 'Autocorrelation detected',
                            'Residuals are normal' if jb_pvalue != 'N/A' and float(jb_pvalue) >= 0.05 else 'Non-normal residuals'
                        ]
                    }
                    diagnostics_df = pd.DataFrame(diagnostics_data)
                    diagnostics_df.to_excel(writer, sheet_name='Diagnostics', index=False)
                except:
                    # Basic diagnostics
                    basic_diagnostics = pd.DataFrame({
                        'Metric': ['Mean Residual', 'Std Residual', 'Min Residual', 'Max Residual'],
                        'Value': [np.mean(model.resid), np.std(model.resid), np.min(model.resid), np.max(model.resid)]
                    })
                    basic_diagnostics.to_excel(writer, sheet_name='Diagnostics', index=False)
            
            # Sheet 4: Residuals Data
            if "Residuals" in excel_sheets:
                residuals_data = pd.DataFrame({
                    'Observation': range(1, len(model.resid) + 1),
                    'Fitted_Values': model.fittedvalues,
                    'Residuals': model.resid,
                    'Standardized_Residuals': model.resid / np.std(model.resid),
                    'Actual_Values': model.fittedvalues + model.resid
                })
                residuals_data.to_excel(writer, sheet_name='Residuals', index=False)
            
            # Sheet 5: Data Sample
            if "Data Sample" in excel_sheets:
                data_sample = st.session_state.data[[dependent_var] + independent_vars].head(1000)
                data_sample.to_excel(writer, sheet_name='Data_Sample', index=False)
        
        # Style the workbook
        workbook = openpyxl.load_workbook(output)
        
        # Style summary sheet
        if "Summary" in excel_sheets and 'Model_Summary' in workbook.sheetnames:
            ws = workbook['Model_Summary']
            
            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        
        # Save styled workbook
        output = io.BytesIO()
        workbook.save(output)
        
        st.success("✅ Excel workbook generated successfully!")
        
        st.download_button(
            label="📥 **Download Excel Workbook**",
            data=output.getvalue(),
            file_name=f"econometrics_analysis_{dependent_var}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_workbook",
            help="Download comprehensive Excel workbook with multiple sheets"
        )
        
        # Show what's included
        st.info(f"📊 **Excel Workbook includes**: {', '.join(excel_sheets)}")
        
    except Exception as e:
        st.error(f"❌ Error generating Excel workbook: {str(e)}")
        st.info("💡 **Tip**: Make sure openpyxl is installed for full Excel functionality")

def generate_latex_table(model, independent_vars, dependent_var, latex_style, include_stars):
    """Generate professional LaTeX table"""
    
    try:
        # LaTeX table header based on style
        if latex_style == "Booktabs":
            table_start = "\\begin{table}[htbp]\n\\centering\n\\caption{Regression Results}\n\\begin{tabular}{lcccc}\n\\toprule"
            table_end = "\\bottomrule\n\\end{tabular}\n\\label{tab:regression}\n\\end{table}"
            line_sep = "\\midrule"
        elif latex_style == "Professional":
            table_start = "\\begin{table}[H]\n\\centering\n\\caption{Regression Results for " + dependent_var.replace('_', ' ').title() + "}\n\\begin{tabular}{@{}lcccc@{}}\n\\toprule"
            table_end = "\\bottomrule\n\\end{tabular}\n\\label{tab:reg_" + dependent_var.lower() + "}\n\\end{table}"
            line_sep = "\\midrule"
        else:
            table_start = "\\begin{table}[htbp]\n\\centering\n\\caption{Regression Results}\n\\begin{tabular}{|l|c|c|c|c|}\n\\hline"
            table_end = "\\hline\n\\end{tabular}\n\\label{tab:regression}\n\\end{table}"
            line_sep = "\\hline"
        
        latex_code = f"""% Regression Results Table
% Generated by Econometrics Learning Lab - Created by HAMDI Boulanouar
% Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

{table_start}
Variable & Coefficient & Std. Error & t-statistic & p-value \\\\
{line_sep}
"""
        
        # Handle parameter access
        try:
            if hasattr(model.params, 'index'):
                param_items = list(zip(
                    model.params.index, model.params.values,
                    model.bse.values, model.tvalues.values, model.pvalues.values
                ))
            else:
                param_names = ['Intercept'] + independent_vars
                param_values = np.array(model.params).flatten()
                std_errors = np.array(model.bse).flatten()
                t_values = np.array(model.tvalues).flatten()
                p_values = np.array(model.pvalues).flatten()
                
                param_items = list(zip(param_names, param_values, std_errors, t_values, p_values))
        except:
            param_items = [('Error', 0, 0, 0, 1)]
        
        # Add coefficient rows
        for name, coef, se, t_val, p_val in param_items:
            # Add significance stars if requested
            if include_stars:
                if p_val < 0.01:
                    stars = "***"
                elif p_val < 0.05:
                    stars = "**"
                elif p_val < 0.10:
                    stars = "*"
                else:
                    stars = ""
            else:
                stars = ""
            
            # Clean variable name for LaTeX
            clean_name = name.replace('_', '\\_').replace('%', '\\%').replace('&', '\\&')
            
            latex_code += f"{clean_name} & {coef:.4f}{stars} & ({se:.4f}) & {t_val:.3f} & {p_val:.4f} \\\\\n"
        
        # Add model statistics
        latex_code += f"""{line_sep}
\\multicolumn{{5}}{{l}}{{\\textbf{{Model Statistics:}}}} \\\\
R-squared & \\multicolumn{{4}}{{l}}{{{model.rsquared:.4f}}} \\\\
Adjusted R-squared & \\multicolumn{{4}}{{l}}{{{model.rsquared_adj:.4f}}} \\\\
F-statistic & \\multicolumn{{4}}{{l}}{{{model.fvalue:.2f} (p = {model.f_pvalue:.4f})}} \\\\
Observations & \\multicolumn{{4}}{{l}}{{{int(model.nobs)}}} \\\\
{table_end}
"""
        
        # Add notes
        if include_stars:
            latex_code += "\n% Notes:\n% Standard errors in parentheses\n% *** p<0.01, ** p<0.05, * p<0.1\n"
        
        # Add required packages
        packages = """
% Required LaTeX packages:
% \\usepackage{booktabs}  % For professional tables (if using booktabs style)
% \\usepackage{float}     % For H placement specifier (if using professional style)
% \\usepackage{amsmath}   % For mathematical symbols
"""
        
        full_latex = packages + "\n" + latex_code
        
        st.success("✅ LaTeX table generated successfully!")
        
        # Show code preview
        st.code(latex_code, language='latex')
        
        st.download_button(
            label="📥 **Download LaTeX Code**",
            data=full_latex,
            file_name=f"regression_table_{dependent_var}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.tex",
            mime="text/plain",
            key="download_latex_table",
            help="Download publication-ready LaTeX table"
        )
        
        # LaTeX tips
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">📝 LaTeX Usage Tips:</h5>
        <div class="red-list">
        • Add <code>\\usepackage{booktabs}</code> for professional table style<br>
        • Use <code>\\usepackage{float}</code> for H placement specifier<br>
        • Compile with pdflatex or similar LaTeX engine<br>
        • Adjust table width with <code>\\resizebox</code> if needed<br>
        • Standard errors are shown in parentheses below coefficients
        </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"❌ Error generating LaTeX table: {str(e)}")

def generate_csv_export(model, independent_vars, dependent_var, include_diagnostics):
    """Generate CSV export with multiple files"""
    
    try:
        # 1. Coefficients CSV
        coeffs_data = get_coefficients_data()
        coeffs_df = pd.DataFrame(coeffs_data)
        
        st.subheader("📊 CSV Exports Generated")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**📈 Coefficients Data:**")
            st.dataframe(coeffs_df.head(), use_container_width=True)
            
            csv_coeffs = coeffs_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Coefficients CSV",
                data=csv_coeffs,
                file_name=f"coefficients_{dependent_var}.csv",
                mime="text/csv",
                key="download_csv_coeffs"
            )
        
        # 2. Model Statistics CSV
        model_stats = get_model_statistics()
        stats_df = pd.DataFrame([model_stats])
        
        with col2:
            st.markdown("**📊 Model Statistics:**")
            st.dataframe(stats_df.head(), use_container_width=True)
            
            csv_stats = stats_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Model Stats CSV",
                data=csv_stats,
                file_name=f"model_statistics_{dependent_var}.csv",
                mime="text/csv",
                key="download_csv_stats"
            )
        
        # 3. Residuals and Fitted Values CSV
        residuals_data = pd.DataFrame({
            'Observation': range(1, len(model.resid) + 1),
            'Actual_Values': model.fittedvalues + model.resid,
            'Fitted_Values': model.fittedvalues,
            'Residuals': model.resid,
            'Standardized_Residuals': model.resid / np.std(model.resid),
            'Squared_Residuals': model.resid ** 2
        })
        
        st.markdown("**📊 Residuals Analysis:**")
        st.dataframe(residuals_data.head(10), use_container_width=True)
        
        csv_residuals = residuals_data.to_csv(index=False)
        st.download_button(
            label="📥 Download Residuals CSV",
            data=csv_residuals,
            file_name=f"residuals_{dependent_var}.csv",
            mime="text/csv",
            key="download_csv_residuals"
        )
        
        # 4. Combined Summary CSV
        summary_data = {
            'Analysis_Info': [
                'Date', 'Created_By', 'Dataset', 'Model_Type', 'Dependent_Variable', 
                'Independent_Variables', 'Sample_Size', 'R_squared', 'Adj_R_squared',
                'F_statistic', 'F_pvalue', 'AIC', 'BIC', 'RMSE'
            ],
            'Values': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'HAMDI Boulanouar',
                st.session_state.uploaded_file_name,
                'Simple' if len(independent_vars) == 1 else 'Multiple',
                dependent_var,
                ', '.join(independent_vars),
                int(model.nobs),
                model.rsquared,
                model.rsquared_adj,
                model.fvalue,
                model.f_pvalue,
                model.aic,
                model.bic,
                np.sqrt(model.mse_resid)
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        csv_summary = summary_df.to_csv(index=False)
        
        st.download_button(
            label="📥 **Download Complete Summary CSV**",
            data=csv_summary,
            file_name=f"complete_analysis_summary_{dependent_var}.csv",
            mime="text/csv",
            key="download_csv_complete"
        )
        
        st.success("✅ All CSV files generated successfully!")
        
    except Exception as e:
        st.error(f"❌ Error generating CSV exports: {str(e)}")

def generate_python_code(model, independent_vars, dependent_var, include_diagnostics):
    """Generate reproducible Python code"""
    
    try:
        python_code = f'''
# Econometrics Analysis - Reproducible Python Code
# Generated by: Econometrics Learning Lab - Created by HAMDI Boulanouar
# Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# Analysis of: {dependent_var} using {independent_vars}

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import statsmodels.api as sm
from statsmodels.stats.diagnostic import het_breuschpagan, het_white
from statsmodels.stats.stattools import durbin_watson, jarque_bera
from statsmodels.stats.outliers_influence import variance_inflation_factor
from scipy import stats
import warnings
warnings.filterwarnings('ignore')

# ========================
# DATA LOADING
# ========================
# Load your data (replace with your actual data loading code)
# df = pd.read_csv('your_data.csv')
# df = pd.read_excel('your_data.xlsx')

# For demonstration, using the current analysis variables:
# dependent_var = '{dependent_var}'
# independent_vars = {independent_vars}

print("="*60)
print("ECONOMETRICS ANALYSIS REPRODUCTION")
print("Created by HAMDI Boulanouar - Econometrics Learning Lab")
print("="*60)

# ========================
# VARIABLE DEFINITION
# ========================
# Define your variables
dependent_var = '{dependent_var}'
independent_vars = {independent_vars}

# Extract variables from dataframe
y = df[dependent_var].values
X = df[independent_vars].values

# Add constant for intercept
X_with_const = sm.add_constant(X)

print(f"Analysis: {{dependent_var}} ~ {{' + '.join(independent_vars)}}")
print(f"Sample size: {{len(y)}} observations")
print(f"Variables: {{len(independent_vars)}} independent variables")

# ========================
# REGRESSION ANALYSIS
# ========================
print("\\n" + "="*60)
print("RUNNING OLS REGRESSION")
print("="*60)

# Fit OLS model
model = sm.OLS(y, X_with_const).fit()

# Display full results
print(model.summary())

# ========================
# KEY RESULTS EXTRACTION
# ========================
print("\\n" + "="*60)
print("KEY STATISTICAL MEASURES")
print("="*60)

print(f"R-squared: {{model.rsquared:.6f}}")
print(f"Adjusted R-squared: {{model.rsquared_adj:.6f}}")
print(f"F-statistic: {{model.fvalue:.4f}} (p-value: {{model.f_pvalue:.6f}})")
print(f"AIC: {{model.aic:.2f}}")
print(f"BIC: {{model.bic:.2f}}")
print(f"Root Mean Square Error: {{np.sqrt(model.mse_resid):.4f}}")
print(f"Log-Likelihood: {{model.llf:.2f}}")

# Model significance
if model.f_pvalue < 0.001:
    print("\\n🎉 MODEL IS HIGHLY SIGNIFICANT (p < 0.001)")
elif model.f_pvalue < 0.05:
    print("\\n✅ MODEL IS SIGNIFICANT (p < 0.05)")
else:
    print("\\n⚠️ MODEL IS NOT SIGNIFICANT (p >= 0.05)")

# ========================
# COEFFICIENT ANALYSIS
# ========================
print("\\n" + "="*60)
print("COEFFICIENT INTERPRETATION")
print("="*60)

# Extract coefficients with proper handling
try:
    if hasattr(model.params, 'index'):
        param_names = list(model.params.index)
        param_values = list(model.params.values)
        std_errors = list(model.bse.values)
        t_values = list(model.tvalues.values)
        p_values = list(model.pvalues.values)
    else:
        param_names = ['Intercept'] + independent_vars
        param_values = list(np.array(model.params).flatten())
        std_errors = list(np.array(model.bse).flatten())
        t_values = list(np.array(model.tvalues).flatten())
        p_values = list(np.array(model.pvalues).flatten())
        
    # Interpret each coefficient
    for i, (name, coef, se, t_val, p_val) in enumerate(zip(param_names, param_values, std_errors, t_values, p_values)):
        significance = "***" if p_val < 0.01 else "**" if p_val < 0.05 else "*" if p_val < 0.10 else ""
        sig_level = "highly significant" if p_val < 0.01 else ("significant" if p_val < 0.05 else ("marginally significant" if p_val < 0.10 else "not significant"))
        
        print(f"\\n{{name}}: {{coef:.6f}} {{significance}}")
        print(f"  Standard Error: {{se:.6f}}")
        print(f"  t-statistic: {{t_val:.4f}}")
        print(f"  p-value: {{p_val:.6f}} ({{sig_level}})")
        
        if name == 'Intercept':
            print(f"  Interpretation: When all X variables = 0, {{dependent_var}} = {{coef:.4f}}")
        else:
            direction = "increases" if coef > 0 else "decreases"
            print(f"  Interpretation: 1-unit increase in {{name}} {{direction}} {{dependent_var}} by {{abs(coef):.4f}}")
            if len(independent_vars) > 1:
                print("                  (holding other variables constant)")

except Exception as e:
    print(f"Error in coefficient extraction: {{e}}")
'''

        if include_diagnostics:
            python_code += f'''
# ========================
# DIAGNOSTIC TESTS
# ========================
print("\\n" + "="*60)
print("REGRESSION DIAGNOSTICS")
print("="*60)

# 1. Heteroscedasticity test (Breusch-Pagan)
try:
    bp_stat, bp_pvalue, bp_f, bp_f_pvalue = het_breuschpagan(model.resid, X_with_const)
    print(f"\\nBreusch-Pagan Test for Heteroscedasticity:")
    print(f"  LM Statistic: {{bp_stat:.4f}}")
    print(f"  p-value: {{bp_pvalue:.4f}}")
    print(f"  Conclusion: {{'Homoscedasticity' if bp_pvalue >= 0.05 else 'Heteroscedasticity detected'}}")
except Exception as e:
    print(f"Breusch-Pagan test failed: {{e}}")

# 2. Autocorrelation test (Durbin-Watson)
dw_stat = durbin_watson(model.resid)
print(f"\\nDurbin-Watson Test for Autocorrelation:")
print(f"  DW Statistic: {{dw_stat:.4f}}")
if 1.5 <= dw_stat <= 2.5:
    print("  Conclusion: No significant autocorrelation")
elif dw_stat < 1.5:
    print("  Conclusion: Positive autocorrelation detected")
else:
    print("  Conclusion: Negative autocorrelation detected")

# 3. Normality test (Jarque-Bera)
try:
    jb_result = jarque_bera(model.resid)
    jb_stat, jb_pvalue = jb_result[0], jb_result[1]
    print(f"\\nJarque-Bera Test for Normality:")
    print(f"  JB Statistic: {{jb_stat:.4f}}")
    print(f"  p-value: {{jb_pvalue:.4f}}")
    print(f"  Conclusion: {{'Residuals appear normal' if jb_pvalue >= 0.05 else 'Non-normal residuals detected'}}")
except Exception as e:
    print(f"Jarque-Bera test failed: {{e}}")

# 4. Multicollinearity test (VIF) - only if multiple variables
if len(independent_vars) > 1:
    print(f"\\nVariance Inflation Factor (VIF) Analysis:")
    try:
        for i, var_name in enumerate(independent_vars):
            vif = variance_inflation_factor(X, i)
            vif_interpretation = "Low" if vif < 5 else ("Moderate" if vif < 10 else "High")
            print(f"  {{var_name}}: {{vif:.3f}} ({{vif_interpretation}} multicollinearity)")
    except Exception as e:
        print(f"VIF calculation failed: {{e}}")

# 5. Residual Statistics
print(f"\\nResidual Analysis:")
print(f"  Mean of residuals: {{np.mean(model.resid):.8f}} (should be ~0)")
print(f"  Std deviation of residuals: {{np.std(model.resid):.4f}}")
print(f"  Minimum residual: {{np.min(model.resid):.4f}}")
print(f"  Maximum residual: {{np.max(model.resid):.4f}}")
print(f"  Residuals range: {{np.max(model.resid) - np.min(model.resid):.4f}}")
'''

        python_code += f'''
# ========================
# VISUALIZATION
# ========================
print("\\n" + "="*60)
print("CREATING DIAGNOSTIC PLOTS")
print("="*60)

# Create comprehensive diagnostic plots
fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
fig.suptitle('Regression Diagnostic Plots\\nCreated by HAMDI Boulanouar - Econometrics Learning Lab', 
             fontsize=14, fontweight='bold')

# 1. Residuals vs Fitted Values
ax1.scatter(model.fittedvalues, model.resid, alpha=0.6, color='blue')
ax1.axhline(y=0, color='red', linestyle='--', linewidth=2)
ax1.set_xlabel('Fitted Values', fontweight='bold')
ax1.set_ylabel('Residuals', fontweight='bold')
ax1.set_title('Residuals vs Fitted Values\\n(Check for Heteroscedasticity)', fontweight='bold')
ax1.grid(True, alpha=0.3)

# Add trend line
z = np.polyfit(model.fittedvalues, model.resid, 1)
p = np.poly1d(z)
ax1.plot(model.fittedvalues, p(model.fittedvalues), "r--", alpha=0.8, linewidth=2)

# 2. Q-Q Plot for Normality
stats.probplot(model.resid, dist="norm", plot=ax2)
ax2.set_title('Q-Q Plot (Normal Distribution)\\n(Check for Normality)', fontweight='bold')
ax2.grid(True, alpha=0.3)

# 3. Histogram of Residuals with Normal Curve
ax3.hist(model.resid, bins=20, density=True, alpha=0.7, color='skyblue', edgecolor='black')
ax3.set_xlabel('Residuals', fontweight='bold')
ax3.set_ylabel('Density', fontweight='bold')
ax3.set_title('Distribution of Residuals\\n(Should be Normal)', fontweight='bold')

# Add normal distribution curve
x_norm = np.linspace(model.resid.min(), model.resid.max(), 100)
y_norm = stats.norm.pdf(x_norm, loc=np.mean(model.resid), scale=np.std(model.resid))
ax3.plot(x_norm, y_norm, 'r-', linewidth=2, label='Normal Distribution')
ax3.legend()
ax3.grid(True, alpha=0.3)

# 4. Scale-Location Plot (Square root of standardized residuals)
standardized_resid = model.resid / np.std(model.resid)
ax4.scatter(model.fittedvalues, np.sqrt(np.abs(standardized_resid)), alpha=0.6, color='green')
ax4.set_xlabel('Fitted Values', fontweight='bold')
ax4.set_ylabel('√|Standardized Residuals|', fontweight='bold')
ax4.set_title('Scale-Location Plot\\n(Check for Homoscedasticity)', fontweight='bold')
ax4.grid(True, alpha=0.3)

# Add trend line
z2 = np.polyfit(model.fittedvalues, np.sqrt(np.abs(standardized_resid)), 1)
p2 = np.poly1d(z2)
ax4.plot(model.fittedvalues, p2(model.fittedvalues), "r--", alpha=0.8, linewidth=2)

plt.tight_layout()
plt.show()

# ========================
# PREDICTIONS EXAMPLE
# ========================
print("\\n" + "="*60)
print("MAKING PREDICTIONS")
print("="*60)

# Example prediction (modify values as needed)
print("Example: Making a prediction with mean values")

# Calculate mean values for prediction
mean_values = [1]  # Start with intercept
for var in independent_vars:
    mean_val = df[var].mean()
    mean_values.append(mean_val)
    print(f"Using mean value for {{var}}: {{mean_val:.4f}}")

# Make prediction
prediction = model.predict([mean_values])[0]
print(f"\\n🎯 PREDICTION: {{dependent_var}} = {{prediction:.4f}}")

# Calculate prediction interval (95% confidence)
pred_se = np.sqrt(model.mse_resid * (1 + np.sum(np.array(mean_values)**2) / len(model.resid)))
t_crit = stats.t.ppf(0.975, df=model.df_resid)
pred_interval = [prediction - t_crit * pred_se, prediction + t_crit * pred_se]

print(f"95% Prediction Interval: [{{pred_interval[0]:.4f}}, {{pred_interval[1]:.4f}}]")
print(f"Interval Width: {{pred_interval[1] - pred_interval[0]:.4f}}")

# ========================
# SUMMARY AND CONCLUSIONS
# ========================
print("\\n" + "="*60)
print("ANALYSIS SUMMARY & CONCLUSIONS")
print("="*60)

print(f"\\n📊 MODEL PERFORMANCE:")
print(f"   • R-squared: {{model.rsquared:.4f}} ({{model.rsquared*100:.1f}}% of variation explained)")
print(f"   • Adjusted R-squared: {{model.rsquared_adj:.4f}}")
print(f"   • Model is {{'SIGNIFICANT' if model.f_pvalue < 0.05 else 'NOT SIGNIFICANT'}} (F p-value: {{model.f_pvalue:.4f}})")

print(f"\\n🎯 KEY RELATIONSHIPS:")
for i, (name, coef, p_val) in enumerate(zip(param_names, param_values, p_values)):
    if name != 'Intercept' and p_val < 0.10:
        direction = "POSITIVE" if coef > 0 else "NEGATIVE"
        strength = "STRONG" if p_val < 0.01 else ("MODERATE" if p_val < 0.05 else "WEAK")
        print(f"   • {{name}} has {{strength}} {{direction}} relationship with {{dependent_var}}")

print(f"\\n⚠️ DIAGNOSTIC SUMMARY:")
print(f"   • Autocorrelation: {{'OK' if 1.5 <= dw_stat <= 2.5 else 'CHECK REQUIRED'}}")
print(f"   • Homoscedasticity: {{'OK' if bp_pvalue >= 0.05 else 'CHECK REQUIRED'}}")
print(f"   • Normality: {{'OK' if jb_pvalue >= 0.05 else 'CHECK REQUIRED'}}")

print(f"\\n💡 RECOMMENDATIONS:")
if model.rsquared > 0.7:
    print("   ✅ Strong model - suitable for prediction and inference")
elif model.rsquared > 0.5:
    print("   ✅ Good model - reliable for most purposes")
else:
    print("   ⚠️ Weak model - consider additional variables or different approach")

if model.f_pvalue < 0.05:
    print("   ✅ Model is statistically significant - relationships are reliable")
else:
    print("   ⚠️ Model lacks statistical significance - interpret with caution")

print("\\n" + "="*60)
print("ANALYSIS COMPLETED SUCCESSFULLY")
print("Generated by Econometrics Learning Lab - Created by HAMDI Boulanouar")
print("For more advanced features, use the full Streamlit application")
print("="*60)

# Optional: Save results to files
# model.save('regression_model.pickle')  # Save model
# pd.DataFrame({{'fitted': model.fittedvalues, 'residuals': model.resid}}).to_csv('residuals.csv')
# print("\\n💾 Results saved to files")
'''

        st.success("✅ Python code generated successfully!")
        
        # Show code preview
        with st.expander("🐍 **Python Code Preview**", expanded=False):
            st.code(python_code[:2000] + "\n\n# ... (code continues) ..." if len(python_code) > 2000 else python_code, language='python')
        
        st.download_button(
            label="📥 **Download Complete Python Code**",
            data=python_code,
            file_name=f"econometrics_analysis_{dependent_var}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.py",
            mime="text/plain",
            key="download_python_code",
            help="Download fully reproducible Python analysis code"
        )
        
        # Code usage tips
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">🐍 Python Code Usage:</h5>
        <div class="red-list">
        • Replace data loading section with your actual data source<br>
        • Modify prediction example values as needed<br>
        • Run in Jupyter notebook or Python environment<br>
        • All required packages are imported at the top<br>
        • Code includes comprehensive comments and explanations
        </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"❌ Error generating Python code: {str(e)}")

def generate_powerpoint_summary(model, independent_vars, dependent_var):
    """Generate PowerPoint summary template"""
    
    st.info("📊 **PowerPoint Export**: This feature generates a structured text template that you can copy into PowerPoint slides.")
    
    try:
        # Create PowerPoint content structure
        ppt_content = f"""
ECONOMETRICS ANALYSIS PRESENTATION
Generated by HAMDI Boulanouar - Econometrics Learning Lab
Date: {datetime.now().strftime('%Y-%m-%d')}

==========================================
SLIDE 1: TITLE SLIDE
==========================================
Title: Regression Analysis Results
Subtitle: Analysis of {dependent_var}
Author: [Your Name]
Date: {datetime.now().strftime('%B %d, %Y')}
Created with: Econometrics Learning Lab by HAMDI Boulanouar

==========================================
SLIDE 2: RESEARCH QUESTION
==========================================
Title: Research Question

• What factors influence {dependent_var.replace('_', ' ').title()}?
• How strong are these relationships?
• Can we predict {dependent_var.replace('_', ' ').title()} accurately?

Variables Analyzed:
• Dependent Variable: {dependent_var.replace('_', ' ').title()}
• Independent Variables: {', '.join([var.replace('_', ' ').title() for var in independent_vars])}
• Sample Size: {int(model.nobs)} observations

==========================================
SLIDE 3: METHODOLOGY
==========================================
Title: Statistical Methodology

Analysis Approach:
• Method: {'Simple' if len(independent_vars) == 1 else 'Multiple'} Linear Regression
• Estimation: Ordinary Least Squares (OLS)
• Software: Python with statsmodels
• Significance Level: 95% confidence

Model Specification:
{dependent_var} = β₀ + β₁{independent_vars[0].replace('_', ' ')}"""

        if len(independent_vars) > 1:
            for i, var in enumerate(independent_vars[1:], 2):
                ppt_content += f" + β{i}{var.replace('_', ' ')}"
        
        ppt_content += " + ε"

        # Extract key results
        try:
            if hasattr(model.params, 'index'):
                param_names = list(model.params.index)
                param_values = list(model.params.values)
                p_values = list(model.pvalues.values)
            else:
                param_names = ['Intercept'] + independent_vars
                param_values = list(np.array(model.params).flatten())
                p_values = list(np.array(model.pvalues).flatten())
        except:
            param_names = ['Intercept'] + independent_vars
            param_values = [0] * len(param_names)
            p_values = [1] * len(param_names)

        ppt_content += f"""

==========================================
SLIDE 4: KEY FINDINGS
==========================================
Title: Main Results

Model Performance:
• R-squared: {model.rsquared:.3f} ({model.rsquared*100:.1f}% of variation explained)
• F-statistic: {model.fvalue:.2f} (p = {model.f_pvalue:.4f})
• Model is {'STATISTICALLY SIGNIFICANT' if model.f_pvalue < 0.05 else 'NOT STATISTICALLY SIGNIFICANT'}

Key Relationships:"""

        for i, (name, coef, pval) in enumerate(zip(param_names, param_values, p_values)):
            if name != 'Intercept':
                direction = "increases" if coef > 0 else "decreases"
                significance = "significant" if pval < 0.05 else "not significant"
                ppt_content += f"\n• {name.replace('_', ' ').title()}: {direction} {dependent_var.replace('_', ' ')} by {abs(coef):.4f} per unit ({significance})"

        ppt_content += f"""

==========================================
SLIDE 5: DETAILED COEFFICIENTS
==========================================
Title: Regression Coefficients

[TABLE FORMAT - Copy to PowerPoint Table]
Variable | Coefficient | Std. Error | t-stat | p-value | Significance
---------|-------------|------------|---------|---------|-------------"""

        for name, coef, pval in zip(param_names, param_values, p_values):
            try:
                se = list(model.bse.values)[param_names.index(name)] if hasattr(model.bse, 'values') else 0
                tval = list(model.tvalues.values)[param_names.index(name)] if hasattr(model.tvalues, 'values') else 0
            except:
                se, tval = 0, 0
            
            stars = "***" if pval < 0.01 else "**" if pval < 0.05 else "*" if pval < 0.10 else ""
            ppt_content += f"\n{name.replace('_', ' ').title()} | {coef:.4f} | {se:.4f} | {tval:.3f} | {pval:.4f} | {stars}"

        ppt_content += f"""

Note: *** p<0.01, ** p<0.05, * p<0.10

==========================================
SLIDE 6: MODEL DIAGNOSTICS
==========================================
Title: Model Quality Assessment

Statistical Tests:
• Durbin-Watson: {durbin_watson(model.resid):.3f} ({'No autocorrelation' if 1.5 <= durbin_watson(model.resid) <= 2.5 else 'Check autocorrelation'})
• Residual Analysis: {'Acceptable' if abs(np.mean(model.resid)) < 0.001 else 'Check residuals'}

Model Fit Quality:
• {'EXCELLENT' if model.rsquared > 0.8 else ('GOOD' if model.rsquared > 0.6 else ('FAIR' if model.rsquared > 0.4 else 'POOR'))} model fit (R² = {model.rsquared:.3f})
• Prediction accuracy: ±{np.sqrt(model.mse_resid):.3f} units (RMSE)

==========================================
SLIDE 7: PRACTICAL IMPLICATIONS
==========================================
Title: Business/Research Implications

Key Insights:
• {dependent_var.replace('_', ' ').title()} can be {'reliably' if model.rsquared > 0.6 else 'partially'} predicted using the selected variables
• Most important factors: [List top 2-3 significant variables]
• {'Strong statistical evidence' if model.f_pvalue < 0.01 else 'Moderate statistical evidence'} for the relationships found

Actionable Recommendations:
• [Add your domain-specific interpretations]
• [Add business/policy implications]
• [Add suggestions for future research]

==========================================
SLIDE 8: LIMITATIONS & FUTURE WORK
==========================================
Title: Study Limitations

Current Limitations:
• Sample size: {int(model.nobs)} observations
• Time period: [Specify your data period]
• Geographic scope: [Specify if applicable]
• Causality: Results show association, not causation

Future Research:
• Consider additional variables
• Test with larger/different samples
• Apply advanced modeling techniques
• Conduct longitudinal analysis

==========================================
SLIDE 9: CONCLUSIONS
==========================================
Title: Summary & Conclusions

Main Conclusions:
1. The model {'successfully' if model.f_pvalue < 0.05 else 'partially'} explains variation in {dependent_var.replace('_', ' ')}
2. {len([p for p in p_values[1:] if p < 0.05])} out of {len(independent_vars)} variables are statistically significant
3. Model explains {model.rsquared*100:.1f}% of the variation in the outcome

Statistical Reliability: {'HIGH' if model.f_pvalue < 0.01 and model.rsquared > 0.6 else 'MODERATE' if model.f_pvalue < 0.05 else 'LOW'}

==========================================
SLIDE 10: THANK YOU
==========================================
Title: Questions & Discussion

Thank You!

Contact Information:
• Analysis created with: Econometrics Learning Lab
• Developed by: HAMDI Boulanouar
• Professional Statistical Analysis Platform

Questions?
[Your contact information]

Generated: {datetime.now().strftime('%B %d, %Y')}
"""

        st.success("✅ PowerPoint template generated successfully!")
        
        # Show preview
        with st.expander("📊 **PowerPoint Template Preview**", expanded=False):
            st.text(ppt_content[:3000] + "\n\n[... template continues ...]" if len(ppt_content) > 3000 else ppt_content)
        
        st.download_button(
            label="📥 **Download PowerPoint Template**",
            data=ppt_content,
            file_name=f"powerpoint_template_{dependent_var}_{datetime.now().strftime('%Y%m%d')}.txt",
            mime="text/plain",
            key="download_ppt_template"
        )
        
        # PowerPoint tips
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">📊 PowerPoint Usage Tips:</h5>
        <div class="red-list">
        • Copy each slide section into separate PowerPoint slides<br>
        • Add charts from the main app's visualization section<br>
        • Customize content for your specific audience<br>
        • Use professional PowerPoint templates for better visuals<br>
        • Add company/institution branding as needed
        </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"❌ Error generating PowerPoint template: {str(e)}")

def generate_word_document(model, independent_vars, dependent_var, include_diagnostics):
    """Generate Word document template"""
    
    st.info("📋 **Word Document Export**: This feature generates formatted text that can be copied into Microsoft Word.")
    
    try:
        # Create Word document content
        word_content = f"""
ECONOMETRICS ANALYSIS REPORT

Title: Statistical Analysis of {dependent_var.replace('_', ' ').title()}
Author: [Your Name]
Date: {datetime.now().strftime('%B %d, %Y')}
Generated by: Econometrics Learning Lab - Created by HAMDI Boulanouar

EXECUTIVE SUMMARY

This report presents the results of a {'simple' if len(independent_vars) == 1 else 'multiple'} linear regression analysis examining the relationship between {dependent_var.replace('_', ' ')} and {len(independent_vars)} independent variable{'s' if len(independent_vars) > 1 else ''}. The analysis was conducted using a sample of {int(model.nobs)} observations.

Key findings include:
• The model explains {model.rsquared*100:.1f}% of the variation in {dependent_var.replace('_', ' ')}
• The overall model is {'statistically significant' if model.f_pvalue < 0.05 else 'not statistically significant'} (F = {model.fvalue:.2f}, p = {model.f_pvalue:.4f})
• {len([p for p in [list(model.pvalues.values) if hasattr(model.pvalues, 'values') else [1]][0] if p < 0.05])} out of {len(independent_vars)} independent variables show significant relationships

1. INTRODUCTION

1.1 Research Objective
The primary objective of this analysis is to understand the factors that influence {dependent_var.replace('_', ' ')} and to quantify these relationships using statistical methods.

1.2 Variables Under Study
Dependent Variable: {dependent_var.replace('_', ' ').title()}
Independent Variables: {', '.join([var.replace('_', ' ').title() for var in independent_vars])}

1.3 Data Description
The analysis is based on {int(model.nobs)} complete observations. [Add description of your data source, collection method, and time period]

2. METHODOLOGY

2.1 Statistical Approach
This study employs Ordinary Least Squares (OLS) regression to estimate the following model:

{dependent_var} = β₀ + β₁({independent_vars[0].replace('_', ' ')})"""

        if len(independent_vars) > 1:
            for i, var in enumerate(independent_vars[1:], 2):
                word_content += f" + β{i}({var.replace('_', ' ')})"
        
        word_content += " + ε"

        word_content += f"""

Where:
• {dependent_var} represents the dependent variable
• β₀ is the intercept term
"""
        
        for i, var in enumerate(independent_vars, 1):
            word_content += f"• β{i} is the coefficient for {var.replace('_', ' ')}\n"
        
        word_content += "• ε is the error term\n"

        word_content += f"""

2.2 Statistical Software
The analysis was conducted using Python with the statsmodels library, providing robust statistical estimation and diagnostic capabilities.

3. RESULTS

3.1 Model Summary
The regression analysis yields the following overall model statistics:

R-squared: {model.rsquared:.4f}
Adjusted R-squared: {model.rsquared_adj:.4f}
F-statistic: {model.fvalue:.4f}
F p-value: {model.f_pvalue:.6f}
Number of observations: {int(model.nobs)}
Degrees of freedom: {int(model.df_resid)}

3.2 Model Significance
{'The F-test indicates that the overall model is statistically significant at the 95% confidence level, suggesting that the independent variables collectively have a significant relationship with the dependent variable.' if model.f_pvalue < 0.05 else 'The F-test indicates that the overall model is not statistically significant, suggesting weak evidence for the collective relationship between independent and dependent variables.'}

3.3 Coefficient Estimates

[TABLE 1: Regression Coefficients]
Variable | Coefficient | Std. Error | t-statistic | p-value | 95% Confidence Interval
"""

        # Handle parameter access for Word document
        try:
            if hasattr(model.params, 'index'):
                param_names = list(model.params.index)
                param_values = list(model.params.values)
                std_errors = list(model.bse.values)
                t_values = list(model.tvalues.values)
                p_values = list(model.pvalues.values)
                ci_lower = list(model.conf_int()[0].values)
                ci_upper = list(model.conf_int()[1].values)
            else:
                param_names = ['Intercept'] + independent_vars
                param_values = list(np.array(model.params).flatten())
                std_errors = list(np.array(model.bse).flatten())
                t_values = list(np.array(model.tvalues).flatten())
                p_values = list(np.array(model.pvalues).flatten())
                ci_lower = list(np.array(model.conf_int()[0]).flatten())
                ci_upper = list(np.array(model.conf_int()[1]).flatten())
        except:
            param_names = ['Intercept'] + independent_vars
            param_values = [0] * len(param_names)
            std_errors = [0] * len(param_names)
            t_values = [0] * len(param_names)
            p_values = [1] * len(param_names)
            ci_lower = [0] * len(param_names)
            ci_upper = [0] * len(param_names)

        for i, (name, coef, se, t_val, p_val, ci_l, ci_u) in enumerate(zip(param_names, param_values, std_errors, t_values, p_values, ci_lower, ci_upper)):
            word_content += f"{name.replace('_', ' ').title()} | {coef:.4f} | {se:.4f} | {t_val:.3f} | {p_val:.4f} | [{ci_l:.4f}, {ci_u:.4f}]\n"

        word_content += f"""

3.4 Interpretation of Coefficients

"""
        
        for name, coef, p_val in zip(param_names, param_values, p_values):
            if name == 'Intercept':
                word_content += f"Intercept (β₀ = {coef:.4f}): This represents the expected value of {dependent_var.replace('_', ' ')} when all independent variables equal zero. "
            else:
                direction = "increase" if coef > 0 else "decrease"
                word_content += f"{name.replace('_', ' ').title()} (β = {coef:.4f}): A one-unit increase in {name.replace('_', ' ')} is associated with a {abs(coef):.4f} unit {direction} in {dependent_var.replace('_', ' ')}"
                if len(independent_vars) > 1:
                    word_content += ", holding all other variables constant"
                word_content += ". "
            
            if p_val < 0.01:
                word_content += "This relationship is highly statistically significant (p < 0.01).\n\n"
            elif p_val < 0.05:
                word_content += "This relationship is statistically significant (p < 0.05).\n\n"
            elif p_val < 0.10:
                word_content += "This relationship is marginally significant (p < 0.10).\n\n"
            else:
                word_content += "This relationship is not statistically significant (p ≥ 0.10).\n\n"

        word_content += f"""4. MODEL DIAGNOSTICS

4.1 Goodness of Fit
The model's R-squared value of {model.rsquared:.4f} indicates that {model.rsquared*100:.1f}% of the variation in {dependent_var.replace('_', ' ')} is explained by the independent variables. This suggests {'a strong' if model.rsquared > 0.7 else 'a moderate' if model.rsquared > 0.5 else 'a weak'} relationship between the variables.

The adjusted R-squared of {model.rsquared_adj:.4f} accounts for the number of variables in the model and provides a more conservative estimate of model fit.

"""

        if include_diagnostics:
            word_content += f"""4.2 Diagnostic Tests

Durbin-Watson Test for Autocorrelation:
The Durbin-Watson statistic is {durbin_watson(model.resid):.4f}. {'This value is within the acceptable range (1.5-2.5), suggesting no significant autocorrelation in the residuals.' if 1.5 <= durbin_watson(model.resid) <= 2.5 else 'This value suggests possible autocorrelation in the residuals, which may require further investigation.'}

Residual Analysis:
• Mean of residuals: {np.mean(model.resid):.6f} (should be approximately zero)
• Standard deviation of residuals: {np.std(model.resid):.4f}
• Root Mean Square Error: {np.sqrt(model.mse_resid):.4f}

4.3 Assumptions Assessment
The validity of OLS regression depends on several key assumptions:
• Linearity: The relationship between variables is linear
• Independence: Observations are independent of each other
• Homoscedasticity: Constant variance of residuals
• Normality: Residuals are normally distributed

[Note: Include diagnostic plots and formal tests results here]

"""

        word_content += f"""5. DISCUSSION

5.1 Key Findings
This analysis reveals several important insights regarding {dependent_var.replace('_', ' ')}:

"""

        significant_vars = []
        for name, coef, p_val in zip(param_names, param_values, p_values):
            if name != 'Intercept' and p_val < 0.05:
                direction = "positive" if coef > 0 else "negative"
                significant_vars.append((name, direction, coef, p_val))
        
        if significant_vars:
            word_content += f"Statistical Significant Relationships:\n"
            for var_name, direction, coef, p_val in significant_vars:
                word_content += f"• {var_name.replace('_', ' ').title()} shows a {direction} relationship with {dependent_var.replace('_', ' ')} (β = {coef:.4f}, p = {p_val:.4f})\n"
        else:
            word_content += "No statistically significant relationships were found at the 95% confidence level.\n"

        word_content += f"""
5.2 Practical Implications
{'The strong statistical relationships identified in this analysis' if len(significant_vars) > 0 else 'While no strong statistical relationships were identified,'} {'provide' if len(significant_vars) > 0 else 'this analysis provides'} insights for [add your domain-specific implications].

5.3 Limitations
Several limitations should be considered when interpreting these results:
• The analysis is based on observational data and cannot establish causation
• The model assumes linear relationships between variables
• Results are specific to the time period and population studied
• [Add other relevant limitations]

6. CONCLUSIONS AND RECOMMENDATIONS

6.1 Summary of Findings
The regression analysis of {dependent_var.replace('_', ' ')} yields the following main conclusions:

1. Model Performance: The model explains {model.rsquared*100:.1f}% of the variation in the dependent variable
2. Statistical Significance: {'The overall model is statistically significant' if model.f_pvalue < 0.05 else 'The overall model lacks statistical significance'}
3. Key Relationships: {len(significant_vars)} out of {len(independent_vars)} variables show significant associations

6.2 Recommendations
Based on these findings, the following recommendations are made:

• [Add domain-specific recommendations based on significant findings]
• [Suggest areas for further research]
• [Recommend policy or business implications if applicable]

6.3 Future Research
Future studies might consider:
• Including additional variables that may influence {dependent_var.replace('_', ' ')}
• Using larger sample sizes for more robust results
• Applying advanced statistical techniques such as instrumental variables
• Conducting longitudinal analysis to establish temporal relationships

REFERENCES
[Add your references here]

APPENDIX
Technical Details:
• Software: Python with statsmodels library
• Estimation Method: Ordinary Least Squares (OLS)
• Confidence Level: 95%
• Missing Data Treatment: Listwise deletion

Generated by: Econometrics Learning Lab
Created by: HAMDI Boulanouar
Professional Statistical Analysis Platform
Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

[End of Report]
"""

        st.success("✅ Word document template generated successfully!")
        
        # Show preview
        with st.expander("📋 **Word Document Preview**", expanded=False):
            st.text(word_content[:4000] + "\n\n[... document continues ...]" if len(word_content) > 4000 else word_content)
        
        st.download_button(
            label="📥 **Download Word Document Template**",
            data=word_content,
            file_name=f"econometrics_report_{dependent_var}_{datetime.now().strftime('%Y%m%d')}.txt",
            mime="text/plain",
            key="download_word_template"
        )
        
        # Word tips
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">📋 Word Document Tips:</h5>
        <div class="red-list">
        • Copy content into Microsoft Word for proper formatting<br>
        • Use Word's built-in heading styles for better organization<br>
        • Add tables using Word's table formatting features<br>
        • Include diagnostic plots from the main application<br>
        • Customize content for your specific research context
        </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"❌ Error generating Word document: {str(e)}")


def educational_materials_page(language):
    """Complete educational materials and explanations"""
    
    st.markdown('<h2 class="section-header">📚 Educational Materials & Learning Center</h2>', 
                unsafe_allow_html=True)
    
    # Create tabs for different educational topics
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Regression Basics",
        "🧮 Mathematical Theory", 
        "🔍 Diagnostic Tests",
        "📈 Interpretation Guide",
        "🎓 Advanced Topics"
    ])
    
    with tab1:
        display_regression_basics()
    
    with tab2:
        display_mathematical_theory()
    
    with tab3:
        display_diagnostic_tests()
    
    with tab4:
        display_interpretation_guide()
    
    with tab5:
        display_advanced_topics()

def display_regression_basics():
    """Display basic regression concepts"""
    
    st.markdown("## 📊 Regression Analysis Fundamentals")
    
    # Interactive concept selector
    concept = st.selectbox(
        "Choose a concept to learn:",
        [
            "What is Linear Regression?",
            "Simple vs Multiple Regression", 
            "The Regression Equation",
            "Key Assumptions",
            "When to Use Regression"
        ]
    )
    
    if concept == "What is Linear Regression?":
        st.markdown("""
        ### 📈 What is Linear Regression?
        
        Linear regression is a statistical method that models the relationship between variables by finding the best-fitting straight line through data points.
        
        #### 🎯 Main Purpose:
        - **Prediction**: Estimate values of one variable based on others
        - **Understanding**: Quantify relationships between variables
        - **Inference**: Test hypotheses about relationships
        
        #### 📊 Real-World Examples:
        """)
        
        examples_df = pd.DataFrame({
            'Industry': ['Business', 'Economics', 'Healthcare', 'Education', 'Marketing'],
            'Question': [
                'How does advertising spend affect sales?',
                'What factors influence house prices?', 
                'How do lifestyle factors affect health outcomes?',
                'What predicts student performance?',
                'How does price affect demand?'
            ],
            'Dependent Variable': ['Sales', 'House Price', 'Health Score', 'Test Score', 'Demand'],
            'Independent Variables': ['Ad Spend, Season', 'Size, Location, Age', 'Exercise, Diet', 'Study Hours, Class Size', 'Price, Income']
        })
        
        st.dataframe(examples_df, use_container_width=True)
        
        # Interactive visualization
        st.markdown("#### 🎨 Interactive Example:")
        
        # Create sample data for demonstration
        np.random.seed(42)
        n_points = st.slider("Number of data points:", 20, 100, 50)
        noise_level = st.slider("Noise level:", 0.5, 3.0, 1.0)
        
        x = np.linspace(0, 10, n_points)
        y = 2 + 1.5 * x + np.random.normal(0, noise_level, n_points)
        
        # Create interactive plot
        fig = px.scatter(x=x, y=y, title="Interactive Regression Example")
        
        # Add regression line
        slope, intercept = np.polyfit(x, y, 1)
        line_x = np.array([x.min(), x.max()])
        line_y = intercept + slope * line_x
        
        fig.add_trace(go.Scatter(
            x=line_x, y=line_y, mode='lines', 
            name=f'Regression Line: y = {intercept:.2f} + {slope:.2f}x',
            line=dict(color='red', width=3)
        ))
        
        st.plotly_chart(fig, use_container_width=True)
        
        r_squared = np.corrcoef(x, y)[0, 1] ** 2
        st.info(f"📊 **R-squared**: {r_squared:.3f} (explains {r_squared*100:.1f}% of variation)")
    
    elif concept == "Simple vs Multiple Regression":
        st.markdown("""
        ### 🔍 Simple vs Multiple Regression
        
        Understanding the difference between simple and multiple regression is crucial for choosing the right analysis method.
        """)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("""
            #### 📈 Simple Linear Regression
            **One predictor variable**
            
            **Equation:** Y = β₀ + β₁X + ε
            
            **Examples:**
            - Sales vs Advertising Spend
            - Height vs Weight  
            - Study Hours vs Test Score
            
            **When to Use:**
            - ✅ Testing one specific relationship
            - ✅ Simple, clear interpretation needed
            - ✅ Limited data available
            - ✅ Exploratory analysis
            """)
        
        with col2:
            st.markdown("""
            #### 📊 Multiple Linear Regression
            **Multiple predictor variables**
            
            **Equation:** Y = β₀ + β₁X₁ + β₂X₂ + ... + βₖXₖ + ε
            
            **Examples:**
            - House Price vs Size + Location + Age
            - Salary vs Education + Experience + Skills
            - Sales vs Price + Advertising + Season
            
            **When to Use:**
            - ✅ Multiple factors influence outcome
            - ✅ Control for confounding variables
            - ✅ Better prediction accuracy needed  
            - ✅ Realistic modeling of complex relationships
            """)
        
        # Comparison table
        st.markdown("#### 📋 Detailed Comparison:")
        
        comparison_df = pd.DataFrame({
            'Aspect': [
                'Number of Predictors', 'Interpretation', 'Visualization', 
                'Assumptions', 'Statistical Power', 'Model Complexity',
                'Real-world Applicability', 'Risk of Overfitting'
            ],
            'Simple Regression': [
                '1 variable', 'Very straightforward', 'Easy (2D scatter plot)',
                'Easier to check', 'Lower', 'Low',
                'Limited', 'Very low'
            ],
            'Multiple Regression': [
                '2+ variables', 'More complex', 'Difficult (multi-dimensional)',
                'More assumptions', 'Higher', 'Higher', 
                'More realistic', 'Higher risk'
            ]
        })
        
        st.dataframe(comparison_df, use_container_width=True)
        
        # Decision flowchart
        st.markdown("""
        #### 🤔 Decision Guide: Which Should I Use?
        
        **Start Here** → Do you have multiple factors that might influence your outcome?
        - **No** → Use Simple Regression
        - **Yes** ↓
        
        **Next** → Do you need to control for other variables?
        - **No** → Consider Simple Regression first
        - **Yes** ↓
        
        **Finally** → Do you have enough data (at least 10-15 observations per variable)?
        - **No** → Use Simple Regression
        - **Yes** → Use Multiple Regression
        """)
    
    elif concept == "The Regression Equation":
        st.markdown("""
        ### 🧮 Understanding the Regression Equation
        
        The regression equation is the mathematical representation of the relationship between variables.
        """)
        
        # Mathematical breakdown
        st.markdown("""
        #### 📐 Mathematical Components:
        """)
        
        st.latex(r"Y_i = \beta_0 + \beta_1 X_{1i} + \beta_2 X_{2i} + ... + \beta_k X_{ki} + \varepsilon_i")
        
        components_df = pd.DataFrame({
            'Component': ['Yi', 'β₀', 'β₁, β₂, ...βₖ', 'X₁i, X₂i, ...Xₖi', 'εi', 'i'],
            'Name': [
                'Dependent Variable', 'Intercept', 'Slope Coefficients', 
                'Independent Variables', 'Error Term', 'Observation Index'
            ],
            'Meaning': [
                'The outcome we want to predict or explain',
                'Value of Y when all X variables = 0',
                'Change in Y for 1-unit change in each X variable',
                'The predictor/explanatory variables',
                'Everything else that affects Y (randomness, omitted variables)',
                'Refers to individual observations (1st, 2nd, 3rd person, etc.)'
            ]
        })
        
        st.dataframe(components_df, use_container_width=True)
        
        # Interactive equation builder
        st.markdown("#### 🛠️ Build Your Own Equation:")
        
        dependent = st.text_input("Dependent Variable (Y):", value="Sales")
        independent = st.text_area("Independent Variables (comma-separated):", value="Advertising, Price, Season")
        
        if dependent and independent:
            indep_vars = [var.strip() for var in independent.split(',')]
            
            equation_display = f"{dependent} = β₀"
            for i, var in enumerate(indep_vars, 1):
                equation_display += f" + β{i}×{var}"
            equation_display += " + ε"
            
            st.markdown(f"**Your Equation:** `{equation_display}`")
            
            # Interpretation
            st.markdown(f"""
            **Interpretation:**
            - **β₀**: The expected {dependent} when {', '.join(indep_vars)} are all zero
            """)
            
            for i, var in enumerate(indep_vars, 1):
                st.markdown(f"- **β{i}**: Change in {dependent} per unit change in {var} (holding other variables constant)")
        
        # Common equation types
        st.markdown("""
        #### 📊 Common Equation Types:
        
        **Linear:** Y = β₀ + β₁X + ε
        - Direct proportional relationship
        - Example: Sales = 1000 + 50×Advertising
        
        **Quadratic:** Y = β₀ + β₁X + β₂X² + ε  
        - Curved relationship (diminishing returns)
        - Example: Yield = 100 + 10×Fertilizer - 0.1×Fertilizer²
        
        **Logarithmic:** log(Y) = β₀ + β₁log(X) + ε
        - Percentage changes
        - Example: log(Wage) = 2 + 0.1×log(Education)
        """)
    
    elif concept == "Key Assumptions":
        st.markdown("""
        ### ✅ The Five Key Assumptions of Linear Regression
        
        For regression results to be valid and reliable, these assumptions must be met:
        """)
        
        # Interactive assumption checker
        assumption = st.selectbox(
            "Select an assumption to learn more:",
            [
                "1. Linearity",
                "2. Independence", 
                "3. Homoscedasticity (Constant Variance)",
                "4. Normality",
                "5. No Multicollinearity"
            ]
        )
        
        if "Linearity" in assumption:
            st.markdown("""
            #### 📈 Assumption 1: Linearity
            
            **What it means:** The relationship between X and Y is actually linear (straight line).
            
            **Why it matters:** If the relationship is curved, linear regression will give biased results.
            
            **How to check:** 
            - Plot Y vs X (should look roughly linear)
            - Plot residuals vs fitted values (should be random scatter)
            
            **What if violated:**
            - Transform variables (log, square root)
            - Add polynomial terms (X²)
            - Use non-linear regression methods
            """)
            
            # Create example plots
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**✅ Good: Linear Relationship**")
                np.random.seed(42)
                x_good = np.linspace(0, 10, 50)
                y_good = 2 + 1.5 * x_good + np.random.normal(0, 1, 50)
                
                fig_good = px.scatter(x=x_good, y=y_good, title="Linear Relationship")
                slope, intercept = np.polyfit(x_good, y_good, 1)
                line_x = np.array([x_good.min(), x_good.max()])
                line_y = intercept + slope * line_x
                fig_good.add_trace(go.Scatter(x=line_x, y=line_y, mode='lines', name='Linear Fit'))
                st.plotly_chart(fig_good, use_container_width=True)
            
            with col2:
                st.markdown("**❌ Problem: Non-Linear Relationship**")
                x_bad = np.linspace(0, 10, 50)
                y_bad = x_bad**2 + np.random.normal(0, 2, 50)
                
                fig_bad = px.scatter(x=x_bad, y=y_bad, title="Non-Linear Relationship")
                slope, intercept = np.polyfit(x_bad, y_bad, 1)
                line_y = intercept + slope * line_x
                fig_bad.add_trace(go.Scatter(x=line_x, y=line_y, mode='lines', name='Poor Linear Fit'))
                st.plotly_chart(fig_bad, use_container_width=True)
        
    elif "Independence" in assumption:
            st.markdown("""
            #### 🔄 Assumption 2: Independence
            
            **What it means:** Each observation is independent of others (no correlation between residuals).
            
            **Why it matters:** Dependent observations lead to incorrect standard errors and invalid tests.
            
            **Common violations:**
            - Time series data (autocorrelation)
            - Clustered data (students in same school)
            - Spatial data (neighboring locations)
            
            **How to check:**
            - Durbin-Watson test for autocorrelation
            - Plot residuals vs time/order
            - Look for patterns in residual plots
            
            **Solutions:**
            - Use robust standard errors
            - Include time trends
            - Use mixed-effects models for clusters
            - Add lagged variables
            """)
            
            # Show autocorrelation example
            st.markdown("**Example: Autocorrelation in Time Series**")
            
            np.random.seed(42)
            time = np.arange(100)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**✅ Independent Residuals**")
                independent_resid = np.random.normal(0, 1, 100)
                fig1 = px.line(x=time, y=independent_resid, title="Independent Residuals")
                fig1.add_hline(y=0, line_dash="dash", line_color="red")
                st.plotly_chart(fig1, use_container_width=True)
            
            with col2:
                st.markdown("**❌ Autocorrelated Residuals**")
                autocorr_resid = np.zeros(100)
                autocorr_resid[0] = np.random.normal(0, 1)
                for i in range(1, 100):
                    autocorr_resid[i] = 0.7 * autocorr_resid[i-1] + np.random.normal(0, 0.5)
                
                fig2 = px.line(x=time, y=autocorr_resid, title="Autocorrelated Residuals")
                fig2.add_hline(y=0, line_dash="dash", line_color="red")
                st.plotly_chart(fig2, use_container_width=True)
        
    elif "Homoscedasticity" in assumption:
            st.markdown("""
            #### 📊 Assumption 3: Homoscedasticity (Constant Variance)
            
            **What it means:** The variance of residuals is constant across all levels of X.
            
            **Why it matters:** Unequal variances lead to inefficient estimates and incorrect standard errors.
            
            **How to check:**
            - Plot residuals vs fitted values
            - Breusch-Pagan test
            - White's test
            
            **Common patterns:**
            - Funnel shape (increasing variance)
            - Megaphone pattern
            - Groupings of different variance
            
            **Solutions:**
            - Transform dependent variable (log, square root)
            - Use weighted least squares
            - Apply robust standard errors
            - Use generalized least squares
            """)
            
            # Visual examples of heteroscedasticity
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**✅ Homoscedasticity (Good)**")
                np.random.seed(42)
                x_homo = np.random.uniform(0, 10, 100)
                y_homo = 2 + 1.5 * x_homo + np.random.normal(0, 1, 100)
                model_homo = np.polyfit(x_homo, y_homo, 1)
                fitted_homo = model_homo[0] * x_homo + model_homo[1]
                resid_homo = y_homo - fitted_homo
                
                fig_homo = px.scatter(x=fitted_homo, y=resid_homo, title="Constant Variance")
                fig_homo.add_hline(y=0, line_dash="dash", line_color="red")
                st.plotly_chart(fig_homo, use_container_width=True)
            
            with col2:
                st.markdown("**❌ Heteroscedasticity (Problem)**")
                x_hetero = np.random.uniform(0, 10, 100)
                y_hetero = 2 + 1.5 * x_hetero + np.random.normal(0, x_hetero * 0.3, 100)
                model_hetero = np.polyfit(x_hetero, y_hetero, 1)
                fitted_hetero = model_hetero[0] * x_hetero + model_hetero[1]
                resid_hetero = y_hetero - fitted_hetero
                
                fig_hetero = px.scatter(x=fitted_hetero, y=resid_hetero, title="Increasing Variance")
                fig_hetero.add_hline(y=0, line_dash="dash", line_color="red")
                st.plotly_chart(fig_hetero, use_container_width=True)
        
    elif "Normality" in assumption:
            st.markdown("""
            #### 📈 Assumption 4: Normality of Residuals
            
            **What it means:** The residuals (errors) follow a normal distribution.
            
            **Why it matters:** 
            - Needed for valid hypothesis tests
            - Confidence intervals accuracy
            - P-values reliability
            
            **Note:** This is about residuals, NOT the original variables!
            
            **How to check:**
            - Q-Q plot of residuals
            - Histogram of residuals
            - Shapiro-Wilk test
            - Jarque-Bera test
            
            **Common violations:**
            - Skewed residuals
            - Heavy-tailed distributions
            - Bimodal distributions
            
            **Solutions:**
            - Transform dependent variable
            - Use robust regression
            - Bootstrap confidence intervals
            - Larger sample sizes (Central Limit Theorem)
            """)
            
            # Interactive normality demonstration
            st.markdown("**🎛️ Interactive Normality Demo:**")
            
            skewness = st.slider("Residual Skewness:", -2.0, 2.0, 0.0, 0.1)
            
            np.random.seed(42)
            if skewness == 0:
                residuals = np.random.normal(0, 1, 1000)
                title = "Normal Residuals (Good)"
                color = "green"
            else:
                from scipy.stats import skewnorm
                residuals = skewnorm.rvs(skewness, size=1000)
                title = f"Skewed Residuals (Skewness = {skewness})"
                color = "red" if abs(skewness) > 0.5 else "orange"
            
            fig = px.histogram(x=residuals, title=title, nbins=30)
            fig.update_traces(marker_color=color, opacity=0.7)
            fig.add_vline(x=np.mean(residuals), line_dash="dash", line_color="black", 
                         annotation_text=f"Mean: {np.mean(residuals):.3f}")
            st.plotly_chart(fig, use_container_width=True)
            
            # Q-Q plot
            fig_qq = go.Figure()
            from scipy import stats
            
            (osm, osr), (slope, intercept, r) = stats.probplot(residuals, dist="norm")
            fig_qq.add_trace(go.Scatter(x=osm, y=osr, mode='markers', name='Sample Quantiles'))
            fig_qq.add_trace(go.Scatter(x=osm, y=slope * osm + intercept, mode='lines', name='Normal Line'))
            fig_qq.update_layout(title="Q-Q Plot vs Normal Distribution", 
                               xaxis_title="Theoretical Quantiles", 
                               yaxis_title="Sample Quantiles")
            st.plotly_chart(fig_qq, use_container_width=True)
        
    elif "Multicollinearity" in assumption:
            st.markdown("""
            #### 🔗 Assumption 5: No Perfect Multicollinearity
            
            **What it means:** Independent variables should not be perfectly correlated with each other.
            
            **Why it matters:**
            - Makes coefficients unstable
            - Standard errors become very large
            - Difficult to separate individual effects
            - Model becomes unreliable
            
            **Types of multicollinearity:**
            - Perfect: One variable is exact function of others
            - High: Variables are highly correlated (r > 0.8)
            - Moderate: Some correlation present (0.5 < r < 0.8)
            
            **How to detect:**
            - Correlation matrix
            - Variance Inflation Factor (VIF)
            - Condition Index
            - Eigenvalues of correlation matrix
            
            **VIF Interpretation:**
            - VIF = 1: No correlation
            - VIF < 5: Acceptable
            - VIF > 10: Problematic
            - VIF > 100: Severe multicollinearity
            
            **Solutions:**
            - Remove highly correlated variables
            - Combine correlated variables into index
            - Use ridge regression or lasso
            - Principal component analysis
            - Collect more data
            """)
            
            # Interactive multicollinearity demo
            st.markdown("**🎛️ Multicollinearity Simulation:**")
            
            correlation = st.slider("Correlation between X1 and X2:", 0.0, 0.99, 0.5, 0.05)
            
            np.random.seed(42)
            n = 100
            x1 = np.random.normal(0, 1, n)
            x2 = correlation * x1 + np.sqrt(1 - correlation**2) * np.random.normal(0, 1, n)
            
            # Calculate VIF
            # VIF = 1 / (1 - R²) where R² is from regressing X1 on X2
            r_squared = correlation ** 2
            vif = 1 / (1 - r_squared) if r_squared < 0.999 else float('inf')
            
            col1, col2 = st.columns(2)
            
            with col1:
                fig_corr = px.scatter(x=x1, y=x2, title=f"X1 vs X2 (r = {correlation:.3f})")
                fig_corr.add_trace(go.Scatter(x=[-3, 3], y=[-3*correlation, 3*correlation], 
                                            mode='lines', name='Trend'))
                st.plotly_chart(fig_corr, use_container_width=True)
            
            with col2:
                vif_color = "green" if vif < 5 else ("orange" if vif < 10 else "red")
                vif_status = "Good" if vif < 5 else ("Moderate" if vif < 10 else "Problematic")
                
                st.markdown(f"""
                **Multicollinearity Assessment:**
                - **Correlation**: {correlation:.3f}
                - **R²**: {r_squared:.3f}
                - **VIF**: {vif:.2f}
                - **Status**: <span style="color: {vif_color};">{vif_status}</span>
                """, unsafe_allow_html=True)
                
                if vif >= 10:
                    st.error("⚠️ High multicollinearity detected!")
                elif vif >= 5:
                    st.warning("⚠️ Moderate multicollinearity present")
                else:
                    st.success("✅ Acceptable level of correlation")
    
    elif concept == "When to Use Regression":
        st.markdown("""
        ### 🤔 When Should You Use Linear Regression?
        
        Linear regression is powerful but not always the right tool. Here's a decision guide:
        """)
        
        # Decision tree
        st.markdown("""
        #### 🌳 Decision Tree: Is Linear Regression Right for You?
        
        **Step 1: What's your goal?**
        - Prediction → ✅ Regression is good
        - Understanding relationships → ✅ Regression is good  
        - Classification (categories) → ❌ Use logistic regression or other methods
        - Time series forecasting → ⚠️ Consider time series methods first
        
        **Step 2: What type of outcome variable?**
        - Continuous numbers → ✅ Linear regression
        - Binary (yes/no) → ❌ Use logistic regression
        - Count data (0,1,2,3...) → ⚠️ Consider Poisson regression
        - Ordered categories → ⚠️ Consider ordinal regression
        
        **Step 3: Do you have enough data?**
        - Rule of thumb: At least 10-15 observations per variable
        - Example: 3 variables needs 30-45 observations minimum
        - More data = more reliable results
        
        **Step 4: Are relationships likely linear?**
        - Plot your data first!
        - If curved, consider transformations or non-linear methods
        """)
        
        # Scenarios table
        st.markdown("#### 📋 Common Scenarios:")
        
        scenarios_df = pd.DataFrame({
            'Scenario': [
                'Predicting house prices',
                'Customer satisfaction survey',
                'Medical diagnosis',
                'Stock price prediction',
                'A/B testing results',
                'Academic performance',
                'Marketing campaign effectiveness',
                'Quality control'
            ],
            'Outcome Type': [
                'Continuous', 'Ordinal scale', 'Binary', 'Time series',
                'Binary/Continuous', 'Continuous', 'Binary/Continuous', 'Pass/Fail'
            ],
            'Best Method': [
                '✅ Linear Regression', '⚠️ Ordinal Regression', '❌ Logistic Regression',
                '⚠️ Time Series Methods', '✅ Depends on metric', '✅ Linear Regression',
                '✅ Depends on outcome', '❌ Logistic Regression'
            ],
            'Why': [
                'Price is continuous', 'Ordered categories', 'Yes/No outcome',
                'Time dependency matters', 'Choose based on KPI', 'Scores are continuous',
                'Revenue (continuous) vs Success (binary)', 'Pass/fail is binary'
            ]
        })
        
        st.dataframe(scenarios_df, use_container_width=True)
        
        # Interactive scenario checker
        st.markdown("#### 🧪 Check Your Scenario:")
        
        user_outcome = st.selectbox(
            "What type of outcome are you trying to predict?",
            ["Continuous numbers (price, score, weight)", 
             "Binary categories (yes/no, pass/fail)",
             "Multiple categories (brand choice, diagnosis)",
             "Counts (number of events)",
             "Time-ordered data"]
        )
        
        user_goal = st.selectbox(
            "What's your primary goal?",
            ["Prediction accuracy", "Understanding relationships", 
             "Hypothesis testing", "Causal inference"]
        )
        
        user_data_size = st.number_input(
            "How many observations do you have?", 
            min_value=10, max_value=10000, value=100
        )
        
        user_variables = st.number_input(
            "How many predictor variables?", 
            min_value=1, max_value=50, value=3
        )
        
        # Provide recommendation
        recommendations = []
        
        if "Continuous numbers" in user_outcome:
            recommendations.append("✅ Linear regression is appropriate for your outcome type")
        elif "Binary categories" in user_outcome:
            recommendations.append("❌ Consider logistic regression instead of linear regression")
        elif "Multiple categories" in user_outcome:
            recommendations.append("❌ Consider multinomial logistic regression or other classification methods")
        elif "Counts" in user_outcome:
            recommendations.append("⚠️ Consider Poisson or negative binomial regression")
        else:
            recommendations.append("⚠️ Consider time series analysis methods")
        
        ratio = user_data_size / user_variables
        if ratio >= 15:
            recommendations.append("✅ You have sufficient data for reliable results")
        elif ratio >= 10:
            recommendations.append("⚠️ Adequate data, but more would be better")
        else:
            recommendations.append("❌ Consider collecting more data or reducing variables")
        
        if user_goal == "Prediction accuracy":
            recommendations.append("✅ Regression is excellent for prediction")
        elif user_goal == "Understanding relationships":
            recommendations.append("✅ Regression provides clear interpretation of relationships")
        elif user_goal == "Hypothesis testing":
            recommendations.append("✅ Regression provides statistical tests for hypotheses")
        else:
            recommendations.append("⚠️ Causal inference requires careful consideration of confounders")
        
        st.markdown("**📋 Recommendations for your scenario:**")
        for rec in recommendations:
            st.markdown(f"- {rec}")

def display_mathematical_theory():
    """Display mathematical foundations"""
    
    st.markdown("## 🧮 Mathematical Theory & Foundations")
    
    theory_topic = st.selectbox(
        "Choose a mathematical topic:",
        [
            "OLS Estimation Theory",
            "Matrix Algebra Approach", 
            "Gauss-Markov Theorem",
            "Maximum Likelihood Estimation",
            "Hypothesis Testing Theory"
        ]
    )
    
    if theory_topic == "OLS Estimation Theory":
        st.markdown("""
        ### 📐 Ordinary Least Squares (OLS) Estimation
        
        OLS finds the coefficients that minimize the sum of squared residuals.
        """)
        
        st.latex(r"\min_{\beta_0, \beta_1} \sum_{i=1}^{n} (Y_i - \beta_0 - \beta_1 X_i)^2")
        
        st.markdown("""
        #### 🎯 The Optimization Problem
        
        We want to find the values of β₀ and β₁ that make the prediction errors as small as possible.
        
        **Why squared errors?**
        - Positive and negative errors don't cancel out
        - Larger errors get penalized more heavily
        - Mathematically convenient (differentiable)
        - Leads to unique solution
        """)
        
        # Derive the normal equations
        st.markdown("#### 📊 Deriving the Solution:")
        
        st.markdown("**Step 1: Take partial derivatives**")
        st.latex(r"\frac{\partial}{\partial \beta_0} \sum_{i=1}^{n} (Y_i - \beta_0 - \beta_1 X_i)^2 = -2\sum_{i=1}^{n} (Y_i - \beta_0 - \beta_1 X_i) = 0")
        st.latex(r"\frac{\partial}{\partial \beta_1} \sum_{i=1}^{n} (Y_i - \beta_0 - \beta_1 X_i)^2 = -2\sum_{i=1}^{n} X_i(Y_i - \beta_0 - \beta_1 X_i) = 0")
        
        st.markdown("**Step 2: Solve the normal equations**")
        st.latex(r"\sum_{i=1}^{n} Y_i = n\beta_0 + \beta_1 \sum_{i=1}^{n} X_i")
        st.latex(r"\sum_{i=1}^{n} X_i Y_i = \beta_0 \sum_{i=1}^{n} X_i + \beta_1 \sum_{i=1}^{n} X_i^2")
        
        st.markdown("**Step 3: Final formulas**")
        st.latex(r"\hat{\beta}_1 = \frac{\sum_{i=1}^{n}(X_i - \bar{X})(Y_i - \bar{Y})}{\sum_{i=1}^{n}(X_i - \bar{X})^2} = \frac{Cov(X,Y)}{Var(X)}")
        st.latex(r"\hat{\beta}_0 = \bar{Y} - \hat{\beta}_1\bar{X}")
        
        # Interactive demonstration
        st.markdown("#### 🎮 Interactive OLS Demonstration:")
        
        # Generate sample data
        np.random.seed(42)
        true_intercept = st.slider("True Intercept (β₀):", 0.0, 5.0, 2.0, 0.1)
        true_slope = st.slider("True Slope (β₁):", 0.5, 3.0, 1.5, 0.1)
        noise_level = st.slider("Noise Level:", 0.1, 2.0, 1.0, 0.1)
        
        x = np.linspace(0, 10, 50)
        y = true_intercept + true_slope * x + np.random.normal(0, noise_level, 50)
        
        # Calculate OLS estimates
        x_mean = np.mean(x)
        y_mean = np.mean(y)
        
        beta1_hat = np.sum((x - x_mean) * (y - y_mean)) / np.sum((x - x_mean)**2)
        beta0_hat = y_mean - beta1_hat * x_mean
        
        # Create plot
        fig = px.scatter(x=x, y=y, title="OLS Estimation Demo")
        
        # Add OLS line
        y_pred = beta0_hat + beta1_hat * x
        fig.add_trace(go.Scatter(x=x, y=y_pred, mode='lines', 
                               name=f'OLS: y = {beta0_hat:.3f} + {beta1_hat:.3f}x'))
        
        # Add true line
        y_true = true_intercept + true_slope * x
        fig.add_trace(go.Scatter(x=x, y=y_true, mode='lines', 
                               name=f'True: y = {true_intercept:.1f} + {true_slope:.1f}x',
                               line=dict(dash='dash')))
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Show estimation accuracy
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Estimated Intercept", f"{beta0_hat:.4f}", 
                     delta=f"{beta0_hat - true_intercept:.4f}")
        with col2:
            st.metric("Estimated Slope", f"{beta1_hat:.4f}", 
                     delta=f"{beta1_hat - true_slope:.4f}")
    
    elif theory_topic == "Matrix Algebra Approach":
        st.markdown("""
        ### 🔢 Matrix Algebra Approach to Regression
        
        For multiple regression, matrix algebra provides an elegant solution.
        """)
        
        st.markdown("#### 📐 Matrix Formulation:")
        
        st.latex(r"\mathbf{Y} = \mathbf{X}\boldsymbol{\beta} + \boldsymbol{\varepsilon}")
        
        st.markdown("""
        Where:
        - **Y** is an n×1 vector of dependent variables
        - **X** is an n×(k+1) design matrix (includes intercept column)
        - **β** is a (k+1)×1 vector of coefficients
        - **ε** is an n×1 vector of errors
        """)
        
        # Show matrix structure
        st.markdown("#### 📊 Matrix Structure Example (n=4, k=2):")
        
        matrix_example = r"""
        \begin{bmatrix} Y_1 \\ Y_2 \\ Y_3 \\ Y_4 \end{bmatrix} = 
        \begin{bmatrix} 1 & X_{11} & X_{12} \\ 1 & X_{21} & X_{22} \\ 1 & X_{31} & X_{32} \\ 1 & X_{41} & X_{42} \end{bmatrix}
        \begin{bmatrix} \beta_0 \\ \beta_1 \\ \beta_2 \end{bmatrix} + 
        \begin{bmatrix} \varepsilon_1 \\ \varepsilon_2 \\ \varepsilon_3 \\ \varepsilon_4 \end{bmatrix}
        """
        
        st.latex(matrix_example)
        
        st.markdown("#### 🎯 OLS Solution:")
        
        st.latex(r"\hat{\boldsymbol{\beta}} = (\mathbf{X}'\mathbf{X})^{-1}\mathbf{X}'\mathbf{Y}")
        
        st.markdown("""
        **Derivation:**
        1. Minimize: **ε'ε = (Y - Xβ)'(Y - Xβ)**
        2. Take derivative: **∂(ε'ε)/∂β = -2X'Y + 2X'Xβ = 0**
        3. Solve: **X'Xβ = X'Y**
        4. Final solution: **β̂ = (X'X)⁻¹X'Y**
        """)
        
        # Properties section
        st.markdown("#### ✅ Key Properties:")
        
        properties_df = pd.DataFrame({
            'Property': [
                'Unbiasedness', 'Efficiency', 'Consistency', 'Normality'
            ],
            'Mathematical_Statement': [
                'E[β̂] = β', 'Var(β̂) is minimum', 'β̂ → β as n → ∞', 'β̂ ~ N(β, σ²(X′X)⁻¹)'
            ],
            'Intuitive_Meaning': [
                'Estimates are correct on average',
                'Most precise among unbiased estimators',
                'Gets better with more data',
                'Follows normal distribution (large samples)'
            ]
        })
        
        st.dataframe(properties_df, use_container_width=True)
        
        # Variance-covariance matrix
        st.markdown("#### 📊 Variance-Covariance Matrix:")
        
        st.latex(r"Var(\hat{\boldsymbol{\beta}}) = \sigma^2(\mathbf{X}'\mathbf{X})^{-1}")
        
        st.markdown("""
        **This matrix tells us:**
        - Diagonal elements: Variances of each coefficient
        - Off-diagonal elements: Covariances between coefficients
        - Used to calculate standard errors and confidence intervals
        """)
    
    elif theory_topic == "Gauss-Markov Theorem":
        st.markdown("""
        ### 🏆 The Gauss-Markov Theorem
        
        This fundamental theorem tells us when OLS is the **Best Linear Unbiased Estimator (BLUE)**.
        """)
        
        st.markdown("#### 📋 Required Assumptions:")
        
        assumptions_df = pd.DataFrame({
            'Assumption': [
                'A1: Linearity', 'A2: Strict Exogeneity', 'A3: No Perfect Multicollinearity',
                'A4: Spherical Errors', 'A5: (Optional) Normality'
            ],
            'Mathematical_Form': [
                'Y = Xβ + ε', 'E[ε|X] = 0', 'rank(X) = k+1',
                'E[εε′|X] = σ²I', 'ε ~ N(0, σ²I)'
            ],
            'Plain_English': [
                'Model is correctly specified',
                'No correlation between X and errors',
                'No perfect correlation among X variables',
                'Constant variance, no autocorrelation',
                'Errors are normally distributed'
            ]
        })
        
        st.dataframe(assumptions_df, use_container_width=True)
        
        st.markdown("#### 🎯 The Theorem Statement:")
        
        st.markdown("""
        <div class="interpretation-box">
        <h5 class="red-text">🏆 Gauss-Markov Theorem:</h5>
        <div class="red-text">
        Under assumptions A1-A4, the OLS estimator β̂ is the <strong>Best Linear Unbiased Estimator</strong>:
        <br><br>
        • <strong>Linear:</strong> β̂ is a linear function of Y<br>
        • <strong>Unbiased:</strong> E[β̂] = β<br>
        • <strong>Best:</strong> Has minimum variance among all linear unbiased estimators
        </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Interactive demonstration of efficiency
        st.markdown("#### 🎮 Efficiency Demonstration:")
        
        sample_size = st.slider("Sample Size:", 20, 200, 100, 10)
        num_simulations = 1000
        
        if st.button("Run Simulation", key="gauss_markov_sim"):
            with st.spinner("Running simulation..."):
                np.random.seed(42)
                
                # True parameters
                true_beta0, true_beta1 = 2.0, 1.5
                
                # Storage for estimates
                ols_estimates = []
                alternative_estimates = []  # Some other linear estimator
                
                for _ in range(num_simulations):
                    # Generate data
                    x = np.random.uniform(0, 10, sample_size)
                    epsilon = np.random.normal(0, 1, sample_size)
                    y = true_beta0 + true_beta1 * x + epsilon
                    
                    # OLS estimates
                    x_mean = np.mean(x)
                    y_mean = np.mean(y)
                    beta1_ols = np.sum((x - x_mean) * (y - y_mean)) / np.sum((x - x_mean)**2)
                    beta0_ols = y_mean - beta1_ols * x_mean
                    
                    ols_estimates.append([beta0_ols, beta1_ols])
                    
                    # Alternative estimator (just for demonstration - less efficient)
                    # Using only first half of data (wasteful but still unbiased)
                    half_n = sample_size // 2
                    x_half = x[:half_n]
                    y_half = y[:half_n]
                    x_half_mean = np.mean(x_half)
                    y_half_mean = np.mean(y_half)
                    beta1_alt = np.sum((x_half - x_half_mean) * (y_half - y_half_mean)) / np.sum((x_half - x_half_mean)**2)
                    beta0_alt = y_half_mean - beta1_alt * x_half_mean
                    
                    alternative_estimates.append([beta0_alt, beta1_alt])
                
                ols_estimates = np.array(ols_estimates)
                alt_estimates = np.array(alternative_estimates)
                
                # Calculate variances
                ols_var_beta1 = np.var(ols_estimates[:, 1])
                alt_var_beta1 = np.var(alt_estimates[:, 1])
                
                # Display results
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**OLS Estimator (Efficient)**")
                    fig1 = px.histogram(x=ols_estimates[:, 1], title="Distribution of β₁ estimates (OLS)")
                    fig1.add_vline(x=true_beta1, line_dash="dash", line_color="red", 
                                  annotation_text="True value")
                    st.plotly_chart(fig1, use_container_width=True)
                    st.metric("Variance of β₁", f"{ols_var_beta1:.6f}")
                
                with col2:
                    st.markdown("**Alternative Estimator (Inefficient)**")
                    fig2 = px.histogram(x=alt_estimates[:, 1], title="Distribution of β₁ estimates (Alternative)")
                    fig2.add_vline(x=true_beta1, line_dash="dash", line_color="red", 
                                  annotation_text="True value")
                    st.plotly_chart(fig2, use_container_width=True)
                    st.metric("Variance of β₁", f"{alt_var_beta1:.6f}")
                
                efficiency_gain = alt_var_beta1 / ols_var_beta1
                st.success(f"🏆 OLS is {efficiency_gain:.2f} times more efficient than the alternative estimator!")
    
    elif theory_topic == "Maximum Likelihood Estimation":
        st.markdown("""
        ### 📊 Maximum Likelihood Estimation (MLE)
        
        Under normality assumption, OLS estimates are identical to MLE estimates.
        """)
        
        st.markdown("#### 🎯 The Likelihood Function:")
        
        st.markdown("If errors are normally distributed:")
        st.latex(r"\varepsilon_i \sim N(0, \sigma^2)")
        
        st.markdown("Then:")
        st.latex(r"Y_i \sim N(\beta_0 + \beta_1 X_i, \sigma^2)")
        
        st.markdown("The likelihood function is:")
        st.latex(r"L(\beta_0, \beta_1, \sigma^2) = \prod_{i=1}^{n} \frac{1}{\sqrt{2\pi\sigma^2}} \exp\left(-\frac{(Y_i - \beta_0 - \beta_1 X_i)^2}{2\sigma^2}\right)")
        
        st.markdown("Taking the log-likelihood:")
        st.latex(r"\ell(\beta_0, \beta_1, \sigma^2) = -\frac{n}{2}\ln(2\pi) - \frac{n}{2}\ln(\sigma^2) - \frac{1}{2\sigma^2}\sum_{i=1}^{n}(Y_i - \beta_0 - \beta_1 X_i)^2")
        
        st.markdown("""
        #### 🔍 Key Insight:
        
        Maximizing the log-likelihood is equivalent to minimizing the sum of squared residuals!
        
        This is why **OLS = MLE** under normality assumption.
        """)
        
        # Interactive likelihood visualization
        st.markdown("#### 🎮 Interactive Likelihood Surface:")
        
        # Generate sample data
        np.random.seed(42)
        x_data = np.array([1, 2, 3, 4, 5])
        y_data = np.array([2.1, 3.9, 6.1, 7.8, 10.2])
        
        # Create grid for likelihood surface
        beta0_range = np.linspace(0, 2, 20)
        beta1_range = np.linspace(1.5, 2.5, 20)
        
        likelihood_surface = np.zeros((20, 20))
        
        for i, b0 in enumerate(beta0_range):
            for j, b1 in enumerate(beta1_range):
                residuals = y_data - (b0 + b1 * x_data)
                ss_res = np.sum(residuals**2)
                likelihood_surface[i, j] = -ss_res  # Negative for maximization
        
        # Plot likelihood surface
        fig = go.Figure(data=go.Contour(
            x=beta1_range,
            y=beta0_range,
            z=likelihood_surface,
            colorscale='Viridis',
            contours=dict(showlabels=True)
        ))
        
        # Add OLS solution
        beta1_ols = np.sum((x_data - np.mean(x_data)) * (y_data - np.mean(y_data))) / np.sum((x_data - np.mean(x_data))**2)
        beta0_ols = np.mean(y_data) - beta1_ols * np.mean(x_data)
        
        fig.add_trace(go.Scatter(
            x=[beta1_ols], y=[beta0_ols],
            mode='markers',
            marker=dict(size=15, color='red', symbol='star'),
            name='OLS/MLE Solution'
        ))
        
        fig.update_layout(
            title='Log-Likelihood Surface',
            xaxis_title='β₁ (Slope)',
            yaxis_title='β₀ (Intercept)'
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        st.info(f"🎯 **OLS/MLE Solution**: β₀ = {beta0_ols:.3f}, β₁ = {beta1_ols:.3f}")
    
    elif theory_topic == "Hypothesis Testing Theory":
        st.markdown("""
        ### 🧪 Hypothesis Testing in Regression
        
        Statistical inference allows us to test hypotheses about population parameters.
        """)
        
        st.markdown("#### 🎯 t-Test for Individual Coefficients:")
        
        st.latex(r"H_0: \beta_j = 0 \quad \text{vs} \quad H_A: \beta_j \neq 0")
        
        st.markdown("Test statistic:")
        st.latex(r"t = \frac{\hat{\beta}_j - 0}{SE(\hat{\beta}_j)} \sim t_{n-k-1}")
        
        st.markdown("Where:")
        st.latex(r"SE(\hat{\beta}_j) = \sqrt{\hat{\sigma}^2 \cdot [(X'X)^{-1}]_{jj}}")
        
        # Interactive t-test demonstration
        st.markdown("#### 🎮 Interactive t-Test Demo:")
        
        true_effect = st.slider("True Effect Size (β₁):", 0.0, 3.0, 1.0, 0.1)
        sample_size_t = st.slider("Sample Size:", 20, 200, 50, 10)
        alpha_level = st.selectbox("Significance Level (α):", [0.01, 0.05, 0.10], index=1)
        
        # Simulate data
        np.random.seed(42)
        x_sim = np.random.normal(0, 1, sample_size_t)
        y_sim = 2 + true_effect * x_sim + np.random.normal(0, 1, sample_size_t)
        
        # Calculate t-statistic
        x_mean = np.mean(x_sim)
        y_mean = np.mean(y_sim)
        beta1_hat = np.sum((x_sim - x_mean) * (y_sim - y_mean)) / np.sum((x_sim - x_mean)**2)
        
        # Calculate standard error
        residuals = y_sim - (y_mean - beta1_hat * x_mean + beta1_hat * x_sim)
        mse = np.sum(residuals**2) / (sample_size_t - 2)
        se_beta1 = np.sqrt(mse / np.sum((x_sim - x_mean)**2))
        
        # t-statistic and p-value
        t_stat = beta1_hat / se_beta1
        df = sample_size_t - 2
        p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df))
        
        # Critical value
        t_critical = stats.t.ppf(1 - alpha_level/2, df)
        
        # Results
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Estimated β₁", f"{beta1_hat:.4f}")
            st.metric("Standard Error", f"{se_beta1:.4f}")
        
        with col2:
            st.metric("t-statistic", f"{t_stat:.3f}")
            st.metric("p-value", f"{p_value:.4f}")
        
        with col3:
            st.metric("Critical Value", f"±{t_critical:.3f}")
            if p_value < alpha_level:
                st.success("✅ Significant")
            else:
                st.error("❌ Not Significant")
        
        # Visualize t-distribution
        t_range = np.linspace(-4, 4, 1000)
        t_density = stats.t.pdf(t_range, df)
        
        fig_t = go.Figure()
        fig_t.add_trace(go.Scatter(x=t_range, y=t_density, mode='lines', name='t-distribution'))
        
        # Add critical regions
        critical_mask = (t_range <= -t_critical) | (t_range >= t_critical)
        fig_t.add_trace(go.Scatter(
            x=t_range[critical_mask], 
            y=t_density[critical_mask], 
            fill='tonexty',
            mode='none',
            name=f'Critical Region (α = {alpha_level})',
            fillcolor='rgba(255,0,0,0.3)'
        ))
        
        # Add observed t-statistic
        fig_t.add_vline(x=t_stat, line_dash="dash", line_color="blue", 
                       annotation_text=f"Observed t = {t_stat:.3f}")
        
        fig_t.update_layout(
            title=f't-Distribution with {df} degrees of freedom',
            xaxis_title='t-value',
            yaxis_title='Density'
        )
        
        st.plotly_chart(fig_t, use_container_width=True)
        
        # F-test for overall significance
        st.markdown("#### 🧪 F-Test for Overall Model Significance:")
        
        st.latex(r"H_0: \beta_1 = \beta_2 = ... = \beta_k = 0")
        st.latex(r"H_A: \text{At least one } \beta_j \neq 0")
        
        st.markdown("F-statistic:")
        st.latex(r"F = \frac{MSR}{MSE} = \frac{SSR/k}{SSE/(n-k-1)} \sim F_{k, n-k-1}")
        
        st.markdown("""
        Where:
        - **SSR**: Sum of Squares Regression (explained variation)
        - **SSE**: Sum of Squares Error (unexplained variation)
        - **MSR**: Mean Square Regression
        - **MSE**: Mean Square Error
        """)

def display_diagnostic_tests():
    """Display comprehensive diagnostic tests explanation"""
    
    st.markdown("## 🔍 Diagnostic Tests Deep Dive")
    
    diagnostic_test = st.selectbox(
        "Choose a diagnostic test to explore:",
        [
            "Heteroscedasticity Tests",
            "Autocorrelation Tests",
            "Normality Tests", 
            "Multicollinearity Detection",
            "Linearity Assessment",
            "Outlier and Influence Detection"
        ]
    )
    
    if diagnostic_test == "Heteroscedasticity Tests":
        st.markdown("""
        ### 🔥 Heteroscedasticity Tests
        
        Testing whether error variance is constant across all observations.
        """)
        
        # Breusch-Pagan Test
        st.markdown("#### 📊 Breusch-Pagan Test")
        
        st.markdown("""
        **Null Hypothesis:** Homoscedasticity (constant variance)
        **Alternative:** Heteroscedasticity (non-constant variance)
        """)
        
        st.markdown("**Procedure:**")
        st.markdown("1. Run original regression: Y = Xβ + ε")
        st.markdown("2. Calculate squared residuals: ε²")
        st.markdown("3. Regress ε² on X variables")
        st.markdown("4. Test significance of this auxiliary regression")
        
        st.latex(r"LM = n \cdot R^2_{aux} \sim \chi^2_k")
        
        # White Test
        st.markdown("#### 🎯 White's Test")
        
        st.markdown("""
        **More general test** - includes squares and cross-products of X variables.
        
        **Procedure:**
        1. Run: ε² = α₀ + α₁X₁ + α₂X₂ + α₃X₁² + α₄X₂² + α₅X₁X₂ + u
        2. Test joint significance of all α coefficients
        """)
        
        # Visual comparison
        st.markdown("#### 👀 Visual Detection:")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**✅ Homoscedasticity (Good)**")
            np.random.seed(42)
            x_homo = np.random.uniform(0, 10, 100)
            y_homo = 2 + 1.5 * x_homo + np.random.normal(0, 1, 100)
            residuals_homo = y_homo - (2 + 1.5 * x_homo)
            
            fig1 = px.scatter(x=x_homo, y=residuals_homo, title="Constant Variance")
            fig1.add_hline(y=0, line_dash="dash", line_color="red")
            st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            st.markdown("**❌ Heteroscedasticity (Problem)**")
            x_hetero = np.random.uniform(0, 10, 100)
            y_hetero = 2 + 1.5 * x_hetero + np.random.normal(0, x_hetero * 0.3, 100)
            residuals_hetero = y_hetero - (2 + 1.5 * x_hetero)
            
            fig2 = px.scatter(x=x_hetero, y=residuals_hetero, title="Increasing Variance")
            fig2.add_hline(y=0, line_dash="dash", line_color="red")
            st.plotly_chart(fig2, use_container_width=True)
        
        # Solutions
        st.markdown("""
        #### 🛠️ Solutions for Heteroscedasticity:
        
        **1. Robust Standard Errors (White/Huber)**
        - Don't change coefficients
        - Correct standard errors and p-values
        - Most common solution
        
        **2. Variable Transformation**
        - Log transformation: log(Y) if variance increases with level
        - Square root: √Y for count data
        - Box-Cox transformation for optimal choice
        
        **3. Weighted Least Squares (WLS)**
        - Weight observations by inverse of variance
        - Need to know or estimate variance function
        
        **4. Generalized Least Squares (GLS)**
        - Full specification of variance-covariance matrix
        - More complex but most efficient
        """)
    
    elif diagnostic_test == "Autocorrelation Tests":
        st.markdown("""
        ### 🔄 Autocorrelation Tests
        
        Testing whether error terms are correlated with each other.
        """)
        
        # Durbin-Watson Test
        st.markdown("#### ⚡ Durbin-Watson Test")
        
        st.latex(r"DW = \frac{\sum_{t=2}^{n}(e_t - e_{t-1})^2}{\sum_{t=1}^{n}e_t^2}")
        
        st.markdown("""
        **Interpretation:**
        - DW ≈ 2: No autocorrelation
        - DW < 2: Positive autocorrelation  
        - DW > 2: Negative autocorrelation
        - Range: 0 to 4
        """)
        
        # Interactive DW demonstration
        st.markdown("#### 🎮 Interactive Autocorrelation Demo:")
        
        autocorr_level = st.slider("Autocorrelation Level (ρ):", -0.9, 0.9, 0.0, 0.1)
        n_obs = st.slider("Number of Observations:", 50, 200, 100, 10)
        
        # Generate autocorrelated residuals
        np.random.seed(42)
        residuals_auto = np.zeros(n_obs)
        residuals_auto[0] = np.random.normal(0, 1)
        
        for t in range(1, n_obs):
            residuals_auto[t] = autocorr_level * residuals_auto[t-1] + np.random.normal(0, 1)
        
        # Calculate DW statistic
        dw_stat = np.sum(np.diff(residuals_auto)**2) / np.sum(residuals_auto**2)
        
        # Plot residuals
        fig_auto = px.line(x=range(n_obs), y=residuals_auto, 
                          title=f"Autocorrelated Residuals (ρ = {autocorr_level}, DW = {dw_stat:.3f})")
        fig_auto.add_hline(y=0, line_dash="dash", line_color="red")
        st.plotly_chart(fig_auto, use_container_width=True)
        
        # Interpretation
        if 1.5 <= dw_stat <= 2.5:
            st.success(f"✅ DW = {dw_stat:.3f}: No significant autocorrelation")
        elif dw_stat < 1.5:
            st.error(f"❌ DW = {dw_stat:.3f}: Positive autocorrelation detected")
        else:
            st.error(f"❌ DW = {dw_stat:.3f}: Negative autocorrelation detected")
        
        # Breusch-Godfrey Test
        st.markdown("#### 🧪 Breusch-Godfrey LM Test")
        
        st.markdown("""
        **More general than Durbin-Watson:**
        - Can test for higher-order autocorrelation
        - Works with lagged dependent variables
        - More powerful test
        
        **Procedure:**
        1. Run original regression: Y = Xβ + ε
        2. Run auxiliary regression: ε̂ₜ = α₀ + α₁X₁ₜ + ... + αₖXₖₜ + ρ₁ε̂ₜ₋₁ + ... + ρₚε̂ₜ₋ₚ + uₜ
        3. Test H₀: ρ₁ = ρ₂ = ... = ρₚ = 0
        """)
        
        st.latex(r"LM = (n-p) \cdot R^2_{aux} \sim \chi^2_p")
        
        # Solutions
        st.markdown("""
        #### 🛠️ Solutions for Autocorrelation:
        
        **1. Robust Standard Errors (Newey-West)**
        - HAC (Heteroskedasticity and Autocorrelation Consistent)
        - Corrects standard errors
        
        **2. Add Lagged Variables**
        - Include Yₜ₋₁, Xₜ₋₁ as explanatory variables
        - Captures dynamic relationships
        
        **3. Cochrane-Orcutt Procedure**
        - Estimate ρ and transform variables
        - Iterative procedure
        
        **4. Time Series Models**
        - ARIMA models
        - Vector Autoregression (VAR)
        """)
    
    elif diagnostic_test == "Normality Tests":
        st.markdown("""
        ### 📊 Normality Tests for Residuals
        
        Testing whether residuals follow a normal distribution.
        """)
        
        # Jarque-Bera Test
        st.markdown("#### 🎯 Jarque-Bera Test")
        
        st.latex(r"JB = \frac{n}{6}\left[S^2 + \frac{1}{4}(K-3)^2\right] \sim \chi^2_2")
        
        st.markdown("""
        Where:
        - **S**: Skewness of residuals
        - **K**: Kurtosis of residuals  
        - **Normal distribution**: S = 0, K = 3
        """)
        
        # Interactive normality demonstration
        st.markdown("#### 🎮 Interactive Normality Test:")
        
        distribution_type = st.selectbox(
            "Choose residual distribution:",
            ["Normal", "Skewed Right", "Skewed Left", "Heavy Tails", "Light Tails"]
        )
        
        sample_size_norm = st.slider("Sample Size:", 50, 500, 200, 50)
        
        np.random.seed(42)
        
        if distribution_type == "Normal":
            residuals_test = np.random.normal(0, 1, sample_size_norm)
        elif distribution_type == "Skewed Right":
            residuals_test = np.random.exponential(1, sample_size_norm) - 1
        elif distribution_type == "Skewed Left":
            residuals_test = -(np.random.exponential(1, sample_size_norm) - 1)
        elif distribution_type == "Heavy Tails":
            residuals_test = np.random.standard_t(3, sample_size_norm)
        else:  # Light Tails
            residuals_test = np.random.uniform(-1.5, 1.5, sample_size_norm)
        
        # Calculate statistics
        skewness_val = stats.skew(residuals_test)
        kurtosis_val = stats.kurtosis(residuals_test) + 3  # Add 3 for normal kurtosis
        jb_stat = sample_size_norm/6 * (skewness_val**2 + (kurtosis_val-3)**2/4)
        jb_pvalue = 1 - stats.chi2.cdf(jb_stat, 2)
        
        # Shapiro-Wilk test (if sample size reasonable)
        if sample_size_norm <= 300:
            sw_stat, sw_pvalue = stats.shapiro(residuals_test)
        else:
            sw_stat, sw_pvalue = None, None
        
        # Display results
        col1, col2 = st.columns(2)
        
        with col1:
            # Histogram
            fig_hist = px.histogram(x=residuals_test, title=f"Distribution: {distribution_type}", 
                                   nbins=30, marginal="box")
            st.plotly_chart(fig_hist, use_container_width=True)
        
        with col2:
            # Q-Q plot
            (osm, osr), (slope, intercept, r) = stats.probplot(residuals_test, dist="norm")
            fig_qq = go.Figure()
            fig_qq.add_trace(go.Scatter(x=osm, y=osr, mode='markers', name='Sample Quantiles'))
            fig_qq.add_trace(go.Scatter(x=osm, y=slope * osm + intercept, mode='lines', name='Normal Line'))
            fig_qq.update_layout(title="Q-Q Plot", 
                               xaxis_title="Theoretical Quantiles", 
                               yaxis_title="Sample Quantiles")
            st.plotly_chart(fig_qq, use_container_width=True)
        
        # Test results
        st.markdown("#### 📋 Test Results:")
        
        results_df = pd.DataFrame({
            'Statistic': ['Skewness', 'Kurtosis', 'Jarque-Bera', 'JB p-value'],
            'Value': [skewness_val, kurtosis_val, jb_stat, jb_pvalue],
            'Normal_Expectation': [0, 3, 'Small', '>0.05'],
            'Interpretation': [
                'Symmetric' if abs(skewness_val) < 0.5 else 'Skewed',
                'Normal tails' if 2.5 < kurtosis_val < 3.5 else 'Non-normal tails',
                'Normal' if jb_pvalue > 0.05 else 'Non-normal',
                'Normal' if jb_pvalue > 0.05 else 'Non-normal'
            ]
        })
        
        st.dataframe(results_df, use_container_width=True)
        
        if sw_stat is not None:
            st.info(f"🔍 **Shapiro-Wilk Test**: W = {sw_stat:.4f}, p-value = {sw_pvalue:.4f}")
            if sw_pvalue > 0.05:
                st.success("✅ Shapiro-Wilk: Residuals appear normal")
            else:
                st.error("❌ Shapiro-Wilk: Residuals not normal")
        
        # Solutions
        st.markdown("""
        #### 🛠️ Solutions for Non-Normal Residuals:
        
        **1. Variable Transformations**
        - Log transformation for right skew
        - Square root for moderate skew
        - Box-Cox transformation for optimal choice
        
        **2. Robust Regression Methods**
        - Huber regression (less sensitive to outliers)
        - LAD regression (minimizes absolute deviations)
        
        **3. Bootstrap Methods**
        - Bootstrap confidence intervals
        - Permutation tests
        
        **4. Larger Sample Sizes**
        - Central Limit Theorem helps
        - Tests become more robust
        """)
    
    elif diagnostic_test == "Multicollinearity Detection":
        st.markdown("""
        ### 🔗 Multicollinearity Detection Methods
        
        Multiple ways to detect problematic correlations between predictors.
        """)
        
        # Correlation Matrix
        st.markdown("#### 📊 Correlation Matrix Analysis")
        
        st.markdown("""
        **Simple but limited approach:**
        - Look for high pairwise correlations (|r| > 0.8)
        - Problem: Doesn't catch complex multicollinearity
        - Multiple variables can be collectively problematic
        """)
        
        # VIF
        st.markdown("#### 🎯 Variance Inflation Factor (VIF)")
        
        st.latex(r"VIF_j = \frac{1}{1-R_j^2}")
        
        st.markdown("""
        Where R²ⱼ comes from regressing Xⱼ on all other X variables.
        
        **Interpretation:**
        - VIF = 1: No correlation
        - VIF < 5: Acceptable  
        - 5 ≤ VIF < 10: Moderate concern
        - VIF ≥ 10: Serious problem
        """)
        
        # Interactive VIF demonstration
        st.markdown("#### 🎮 Interactive Multicollinearity Demo:")
        
        n_vars = st.selectbox("Number of Variables:", [2, 3, 4], index=1)
        correlation_strength = st.slider("Correlation Strength:", 0.0, 0.95, 0.7, 0.05)
        
        np.random.seed(42)
        n_obs = 100
        
        # Generate correlated variables
        if n_vars == 2:
            X1 = np.random.normal(0, 1, n_obs)
            X2 = correlation_strength * X1 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X_matrix = np.column_stack([X1, X2])
            var_names = ['X1', 'X2']
        
        elif n_vars == 3:
            X1 = np.random.normal(0, 1, n_obs)
            X2 = correlation_strength * X1 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X3 = correlation_strength * X1 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X_matrix = np.column_stack([X1, X2, X3])
            var_names = ['X1', 'X2', 'X3']
        
        else:  # 4 variables
            X1 = np.random.normal(0, 1, n_obs)
            X2 = correlation_strength * X1 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X3 = correlation_strength * X1 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X4 = correlation_strength * X2 + np.sqrt(1 - correlation_strength**2) * np.random.normal(0, 1, n_obs)
            X_matrix = np.column_stack([X1, X2, X3, X4])
            var_names = ['X1', 'X2', 'X3', 'X4']
        
        # Calculate correlation matrix
        corr_matrix = np.corrcoef(X_matrix.T)
        
        # Calculate VIF
        vif_values = []
        for i in range(n_vars):
            # R² from regressing Xi on all other X variables
            y_temp = X_matrix[:, i]
            X_temp = np.delete(X_matrix, i, axis=1)
            if X_temp.shape[1] > 0:
                # Add constant for regression
                X_temp_const = sm.add_constant(X_temp)
                try:
                    model_temp = sm.OLS(y_temp, X_temp_const).fit()
                    r_squared = model_temp.rsquared
                    vif = 1 / (1 - r_squared) if r_squared < 0.999 else float('inf')
                except:
                    vif = 1.0
            else:
                vif = 1.0
            vif_values.append(vif)
        
        # Display results
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Correlation Matrix:**")
            corr_df = pd.DataFrame(corr_matrix, columns=var_names, index=var_names)
            st.dataframe(corr_df.round(3), use_container_width=True)
        
        with col2:
            st.markdown("**VIF Analysis:**")
            vif_df = pd.DataFrame({
                'Variable': var_names,
                'VIF': vif_values,
                'Status': ['✅ Good' if vif < 5 else ('⚠️ Moderate' if vif < 10 else '❌ High') for vif in vif_values]
            })
            st.dataframe(vif_df, use_container_width=True)
        
        # Condition Number
        st.markdown("#### 📐 Condition Number")
        
        eigenvals = np.linalg.eigvals(corr_matrix)
        condition_number = np.sqrt(np.max(eigenvals) / np.min(eigenvals))
        
        st.info(f"🔢 **Condition Number**: {condition_number:.2f}")
        
        if condition_number < 10:
            st.success("✅ Low multicollinearity")
        elif condition_number < 30:
            st.warning("⚠️ Moderate multicollinearity")
        else:
            st.error("❌ High multicollinearity")
        
        # Solutions
        st.markdown("""
        #### 🛠️ Solutions for Multicollinearity:
        
        **1. Remove Variables**
        - Drop variables with highest VIF
        - Use domain knowledge to choose
        
        **2. Combine Variables**
        - Create index/composite variables
        - Principal Component Analysis (PCA)
        
        **3. Regularization Methods**
        - Ridge regression (penalizes large coefficients)
        - LASSO (automatic variable selection)
        
        **4. Collect More Data**
        - Sometimes helps separate effects
        - Not always feasible
        """)
    
    elif diagnostic_test == "Linearity Assessment":
        st.markdown("""
        ### 📈 Linearity Assessment Methods
        
        Testing whether relationships are actually linear.
        """)
        
        # Visual methods
        st.markdown("#### 👀 Visual Assessment Methods")
        
        st.markdown("""
        **1. Scatterplots (Y vs each X)**
        - Look for curved patterns
        - Most intuitive method
        
        **2. Residuals vs Fitted Plot**
        - Should show random scatter
        - Patterns indicate non-linearity
        
        **3. Component-Plus-Residual Plots**
        - Partial regression plots
        - Shows relationship controlling for other variables
        """)
        
        # Interactive linearity demo
        st.markdown("#### 🎮 Interactive Linearity Demo:")
        
        relationship_type = st.selectbox(
            "Choose relationship type:",
            ["Linear", "Quadratic", "Logarithmic", "Exponential", "Sine Wave"]
        )
        
        noise_level = st.slider("Noise Level:", 0.1, 2.0, 0.5, 0.1)
        
        np.random.seed(42)
        x_lin = np.linspace(0, 10, 100)
        
        if relationship_type == "Linear":
            y_true = 2 + 1.5 * x_lin
        elif relationship_type == "Quadratic":
            y_true = 2 + 1.5 * x_lin - 0.1 * x_lin**2
        elif relationship_type == "Logarithmic":
            y_true = 2 + 3 * np.log(x_lin + 1)
        elif relationship_type == "Exponential":
            y_true = 2 + 0.3 * np.exp(0.2 * x_lin)
        else:  # Sine Wave
            y_true = 2 + 1.5 * x_lin + 2 * np.sin(x_lin)
        
        y_observed = y_true + np.random.normal(0, noise_level, 100)
        
        # Fit linear model
        slope_lin, intercept_lin = np.polyfit(x_lin, y_observed, 1)
        y_linear_fit = intercept_lin + slope_lin * x_lin
        
        # Calculate residuals
        residuals_lin = y_observed - y_linear_fit
        
        # Create plots
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Data with Linear Fit**")
            fig1 = px.scatter(x=x_lin, y=y_observed, title=f"{relationship_type} Relationship")
            fig1.add_trace(go.Scatter(x=x_lin, y=y_linear_fit, mode='lines', name='Linear Fit'))
            fig1.add_trace(go.Scatter(x=x_lin, y=y_true, mode='lines', name='True Relationship', line=dict(dash='dash')))
            st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            st.markdown("**Residuals vs Fitted**")
            fig2 = px.scatter(x=y_linear_fit, y=residuals_lin, title="Residuals vs Fitted Values")
            fig2.add_hline(y=0, line_dash="dash", line_color="red")
            st.plotly_chart(fig2, use_container_width=True)
        
        # Interpretation
        if relationship_type == "Linear":
            st.success("✅ Residuals show random scatter - linearity assumption satisfied")
        else:
            st.error("❌ Residuals show systematic pattern - linearity assumption violated")
        
        # Formal tests
        st.markdown("#### 🧪 Formal Tests for Linearity")
        
        st.markdown("""
        **1. Ramsey RESET Test**
        - Add powers of fitted values to model
        - Test if they're significant
        - H₀: Linear specification is correct
        """)
        
        st.markdown("""
        **2. Box-Cox Test**
        - Tests for optimal transformation
        - λ = 1 means no transformation needed
        - λ = 0 means log transformation
        """)
        
        # Solutions
        st.markdown("""
        #### 🛠️ Solutions for Non-Linearity:
        
        **1. Variable Transformations**
        - Log transformation: Y = β₀ + β₁ln(X) + ε
        - Square root: Y = β₀ + β₁√X + ε
        - Reciprocal: Y = β₀ + β₁(1/X) + ε
        
        **2. Polynomial Terms**
        - Quadratic: Y = β₀ + β₁X + β₂X² + ε
        - Cubic: Y = β₀ + β₁X + β₂X² + β₃X³ + ε
        
        **3. Interaction Terms**
        - Y = β₀ + β₁X₁ + β₂X₂ + β₃X₁X₂ + ε
        
        **4. Non-Linear Regression**
        - Exponential models
        - Logistic growth models
        - Spline regression
        """)
    
    elif diagnostic_test == "Outlier and Influence Detection":
        st.markdown("""
        ### 🎯 Outlier and Influence Detection
        
        Identifying observations that don't fit the pattern or have excessive influence.
        """)
        
        # Types of unusual observations
        st.markdown("#### 📋 Types of Unusual Observations")
        
        types_df = pd.DataFrame({
            'Type': ['Outlier', 'High Leverage', 'Influential'],
            'Definition': [
                'Large residual (unusual Y value)',
                'Unusual X values (far from center)',
                'Changes results when removed'
            ],
            'Detection_Method': [
                'Standardized residuals > |2|',
                'Leverage > 2(k+1)/n',
                "Cook's Distance > 4/n"
            ],
            'Concern_Level': [
                'Medium', 'Low-Medium', 'High'
            ]
        })
        
        st.dataframe(types_df, use_container_width=True)
        
        # Interactive influence demonstration
        st.markdown("#### 🎮 Interactive Influence Demo:")
        
        include_outlier = st.checkbox("Include Influential Outlier", value=False)
        
        np.random.seed(42)
        n_points = 50
        
        # Generate base data
        x_base = np.random.normal(5, 2, n_points)
        y_base = 2 + 1.5 * x_base + np.random.normal(0, 1, n_points)
        
        if include_outlier:
            # Add influential outlier
            x_data = np.append(x_base, 15)  # High leverage
            y_data = np.append(y_base, 8)   # Unusual Y value
        else:
            x_data = x_base
            y_data = y_base
        
        # Fit model
        slope, intercept = np.polyfit(x_data, y_data, 1)
        
        # Calculate diagnostics
        y_fitted = intercept + slope * x_data
        residuals = y_data - y_fitted
        
        # Standardized residuals
        residual_std = np.std(residuals)
        standardized_residuals = residuals / residual_std
        
        # Leverage (simplified calculation)
        x_mean = np.mean(x_data)
        leverage = 1/len(x_data) + (x_data - x_mean)**2 / np.sum((x_data - x_mean)**2)
        
        # Cook's Distance (simplified)
        cooks_d = standardized_residuals**2 * leverage / (2 * (1 - leverage))
        
        # Create plots
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Regression with/without Outlier**")
            fig1 = px.scatter(x=x_data, y=y_data, title="Data with Regression Line")
            fig1.add_trace(go.Scatter(x=x_data, y=y_fitted, mode='lines', name='Regression Line'))
            
            if include_outlier:
                fig1.add_trace(go.Scatter(x=[x_data[-1]], y=[y_data[-1]], 
                                        mode='markers', marker=dict(size=15, color='red'),
                                        name='Potential Outlier'))
            
            st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            st.markdown("**Cook's Distance**")
            fig2 = px.bar(x=range(len(cooks_d)), y=cooks_d, title="Cook's Distance by Observation")
            fig2.add_hline(y=4/len(x_data), line_dash="dash", line_color="red", 
                          annotation_text="Threshold (4/n)")
            st.plotly_chart(fig2, use_container_width=True)
        
        # Show diagnostics table
        st.markdown("#### 📊 Diagnostic Statistics:")
        
        diagnostic_df = pd.DataFrame({
            'Observation': range(1, len(x_data) + 1),
            'X_Value': x_data.round(3),
            'Y_Value': y_data.round(3),
            'Residual': residuals.round(3),
            'Std_Residual': standardized_residuals.round(3),
            'Leverage': leverage.round(4),
            'Cooks_D': cooks_d.round(4),
            'Flags': [
                ('High Residual' if abs(sr) > 2 else '') + 
                (' High Leverage' if lev > 2*2/len(x_data) else '') +
                (' High Influence' if cd > 4/len(x_data) else '')
                for sr, lev, cd in zip(standardized_residuals, leverage, cooks_d)
            ]
        })
        
        # Highlight problematic observations
        problematic = diagnostic_df[
            (abs(diagnostic_df['Std_Residual']) > 2) |
            (diagnostic_df['Leverage'] > 2*2/len(x_data)) |
            (diagnostic_df['Cooks_D'] > 4/len(x_data))
        ]
        
        if len(problematic) > 0:
            st.warning(f"⚠️ **{len(problematic)} Potentially Problematic Observations:**")
            st.dataframe(problematic, use_container_width=True)
        else:
            st.success("✅ No problematic observations detected")
        
        # Show model impact
        if include_outlier:
            # Fit model without outlier
            slope_no_outlier, intercept_no_outlier = np.polyfit(x_base, y_base, 1)
            
            st.markdown("#### 📈 Impact on Model:")
            
            impact_df = pd.DataFrame({
                'Coefficient': ['Intercept', 'Slope'],
                'With_Outlier': [intercept, slope],
                'Without_Outlier': [intercept_no_outlier, slope_no_outlier],
                'Change': [intercept - intercept_no_outlier, slope - slope_no_outlier],
                'Percent_Change': [
                    100 * abs(intercept - intercept_no_outlier) / abs(intercept_no_outlier),
                    100 * abs(slope - slope_no_outlier) / abs(slope_no_outlier)
                ]
            })
            
            st.dataframe(impact_df.round(4), use_container_width=True)
        
        # Solutions
        st.markdown("""
        #### 🛠️ Dealing with Outliers and Influential Points:
        
        **1. Investigate the Data Point**
        - Check for data entry errors
        - Verify measurement accuracy
        - Consider contextual factors
        
        **2. Statistical Approaches**
        - Robust regression (less sensitive to outliers)
        - Winsorization (cap extreme values)
        - Transformation of variables
        
        **3. Sensitivity Analysis**
        - Run analysis with and without outliers
        - Report both sets of results
        - Check if conclusions change
        
        **4. Model Improvement**
        - Add variables that explain the outlier
        - Use different functional form
        - Consider interaction terms
        """)

def display_interpretation_guide():
    """Display comprehensive interpretation guide"""
    
    st.markdown("## 📈 Interpretation Guide")
    
    interp_topic = st.selectbox(
        "Choose an interpretation topic:",
        [
            "Coefficient Interpretation",
            "Statistical Significance",
            "R-squared and Model Fit",
            "Prediction vs Inference", 
            "Causation vs Correlation",
            "Real-World Applications"
        ]
    )
    
    if interp_topic == "Coefficient Interpretation":
        st.markdown("""
        ### 📊 How to Interpret Regression Coefficients
        
        The art of translating statistical results into meaningful insights.
        """)
        
        # Simple regression interpretation
        st.markdown("#### 🔍 Simple Regression: Y = β₀ + β₁X + ε")
        
        st.markdown("""
        **β₀ (Intercept):**
        - "When X = 0, the expected value of Y is β₀"
        - May or may not be meaningful (depends on context)
        - Example: "When advertising spend = 0, expected sales = $10,000"
        
        **β₁ (Slope):**
        - "A 1-unit increase in X is associated with a β₁ change in Y"
        - This is the marginal effect
        - Example: "Each additional $1,000 in advertising increases sales by $3,500"
        """)
        
        # Interactive example
        st.markdown("#### 🎮 Interactive Interpretation Example:")
        
        # User inputs
        context = st.selectbox(
            "Choose a context:",
            ["Sales vs Advertising", "Wage vs Education", "House Price vs Size", "Test Score vs Study Hours"]
        )
        
        if context == "Sales vs Advertising":
            y_var, x_var = "Sales ($)", "Advertising Spend ($1000s)"
            intercept_val = 50000
            slope_val = 2500
            units_y = "dollars"
            units_x = "thousand dollars"
        elif context == "Wage vs Education":
            y_var, x_var = "Annual Wage ($)", "Years of Education"
            intercept_val = 25000
            slope_val = 3500
            units_y = "dollars per year"
            units_x = "years"
        elif context == "House Price vs Size":
            y_var, x_var = "House Price ($)", "Size (sq ft)"
            intercept_val = 50000
            slope_val = 150
            units_y = "dollars"
            units_x = "square feet"
        else:
            y_var, x_var = "Test Score (%)", "Study Hours per Week"
            intercept_val = 60
            slope_val = 4.5
            units_y = "percentage points"
            units_x = "hours per week"
        
        st.markdown(f"**Estimated Equation:** {y_var} = {intercept_val:,} + {slope_val}×{x_var}")
        
        st.markdown(f"""
        **Interpretation:**
        
        **Intercept ({intercept_val:,}):** 
        When {x_var} = 0, the expected {y_var} is {intercept_val:,} {units_y}.
        """)
        
        if "Advertising" in context:
            st.markdown("💡 *This represents baseline sales with zero advertising.*")
        elif "Education" in context:
            st.markdown("💡 *This represents expected wage with zero years of education (hypothetical).*")
        elif "House" in context:
            st.markdown("💡 *This represents the baseline house price for zero square feet (unrealistic).*")
        else:
            st.markdown("💡 *This represents expected test score with zero study hours.*")
        
        st.markdown(f"""
        **Slope ({slope_val}):**
        A 1-{units_x.split()[-1] if ' ' in units_x else units_x} increase in {x_var} is associated with a {slope_val} {units_y} increase in {y_var}.
        """)
        
        # Practical examples
        st.markdown("**🎯 Practical Examples:**")
        
        if "Advertising" in context:
            st.markdown(f"""
            - Increasing advertising by $5,000 → Sales increase by ${slope_val * 5:,}
            - Doubling current advertising budget → Sales increase by (current budget × ${slope_val:,})
            - ROI of advertising: ${slope_val:,} return per $1,000 invested
            """)
        elif "Education" in context:
            st.markdown(f"""
            - College degree (4 extra years) → Wage increases by ${slope_val * 4:,}/year
            - Master's degree (2 more years) → Additional ${slope_val * 2:,}/year
            - Lifetime earnings difference: 40 years × ${slope_val * 4:,} = ${slope_val * 4 * 40:,}
            """)
        elif "House" in context:
            st.markdown(f"""
            - 500 sq ft larger house → Price increases by ${slope_val * 500:,}
            - 1000 sq ft addition → Price increases by ${slope_val * 1000:,}
            - Price per square foot: ${slope_val}
            """)
        else:
            st.markdown(f"""
            - 5 more study hours per week → Test score increases by {slope_val * 5} percentage points
            - Doubling study time (from 10 to 20 hours) → Score increases by {slope_val * 10} points
            - Each hour of study is worth {slope_val} points
            """)
        
        # Multiple regression interpretation
        st.markdown("#### 🔗 Multiple Regression: Y = β₀ + β₁X₁ + β₂X₂ + ε")
        
        st.markdown("""
        **Key Addition: "Holding other variables constant"**
        
        **β₁:** A 1-unit increase in X₁ is associated with a β₁ change in Y, **holding X₂ constant**.
        
        This is crucial because:
        - It isolates the effect of X₁
        - Controls for confounding variables
        - Allows cleaner causal interpretation
        """)
        
        # Multiple regression example
        st.markdown("**📊 Example: House Price = β₀ + β₁(Size) + β₂(Age) + ε**")
        
        st.markdown("""
        Result: Price = 50,000 + 150×Size - 2,000×Age
        
        **Interpretation:**
        - **Size coefficient (150):** Each additional square foot increases house price by $150, **holding age constant**
        - **Age coefficient (-2,000):** Each additional year of age decreases house price by $2,000, **holding size constant**
        
        **Why "holding constant" matters:**
        Without controlling for age, size might appear more valuable because newer houses are often larger.
        """)
    
    elif interp_topic == "Statistical Significance":
        st.markdown("""
        ### 🧪 Understanding Statistical Significance
        
        What p-values really mean and how to interpret them correctly.
        """)
        
        # Basic concepts
        st.markdown("#### 📋 Basic Concepts")
        
        significance_df = pd.DataFrame({
            'Concept': ['p-value', 'Significance Level (α)', 'Type I Error', 'Type II Error'],
            'Definition': [
                'Probability of observing this result if null hypothesis is true',
                'Threshold for rejecting null hypothesis (usually 0.05)',
                'Rejecting true null hypothesis (false positive)',
                'Failing to reject false null hypothesis (false negative)'
            ],
            'Common_Values': ['0.05, 0.01, 0.10', '0.05, 0.01', '5% with α=0.05', 'Depends on power'],
            'Interpretation': [
                'Lower = stronger evidence against H₀',
                'Researcher choice based on consequences',
                'Controlled by significance level',
                'Reduced by larger samples'
            ]
        })
        
        st.dataframe(significance_df, use_container_width=True)
        
        # Interactive significance demonstration
        st.markdown("#### 🎮 Interactive Significance Demo:")
        
        true_effect = st.slider("True Effect Size:", 0.0, 2.0, 1.0, 0.1)
        sample_size_sig = st.slider("Sample Size:", 20, 200, 50, 10)
        alpha_level_sig = st.selectbox("Significance Level:", [0.01, 0.05, 0.10], index=1)
        
                # Simulate many studies
        n_simulations = 1000
        np.random.seed(42)
        
        p_values = []
        for _ in range(n_simulations):
            # Generate data
            x_sim = np.random.normal(0, 1, sample_size_sig)
            y_sim = true_effect * x_sim + np.random.normal(0, 1, sample_size_sig)
            
            # Calculate t-test
            if np.std(x_sim) > 0:
                correlation = np.corrcoef(x_sim, y_sim)[0, 1]
                t_stat = correlation * np.sqrt(sample_size_sig - 2) / np.sqrt(1 - correlation**2)
                p_val = 2 * (1 - stats.t.cdf(abs(t_stat), sample_size_sig - 2))
                p_values.append(p_val)
        
        # Calculate power and Type I error rate
        significant_results = sum(1 for p in p_values if p < alpha_level_sig)
        power = significant_results / n_simulations if true_effect > 0 else None
        type_i_error = significant_results / n_simulations if true_effect == 0 else None
        
        # Display results
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Significant Results", f"{significant_results}/{n_simulations}")
            if true_effect > 0:
                st.metric("Statistical Power", f"{power:.3f}")
            else:
                st.metric("Type I Error Rate", f"{type_i_error:.3f}")
        
        with col2:
            fig_hist = px.histogram(x=p_values, title="Distribution of p-values", nbins=20)
            fig_hist.add_vline(x=alpha_level_sig, line_dash="dash", line_color="red", 
                              annotation_text=f"α = {alpha_level_sig}")
            st.plotly_chart(fig_hist, use_container_width=True)
        
        with col3:
            if true_effect == 0:
                st.info("🎯 **No True Effect**: We expect 5% false positives with α=0.05")
            else:
                st.info(f"🎯 **True Effect Present**: Power = {power:.1%} chance of detecting it")
        
        # Common misconceptions
        st.markdown("""
        #### ❌ Common Misconceptions About p-values
        
        **WRONG:** "p = 0.05 means there's a 5% chance the null hypothesis is true"
        **RIGHT:** "p = 0.05 means there's a 5% chance of seeing this data if the null hypothesis were true"
        
        **WRONG:** "p = 0.001 is much better than p = 0.049"
        **RIGHT:** "Both provide evidence against the null, but effect size matters more"
        
        **WRONG:** "Non-significant means there's no effect"
        **RIGHT:** "Non-significant means we don't have enough evidence for an effect"
        
        **WRONG:** "Significant means important"
        **RIGHT:** "Significant means statistically detectable, not necessarily practically important"
        """)
        
        # Effect size vs significance
        st.markdown("#### 📏 Effect Size vs Statistical Significance")
        
        # Interactive effect size demo
        effect_demo = st.selectbox(
            "Choose a scenario:",
            ["Small effect, large sample", "Large effect, small sample", "No effect, large sample"]
        )
        
        if effect_demo == "Small effect, large sample":
            demo_effect = 0.1
            demo_n = 1000
            demo_description = "Tiny effect but highly significant due to large sample"
        elif effect_demo == "Large effect, small sample":
            demo_effect = 1.5
            demo_n = 20
            demo_description = "Large effect but might not reach significance due to small sample"
        else:
            demo_effect = 0.0
            demo_n = 1000
            demo_description = "No effect - should not be significant regardless of sample size"
        
        # Generate demo data
        np.random.seed(42)
        x_demo = np.random.normal(0, 1, demo_n)
        y_demo = demo_effect * x_demo + np.random.normal(0, 1, demo_n)
        
        # Calculate statistics
        correlation_demo = np.corrcoef(x_demo, y_demo)[0, 1]
        t_stat_demo = correlation_demo * np.sqrt(demo_n - 2) / np.sqrt(1 - correlation_demo**2)
        p_val_demo = 2 * (1 - stats.t.cdf(abs(t_stat_demo), demo_n - 2))
        
        st.info(f"📊 **Scenario**: {demo_description}")
        st.info(f"🔢 **Results**: Correlation = {correlation_demo:.3f}, p-value = {p_val_demo:.4f}")
        
        if p_val_demo < 0.05:
            st.success("✅ Statistically significant")
        else:
            st.error("❌ Not statistically significant")
    
    elif interp_topic == "R-squared and Model Fit":
        st.markdown("""
        ### 📊 R-squared and Model Fit Assessment
        
        Understanding how well your model explains the data.
        """)
        
        # R-squared explanation
        st.markdown("#### 🎯 What is R-squared?")
        
        st.latex(r"R^2 = 1 - \frac{SSR}{TSS} = 1 - \frac{\sum(Y_i - \hat{Y}_i)^2}{\sum(Y_i - \bar{Y})^2}")
        
        st.markdown("""
        **Intuitive Explanation:**
        - **TSS (Total Sum of Squares)**: How much Y varies around its mean
        - **SSR (Sum of Squared Residuals)**: How much Y varies around the regression line
        - **R²**: Proportion of variance "explained" by the model
        
        **Scale:** 0 to 1 (sometimes reported as 0% to 100%)
        """)
        
        # Interactive R-squared demo
        st.markdown("#### 🎮 Interactive R-squared Demo:")
        
        true_relationship = st.slider("Strength of True Relationship:", 0.0, 2.0, 1.0, 0.1)
        noise_level_r2 = st.slider("Noise Level:", 0.1, 3.0, 1.0, 0.1)
        
        np.random.seed(42)
        x_r2 = np.random.normal(0, 1, 100)
        y_r2 = true_relationship * x_r2 + np.random.normal(0, noise_level_r2, 100)
        
        # Calculate R-squared
        slope_r2, intercept_r2 = np.polyfit(x_r2, y_r2, 1)
        y_pred_r2 = intercept_r2 + slope_r2 * x_r2
        
        ss_res = np.sum((y_r2 - y_pred_r2)**2)
        ss_tot = np.sum((y_r2 - np.mean(y_r2))**2)
        r_squared = 1 - (ss_res / ss_tot)
        
        # Visualization
        fig_r2 = px.scatter(x=x_r2, y=y_r2, title=f"R² = {r_squared:.3f}")
        fig_r2.add_trace(go.Scatter(x=x_r2, y=y_pred_r2, mode='lines', name='Regression Line'))
        fig_r2.add_hline(y=np.mean(y_r2), line_dash="dash", line_color="green", 
                        annotation_text="Mean of Y")
        st.plotly_chart(fig_r2, use_container_width=True)
        
        # Interpretation guidelines
        st.markdown("""
        #### 📋 R-squared Interpretation Guidelines
        
        **General Rules of Thumb:**
        """)
        
        r2_guide_df = pd.DataFrame({
            'R² Range': ['0.00 - 0.25', '0.25 - 0.50', '0.50 - 0.75', '0.75 - 1.00'],
            'Description': ['Weak', 'Moderate', 'Strong', 'Very Strong'],
            'Interpretation': [
                'Model explains little variation',
                'Model explains some variation',
                'Model explains most variation',
                'Model explains almost all variation'
            ],
            'Typical_Fields': [
                'Social sciences, psychology',
                'Economics, marketing',
                'Physical sciences',
                'Controlled experiments'
            ]
        })
        
        st.dataframe(r2_guide_df, use_container_width=True)
        
        # Context matters
        st.markdown("""
        #### ⚠️ Context Matters!
        
        **High R² Expected:**
        - Physics experiments (R² > 0.9)
        - Chemical processes
        - Controlled laboratory studies
        
        **Low R² Normal:**
        - Human behavior studies (R² = 0.1-0.3 can be meaningful)
        - Stock market predictions
        - Social science research
        
        **Remember:** A low R² doesn't mean your model is useless if the relationships are statistically significant and theoretically meaningful.
        """)
        
        # Adjusted R-squared
        st.markdown("#### 📊 Adjusted R-squared")
        
        st.latex(r"\bar{R}^2 = 1 - \frac{(1-R^2)(n-1)}{n-k-1}")
        
        st.markdown("""
        **Why use Adjusted R²?**
        - Regular R² always increases when you add variables
        - Adjusted R² can decrease if new variable doesn't help enough
        - Penalizes overly complex models
        - Better for model comparison
        
        **Rule:** Adjusted R² should be close to regular R². If not, you might be overfitting.
        """)
    
    elif interp_topic == "Prediction vs Inference":
        st.markdown("""
        ### 🔮 Prediction vs Inference: Different Goals, Different Approaches
        
        Understanding when to focus on prediction accuracy vs understanding relationships.
        """)
        
        # Comparison table
        comparison_df = pd.DataFrame({
            'Aspect': [
                'Primary Goal', 'Key Question', 'Model Complexity', 'Variable Selection',
                'Interpretability', 'Overfitting Concern', 'Validation Method',
                'Success Metric', 'Typical Use Cases'
            ],
            'Prediction Focus': [
                'Accurate forecasting', 'How well can we predict?', 'Complex models OK',
                'Include everything helpful', 'Less important', 'High concern',
                'Out-of-sample testing', 'RMSE, MAE, R²', 'Business forecasting, ML applications'
            ],
            'Inference Focus': [
                'Understanding relationships', 'What causes what?', 'Simpler preferred',
                'Theory-driven selection', 'Very important', 'Moderate concern',
                'Statistical significance', 'p-values, confidence intervals', 'Academic research, policy analysis'
            ]
        })
        
        st.dataframe(comparison_df, use_container_width=True)
        
        # Interactive example
        st.markdown("#### 🎮 Interactive Example: House Prices")
        
        analysis_goal = st.selectbox(
            "What's your goal?",
            ["Predict house prices for real estate app", "Understand what makes houses valuable for policy"]
        )
        
        if "Predict" in analysis_goal:
            st.markdown("""
            ### 🔮 Prediction-Focused Approach
            
            **Goal:** Build the most accurate house price predictor
            
            **Strategy:**
            - Include ALL available variables (size, location, age, bedrooms, bathrooms, garage, etc.)
            - Try different models (linear, polynomial, machine learning)
            - Focus on out-of-sample prediction accuracy
            - Don't worry too much about interpreting individual coefficients
            
            **Variables to Include:**
            ✅ Square footage, bedrooms, bathrooms
            ✅ Location dummies for every neighborhood  
            ✅ Age, age², age³ (capture non-linear effects)
            ✅ Interaction terms (size × location)
            ✅ Seasonal variables, market trends
            ✅ School district ratings, crime statistics
            
            **Success Measures:**
            - RMSE on test set < $20,000
            - R² > 0.85 on new data
            - Consistently beats competitor models
            """)
            
        else:
            st.markdown("""
            ### 🧪 Inference-Focused Approach
            
            **Goal:** Understand what factors affect house values for urban planning
            
            **Strategy:**
            - Include theoretically important variables only
            - Keep model interpretable
            - Focus on statistical significance and effect sizes
            - Control for confounding variables carefully
            
            **Variables to Include:**
            ✅ Key policy variables: School quality, public transit access
            ✅ Standard controls: Size, age, neighborhood fixed effects
            ❌ Too many interaction terms (hard to interpret)
            ❌ Variables you can't policy-control (exact GPS coordinates)
            
            **Success Measures:**
            - Clear, significant coefficients for policy variables
            - Robust results across different specifications
            - Effect sizes large enough to matter for policy
            """)
        
        # Practical tips
        st.markdown("""
        ### 💡 Practical Tips for Each Approach
        
        #### 🔮 For Prediction Projects:
        1. **Split your data:** Train (60%), Validation (20%), Test (20%)
        2. **Try multiple models:** Linear, polynomial, regularized, ensemble
        3. **Feature engineering:** Create interactions, transformations
        4. **Cross-validation:** Use k-fold CV to select best model
        5. **Monitor performance:** Track prediction accuracy over time
        
        #### 🧪 For Inference Projects:
        1. **Start with theory:** What relationships do you expect?
        2. **Control for confounders:** Include variables that affect both X and Y
        3. **Check robustness:** Do results hold with different specifications?
        4. **Report effect sizes:** Don't just focus on significance
        5. **Consider causation:** Use instrumental variables, natural experiments
        """)
        
        # Common mistakes
        st.markdown("""
        ### ❌ Common Mistakes to Avoid
        
        **Prediction Mistakes:**
        - Using training accuracy to evaluate model
        - Overfitting to specific dataset
        - Ignoring data leakage (future information in past predictions)
        - Not updating models as world changes
        
        **Inference Mistakes:**
        - Including too many variables without theory
        - Claiming causation from correlation
        - P-hacking (trying many specifications until something is significant)
        - Ignoring omitted variable bias
        """)
    
    elif interp_topic == "Causation vs Correlation":
        st.markdown("""
        ### ⚡ Causation vs Correlation: The Most Important Distinction
        
        Just because two variables are related doesn't mean one causes the other.
        """)
        
        # Core concepts
        st.markdown("""
        #### 🎯 Core Concepts
        
        **Correlation:** Two variables move together
        **Causation:** One variable actually influences the other
        
        **The Problem:** Regression only shows correlation, but we often want to know about causation.
        """)
        
        # Famous examples
        st.markdown("#### 📚 Famous Examples of Spurious Correlation")
        
        examples_df = pd.DataFrame({
            'Variable 1': [
                'Ice cream sales', 'Number of firefighters', 'Shoe size', 
                'Pirates worldwide', 'Chocolate consumption', 'Height'
            ],
            'Variable 2': [
                'Drowning deaths', 'Fire damage', 'Vocabulary size',
                'Global temperature', 'Nobel prizes per capita', 'Salary'
            ],
            'Correlation': [
                'Positive', 'Positive', 'Positive',
                'Negative', 'Positive', 'Positive'
            ],
            'True_Cause': [
                'Hot weather (confounding)', 'Fire size (reverse causation)',
                'Age (confounding)', 'Economic development (confounding)',
                'Wealth (confounding)', 'Education, skills (confounding)'
            ]
        })
        
        st.dataframe(examples_df, use_container_width=True)
        
        # Interactive causation demo
        st.markdown("#### 🎮 Interactive Causation Demo")
        
        scenario = st.selectbox(
            "Choose a scenario to analyze:",
            [
                "Education → Income",
                "Exercise → Health", 
                "Advertising → Sales",
                "Class Size → Student Performance"
            ]
        )
        
        if scenario == "Education → Income":
            st.markdown("""
            ### 🎓 Education and Income
            
            **Observed:** People with more education earn more money.
            **Question:** Does education cause higher income, or is something else going on?
            
            **Potential Confounders:**
            - **Family background:** Rich families can afford both good education and provide job connections
            - **Natural ability:** Smart people get more education AND higher-paying jobs
            - **Motivation:** Motivated people pursue education AND work harder
            
            **Evidence for Causation:**
            ✅ Natural experiments (draft lottery affecting college attendance)
            ✅ Changes in compulsory schooling laws
            ✅ Scholarship randomization
            
            **Evidence Against Simple Causation:**
            ❌ Returns to education vary greatly by field
            ❌ Signaling theory: Degree signals ability more than teaching skills
            ❌ Sheepskin effects: Big jumps at degree completion
            """)
            
        elif scenario == "Exercise → Health":
            st.markdown("""
            ### 🏃‍♀️ Exercise and Health
            
            **Observed:** People who exercise more are healthier.
            **Question:** Does exercise cause better health?
            
            **Potential Confounders:**
            - **Income:** Rich people can afford gyms AND healthcare
            - **Time availability:** People with less stressful jobs exercise more AND are healthier
            - **Genetics:** Some people naturally active AND naturally healthy
            - **Reverse causation:** Healthy people exercise more (not the reverse)
            
            **Evidence for Causation:**
            ✅ Randomized controlled trials show health improvements
            ✅ Biological mechanisms are well understood
            ✅ Dose-response relationship (more exercise = better health)
            
            **Complications:**
            ⚠️ Optimal amount varies by person
            ⚠️ Too much exercise can be harmful
            ⚠️ Different types of exercise have different effects
            """)
            
        elif scenario == "Advertising → Sales":
            st.markdown("""
            ### 📺 Advertising and Sales
            
            **Observed:** Companies that advertise more sell more.
            **Question:** Does advertising cause higher sales?
            
            **Potential Confounders:**
            - **Company size:** Big companies advertise more AND have more sales
            - **Product quality:** Good products get more advertising budget AND sell more
            - **Market conditions:** Companies advertise more in good markets
            - **Reverse causation:** High sales fund more advertising
            
            **Evidence for Causation:**
            ✅ Randomized experiments by companies
            ✅ Natural experiments (ad bans, strikes)
            ✅ Immediate response to ad campaigns
            
            **Complications:**
            ⚠️ Diminishing returns to advertising
            ⚠️ Brand advertising vs direct response advertising
            ⚠️ Long-term vs short-term effects
            """)
            
        else:  # Class Size
            st.markdown("""
            ### 👥 Class Size and Student Performance
            
            **Observed:** Students in smaller classes often perform better.
            **Question:** Does reducing class size cause better performance?
            
            **Potential Confounders:**
            - **School resources:** Rich schools have smaller classes AND better everything else
            - **Student selection:** Private schools with small classes select better students
            - **Teacher quality:** Good teachers attract smaller classes
            - **Parental involvement:** Engaged parents choose schools with smaller classes
            
            **Evidence for Causation:**
            ✅ Tennessee STAR experiment (random assignment)
            ✅ Natural experiments with enrollment cutoffs
            ✅ Maimonides' rule studies
            
            **Evidence Against Simple Causation:**
            ❌ Effects vary greatly by context
            ❌ Very expensive intervention for modest gains
            ❌ Teacher quality matters more than class size
            """)
        
        # Tools for establishing causation
        st.markdown("""
        ### 🛠️ Tools for Establishing Causation
        
        #### 🥇 Gold Standard: Randomized Controlled Trials (RCTs)
        - Randomly assign treatment (education, exercise, class size)
        - Compare outcomes between treatment and control groups
        - **Problem:** Often impossible or unethical for important questions
        
        #### 🥈 Natural Experiments
        - Find situations where assignment was "as good as random"
        - Examples: Draft lottery, policy changes, weather shocks
        - Use these as instrumental variables
        
        #### 🥉 Quasi-Experimental Methods
        - **Regression Discontinuity:** Sharp cutoffs in treatment assignment
        - **Difference-in-Differences:** Compare changes over time across groups
        - **Fixed Effects:** Control for time-invariant confounders
        
        #### 📊 Observational Data Best Practices
        - Control for all observable confounders
        - Test robustness across specifications
        - Look for dose-response relationships
        - Check for reverse causation
        - Use theory to guide variable selection
        """)
        
        # Warning signs
        st.markdown("""
        ### ⚠️ Warning Signs of Weak Causal Claims
        
        **Red Flags:**
        ❌ "Studies show that X is linked to Y" (correlation language)
        ❌ No discussion of potential confounders
        ❌ Results that seem too good to be true
        ❌ Cherry-picked studies or timeframes
        ❌ Correlation coefficient interpreted as causal effect
        
        **Green Flags:**
        ✅ Multiple independent studies reach same conclusion
        ✅ Researchers acknowledge limitations
        ✅ Biological/economic mechanism is plausible
        ✅ Dose-response relationship
        ✅ Results robust to different specifications
        """)
    
    elif interp_topic == "Real-World Applications":
        st.markdown("""
        ### 🌍 Real-World Applications: From Theory to Practice
        
        How regression analysis is used across different fields and industries.
        """)
        
        # Field selection
        field = st.selectbox(
            "Choose a field to explore:",
            [
                "Business & Marketing",
                "Healthcare & Medicine",
                "Economics & Policy",
                "Sports Analytics",
                "Environmental Science"
            ]
        )
        
        if field == "Business & Marketing":
            st.markdown("""
            ### 💼 Business & Marketing Applications
            
            #### 📊 Customer Analytics
            **Question:** What factors drive customer lifetime value?
            
            **Model:** CLV = β₀ + β₁(Purchase_Frequency) + β₂(Average_Order_Value) + β₃(Customer_Age) + ε
            
            **Business Impact:**
            - Identify high-value customer segments
            - Optimize marketing spend allocation
            - Predict customer churn risk
            - Personalize product recommendations
            
            #### 📈 Pricing Optimization
            **Question:** How does price affect demand?
            
            **Model:** Log(Quantity) = β₀ + β₁Log(Price) + β₂(Competitor_Price) + β₃(Seasonality) + ε
            
            **Key Insights:**
            - β₁ = price elasticity of demand
            - If β₁ = -1.5, then 1% price increase → 1.5% quantity decrease
            - Revenue maximization vs profit maximization
            
            #### 🎯 Marketing Mix Modeling
            **Question:** What's the ROI of different marketing channels?
            
            **Model:** Sales = β₀ + β₁(TV_Ads) + β₂(Digital_Ads) + β₃(Social_Media) + β₄(Email) + ε
            
            **Practical Use:**
            - Budget allocation across channels
            - Measure incremental lift from campaigns
            - Account for diminishing returns
            - Plan optimal media mix
            """)
            
        elif field == "Healthcare & Medicine":
            st.markdown("""
            ### 🏥 Healthcare & Medicine Applications
            
            #### 💊 Drug Efficacy Studies
            **Question:** Does this treatment improve patient outcomes?
            
            **Model:** Recovery_Time = β₀ + β₁(Treatment) + β₂(Age) + β₃(Severity) + β₄(Comorbidities) + ε
            
            **Critical Considerations:**
            - Control for patient characteristics
            - Account for selection bias
            - Multiple comparison corrections
            - Clinical vs statistical significance
            
            #### 🏥 Hospital Performance
            **Question:** What factors affect patient satisfaction?
            
            **Model:** Satisfaction = β₀ + β₁(Wait_Time) + β₂(Nurse_Ratio) + β₃(Room_Quality) + β₄(Communication) + ε
            
            **Applications:**
            - Quality improvement initiatives
            - Resource allocation decisions
            - Benchmarking across hospitals
            - Patient experience optimization
            
            #### 🧬 Epidemiological Studies
            **Question:** What lifestyle factors affect disease risk?
            
            **Model:** Log(Risk) = β₀ + β₁(Smoking) + β₂(Exercise) + β₃(Diet) + β₄(Genetics) + ε
            
            **Challenges:**
            - Long-term follow-up required
            - Multiple confounding factors
            - Measurement error in lifestyle variables
            - Ethical constraints on experiments
            """)
            
        elif field == "Economics & Policy":
            st.markdown("""
            ### 🏛️ Economics & Policy Applications
            
            #### 💰 Minimum Wage Impact
            **Question:** Do minimum wage increases reduce employment?
            
            **Model:** Employment_Change = β₀ + β₁(Min_Wage_Change) + β₂(Economic_Conditions) + State_FE + Time_FE + ε
            
            **Policy Implications:**
            - If β₁ < 0: Minimum wage reduces employment
            - If β₁ ≈ 0: No significant employment effect
            - If β₁ > 0: Possible monopsony effects
            
            #### 🏫 Education Policy
            **Question:** Does class size reduction improve student outcomes?
            
            **Model:** Test_Score = β₀ + β₁(Class_Size) + β₂(Teacher_Quality) + β₃(Student_Background) + School_FE + ε
            
            **Policy Design:**
            - Cost-benefit analysis of class size reduction
            - Optimal resource allocation
            - Teacher hiring vs class size trade-offs
            
            #### 🌍 Environmental Policy
            **Question:** Do carbon taxes reduce emissions?
            
            **Model:** Log(Emissions) = β₀ + β₁(Carbon_Tax) + β₂(GDP) + β₃(Energy_Prices) + Country_FE + Year_FE + ε
            
            **Policy Evaluation:**
            - Effectiveness of environmental regulations
            - Design of carbon pricing systems
            - International competitiveness concerns
            """)
            
        elif field == "Sports Analytics":
            st.markdown("""
            ### ⚽ Sports Analytics Applications
            
            #### 🏀 Player Performance
            **Question:** What stats predict wins in basketball?
            
            **Model:** Win_Probability = β₀ + β₁(Field_Goal_%) + β₂(Rebounds) + β₃(Turnovers) + β₄(Free_Throws) + ε
            
            **Applications:**
            - Player evaluation and contracts
            - Game strategy optimization
            - Draft pick analysis
            - Team building decisions
            
            #### ⚾ Baseball Sabermetrics
            **Question:** What's the relationship between team payroll and wins?
            
            **Model:** Wins = β₀ + β₁(Payroll) + β₂(Previous_Year_Wins) + β₃(Market_Size) + ε
            
            **Insights:**
            - Diminishing returns to spending
            - Market inefficiencies
            - Value of player development
            - "Moneyball" strategies
            
            #### 🏈 Fantasy Sports
            **Question:** How do we predict player fantasy points?
            
            **Model:** Fantasy_Points = β₀ + β₁(Opponent_Defense_Rank) + β₂(Weather) + β₃(Injury_Status) + β₄(Recent_Form) + ε
            
            **Commercial Value:**
            - Daily fantasy sports optimization
            - Player pricing algorithms
            - Risk management for operators
            - User engagement strategies
            """)
            
        else:  # Environmental Science
            st.markdown("""
            ### 🌱 Environmental Science Applications
            
            #### 🌡️ Climate Change Analysis
            **Question:** How do human activities affect global temperature?
            
            **Model:** Temperature_Anomaly = β₀ + β₁(CO2_Levels) + β₂(Solar_Activity) + β₃(Volcanic_Activity) + Time_Trend + ε
            
            **Scientific Impact:**
            - Attribution of climate change
            - Projection of future warming
            - Policy target setting
            - Economic impact assessment
            
            #### 🏭 Air Quality Studies
            **Question:** What factors affect urban air pollution?
            
            **Model:** PM2.5 = β₀ + β₁(Traffic_Volume) + β₂(Industrial_Activity) + β₃(Weather) + β₄(Regulations) + ε
            
            **Policy Applications:**
            - Emission source identification
            - Regulation effectiveness evaluation
            - Public health impact assessment
            - Urban planning decisions
            
            #### 🌊 Ecosystem Health
            **Question:** How does land use affect biodiversity?
            
            **Model:** Species_Count = β₀ + β₁(Forest_Cover) + β₂(Urban_Area) + β₃(Agriculture) + β₄(Water_Quality) + ε
            
            **Conservation Impact:**
            - Habitat protection prioritization
            - Development impact assessment
            - Restoration project planning
            - Species conservation strategies
            """)
        
        # Best practices across fields
        st.markdown("""
        ### 🎯 Best Practices Across All Fields
        
        #### 📋 Before Starting Analysis
        1. **Define clear research question**
        2. **Understand the data generating process**
        3. **Identify potential confounders**
        4. **Consider data limitations**
        5. **Plan robustness checks**
        
        #### 🔍 During Analysis
        1. **Start with descriptive statistics**
        2. **Check all regression assumptions**
        3. **Test multiple model specifications**
        4. **Look for non-linear relationships**
        5. **Validate with out-of-sample data**
        
        #### 📊 Communicating Results
        1. **Report effect sizes, not just p-values**
        2. **Use confidence intervals**
        3. **Discuss practical significance**
        4. **Acknowledge limitations**
        5. **Provide context for non-experts**
        
        #### ⚠️ Common Pitfalls to Avoid
        - Correlation ≠ Causation
        - Sample selection bias
        - Omitted variable bias
        - Measurement error
        - Multiple testing without correction
        - Overfitting to sample data
        """)

def display_advanced_topics():
    """Display advanced econometrics topics"""
    
    st.markdown("## 🎓 Advanced Topics")
    
    advanced_topic = st.selectbox(
        "Choose an advanced topic:",
        [
            "Panel Data Models",
            "Instrumental Variables", 
            "Difference-in-Differences",
            "Regression Discontinuity",
            "Time Series Econometrics",
            "Limited Dependent Variables"
        ]
    )
    
    if advanced_topic == "Panel Data Models":
        st.markdown("""
        ### 📊 Panel Data Models: Following Units Over Time
        
        Panel data combines cross-sectional and time series dimensions, allowing us to control for unobserved heterogeneity.
        """)
        
        st.markdown("""
        #### 🎯 What is Panel Data?
        
        **Structure:** Multiple units (individuals, firms, countries) observed over multiple time periods
        
        **Example:** 1,000 workers observed for 10 years each = 10,000 observations
        
        **Advantages:**
        - Control for time-invariant unobserved factors
        - More variation to identify effects
        - Can study dynamics and changes
        - Reduce omitted variable bias
        """)
        
        # Model types
        st.markdown("#### 🔧 Three Main Panel Data Models:")
        
        models_df = pd.DataFrame({
            'Model': ['Pooled OLS', 'Fixed Effects (FE)', 'Random Effects (RE)'],
            'Equation': [
                'Yᵢₜ = β₀ + β₁Xᵢₜ + εᵢₜ',
                'Yᵢₜ = αᵢ + β₁Xᵢₜ + εᵢₜ',
                'Yᵢₜ = β₀ + β₁Xᵢₜ + (αᵢ + εᵢₜ)'
            ],
            'Assumes': [
                'No individual heterogeneity',
                'Individual effects correlated with X',
                'Individual effects uncorrelated with X'
            ],
            'When_to_Use': [
                'No unobserved heterogeneity',
                'Worried about omitted variable bias',
                'Random sample from population'
            ]
        })
        
        st.dataframe(models_df, use_container_width=True)
        
        # Interactive demonstration
        st.markdown("#### 🎮 Panel Data Simulation:")
        
        individual_effect_size = st.slider("Individual Effect Size:", 0.0, 3.0, 1.0, 0.1)
        within_variation = st.slider("Within-Individual Variation:", 0.1, 2.0, 0.5, 0.1)
        
        # Generate panel data
        np.random.seed(42)
        n_individuals = 50
        n_periods = 5
        
        # Individual effects
        individual_effects = np.random.normal(0, individual_effect_size, n_individuals)
        
        # Generate data
        panel_data = []
        for i in range(n_individuals):
            for t in range(n_periods):
                x = np.random.normal(2, 1) + individual_effects[i] * 0.5  # X correlated with individual effect
                y = 1 + 1.5 * x + individual_effects[i] + np.random.normal(0, within_variation)
                panel_data.append([i+1, t+1, x, y, individual_effects[i]])
        
        panel_df = pd.DataFrame(panel_data, columns=['Individual', 'Period', 'X', 'Y', 'True_Effect'])
        
        # Compare models
        # Pooled OLS
        pooled_slope = np.cov(panel_df['X'], panel_df['Y'])[0,1] / np.var(panel_df['X'])
        
        # Fixed effects (demeaning)
        panel_df['X_demeaned'] = panel_df.groupby('Individual')['X'].transform(lambda x: x - x.mean())
        panel_df['Y_demeaned'] = panel_df.groupby('Individual')['Y'].transform(lambda x: x - x.mean())
        fe_slope = np.cov(panel_df['X_demeaned'], panel_df['Y_demeaned'])[0,1] / np.var(panel_df['X_demeaned'])
        
        # Show results
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Pooled OLS Results:**")
            st.metric("Estimated Slope", f"{pooled_slope:.3f}")
            st.metric("Bias", f"{pooled_slope - 1.5:.3f}")
            if abs(pooled_slope - 1.5) > 0.2:
                st.error("❌ Substantial bias due to omitted individual effects")
            else:
                st.success("✅ Relatively unbiased")
        
        with col2:
            st.markdown("**Fixed Effects Results:**")
            st.metric("Estimated Slope", f"{fe_slope:.3f}")
            st.metric("Bias", f"{fe_slope - 1.5:.3f}")
            if abs(fe_slope - 1.5) > 0.2:
                st.warning("⚠️ Some bias (random variation)")
            else:
                st.success("✅ Unbiased estimate")
        
        # Visualization
        fig_panel = px.scatter(panel_df.head(100), x='X', y='Y', color='Individual', 
                             title="Panel Data Visualization (First 100 observations)")
        st.plotly_chart(fig_panel, use_container_width=True)
        
    elif advanced_topic == "Instrumental Variables":
        st.markdown("""
        ### 🎯 Instrumental Variables: Solving Endogeneity Problems
        
        When X and the error term are correlated, OLS is biased. IV provides a solution.
        """)
        
        st.markdown("""
        #### ⚠️ The Endogeneity Problem
        
        **Problem:** E[X·ε] ≠ 0 (X is correlated with the error term)
        
        **Causes:**
        1. **Omitted variables:** Missing confounders in error term
        2. **Reverse causation:** Y affects X as well as X affecting Y
        3. **Measurement error:** X measured with error
        4. **Sample selection:** Non-random sample selection
        
        **Consequence:** OLS estimates are biased and inconsistent
        """)
        
        # IV requirements
        st.markdown("#### 🔑 Requirements for a Valid Instrument (Z)")
        
        requirements_df = pd.DataFrame({
            'Requirement': ['Relevance', 'Exogeneity', 'Exclusion Restriction'],
            'Mathematical': ['Cov(Z,X) ≠ 0', 'Cov(Z,ε) = 0', 'Z affects Y only through X'],
            'Plain_English': [
                'Instrument is correlated with endogenous variable',
                'Instrument is uncorrelated with error term',
                'Instrument has no direct effect on outcome'
            ],
            'Testable': ['Yes (F-test)', 'No (assumption)', 'Partially (overidentification)']
        })
        
        st.dataframe(requirements_df, use_container_width=True)
        
        # Famous examples
        st.markdown("#### 📚 Famous IV Examples")
        
        examples_df = pd.DataFrame({
            'Study': ['Returns to Education', 'Military Service Effect', 'Institutions & Growth'],
            'Question': [
                'How does education affect wages?',
                'How does military service affect earnings?',
                'How do institutions affect economic growth?'
            ],
            'Endogeneity_Problem': [
                'Ability affects both education and wages',
                'People choose military service',
                'Rich countries choose better institutions'
            ],
            'Instrument': [
                'Compulsory schooling laws',
                'Draft lottery numbers',
                'Settler mortality rates'
            ],
            'Logic': [
                'Laws force some to get more education',
                'Random assignment to military',
                'Geography affects colonial institutions'
            ]
        })
        
        st.dataframe(examples_df, use_container_width=True)
        
        # Two-stage least squares
        st.markdown("#### 🎯 Two-Stage Least Squares (2SLS)")
        
        st.markdown("""
        **Stage 1:** Regress endogenous variable on instrument
        X = π₀ + π₁Z + v
        
        **Stage 2:** Regress Y on predicted X from stage 1
        Y = β₀ + β₁X̂ + ε
        
        **Standard Errors:** Must correct for two-stage procedure
        """)
        
        # Interactive IV demonstration
        st.markdown("#### 🎮 Interactive IV Simulation:")
        
        endogeneity_strength = st.slider("Endogeneity Strength:", 0.0, 1.0, 0.5, 0.1)
        instrument_strength = st.slider("Instrument Strength:", 0.1, 1.0, 0.7, 0.1)
        
        np.random.seed(42)
        n = 1000
        
        # Generate data with endogeneity
        z = np.random.normal(0, 1, n)  # Instrument
        u = np.random.normal(0, 1, n)  # Unobserved confounder
        
        # X is endogenous (correlated with u)
        x = instrument_strength * z + endogeneity_strength * u + np.random.normal(0, 1, n)
        
        # Y depends on X and u (creating endogeneity)
        y = 1 + 2 * x + u + np.random.normal(0, 1, n)  # True effect is 2
        
        # Compare estimates
        # OLS (biased)
        ols_slope = np.cov(x, y)[0,1] / np.var(x)
        
        # 2SLS
        # Stage 1: regress X on Z
        stage1_slope = np.cov(z, x)[0,1] / np.var(z)
        x_predicted = stage1_slope * z
        
        # Stage 2: regress Y on predicted X
        iv_slope = np.cov(x_predicted, y)[0,1] / np.var(x_predicted)
        
        # Show results
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**OLS Results (Biased):**")
            st.metric("Estimated Effect", f"{ols_slope:.3f}")
            st.metric("Bias", f"{ols_slope - 2:.3f}")
            st.metric("First-Stage F-stat", f"{(stage1_slope**2 * np.var(z) / (np.var(x) - stage1_slope**2 * np.var(z))) * (n-2):.1f}")
        
        with col2:
            st.markdown("**IV Results (Consistent):**")
            st.metric("Estimated Effect", f"{iv_slope:.3f}")
            st.metric("Bias", f"{iv_slope - 2:.3f}")
            if (stage1_slope**2 * np.var(z) / (np.var(x) - stage1_slope**2 * np.var(z))) * (n-2) > 10:
                st.success("✅ Strong instrument (F > 10)")
            else:
                st.error("❌ Weak instrument (F < 10)")
        
        # Diagnostics
        st.markdown("#### 🔍 IV Diagnostics")
        
        st.markdown("""
        **Weak Instruments:**
        - F-statistic < 10 in first stage suggests weak instruments
        - Weak instruments lead to large standard errors and bias
        - Rule of thumb: F-statistic > 10
        
        **Overidentification Test:**
        - When you have more instruments than endogenous variables
        - Tests whether instruments are valid (Hansen J-test)
        - Can't test validity with exactly identified models
        """)
    
    elif advanced_topic == "Difference-in-Differences":
        st.markdown("""
        ### 📊 Difference-in-Differences: Natural Experiments
        
        Compare changes over time between treatment and control groups to identify causal effects.
        """)
        
        st.markdown("""
        #### 🎯 The DiD Logic
        
        **Idea:** If treatment and control groups would have followed similar trends absent treatment, then differences in changes identify the treatment effect.
        
        **Model:** Y_{it} = β₀ + β₁(Treatment_Group) + β₂(Post_Period) + β₃(Treatment × Post) + ε_{it}
        
        **Key Parameter:** β₃ = Treatment effect (DiD estimator)
        """)
        
        # DiD assumptions
        st.markdown("#### ⚖️ Key Assumptions")
        
        assumptions_df = pd.DataFrame({
            'Assumption': ['Parallel Trends', 'No Composition Changes', 'Stable Unit Treatment'],
            'Description': [
                'Treatment and control would follow same trend absent treatment',
                'Same units in both periods (no selective migration)',
                'Treatment of one unit does not affect others'
            ],
            'Testable': ['Partially (pre-trends)', 'Yes (check composition)', 'Context-dependent'],
            'Solutions_if_Violated': [
                'Control for different trends', 'Weight by composition', 'Account for spillovers'
            ]
        })
        
        st.dataframe(assumptions_df, use_container_width=True)
        
        # Interactive DiD demonstration
        st.markdown("#### 🎮 Interactive DiD Simulation:")
        
        treatment_effect = st.slider("True Treatment Effect:", -2.0, 3.0, 1.5, 0.1)
        parallel_trends = st.checkbox("Parallel Trends Hold", value=True)
        
        # Generate DiD data
        np.random.seed(42)
        periods = ['Pre', 'Post']
        groups = ['Control', 'Treatment']
        
        # Base values
        control_pre = 10
        treatment_pre = 12  # Treatment group starts higher
        trend = 0.5  # Common trend
        
        if parallel_trends:
            control_post = control_pre + trend
            treatment_post = treatment_pre + trend + treatment_effect
        else:
            # Violation: Different trends
            control_post = control_pre + trend
            treatment_post = treatment_pre + trend * 1.5 + treatment_effect  # Different trend
        
        # Create visualization data
        did_data = pd.DataFrame({
            'Period': ['Pre', 'Post', 'Pre', 'Post'],
            'Group': ['Control', 'Control', 'Treatment', 'Treatment'],
            'Outcome': [control_pre, control_post, treatment_pre, treatment_post]
        })
        
        # Calculate DiD estimate
        control_change = control_post - control_pre
        treatment_change = treatment_post - treatment_pre
        did_estimate = treatment_change - control_change
        
        # Visualization
        fig_did = px.line(did_data, x='Period', y='Outcome', color='Group', 
                         title=f"Difference-in-Differences Design (True Effect = {treatment_effect})")
        fig_did.update_traces(mode='lines+markers', line=dict(width=3), marker=dict(size=8))
        
        # Add counterfactual line
        if not parallel_trends:
            fig_did.add_trace(go.Scatter(
                x=['Pre', 'Post'], 
                y=[treatment_pre, treatment_pre + trend],
                mode='lines',
                line=dict(dash='dash', color='red'),
                name='Treatment Counterfactual (if parallel trends)'
            ))
        
        st.plotly_chart(fig_did, use_container_width=True)
        
        # Results
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**DiD Calculation:**")
            st.write(f"Control Change: {control_change:.2f}")
            st.write(f"Treatment Change: {treatment_change:.2f}")
            st.write(f"**DiD Estimate: {did_estimate:.2f}**")
        
        with col2:
            bias = did_estimate - treatment_effect
            st.metric("Estimation Bias", f"{bias:.2f}")
            if abs(bias) < 0.1:
                st.success("✅ Unbiased estimate")
            elif parallel_trends:
                st.info("ℹ️ Small bias (random variation)")
            else:
                st.error("❌ Biased due to different trends")
        
        # Famous DiD studies
        st.markdown("#### 📚 Famous DiD Studies")
        
        studies_df = pd.DataFrame({
            'Study': ['Card & Krueger (1994)', 'Acemoglu & Angrist (2000)', 'Autor & Duggan (2003)'],
            'Policy': ['Minimum wage increase', 'Compulsory schooling', 'Disability insurance expansion'],
            'Treatment': ['New Jersey', 'States changing laws', 'States with easier access'],
            'Control': ['Pennsylvania', 'States with no change', 'States with no change'],
            'Outcome': ['Employment', 'Education & earnings', 'Disability applications'],
            'Finding': ['No employment reduction', 'Increased education', 'Increased applications']
        })
        
        st.dataframe(studies_df, use_container_width=True)
    
    elif advanced_topic == "Regression Discontinuity":
        st.markdown("""
        ### 📏 Regression Discontinuity: Sharp Cutoffs as Experiments
        
        When treatment assignment has a sharp cutoff, we can identify causal effects by comparing units just above and below the threshold.
        """)
        
        st.markdown("""
        #### 🎯 RD Logic
        
        **Idea:** Units just above and below a cutoff are nearly identical except for treatment status.
        
        **Key Insight:** If assignment is based on an arbitrary cutoff, treatment is "as good as random" near the threshold.
        
        **Types:**
        - **Sharp RD:** All units above cutoff get treatment
        - **Fuzzy RD:** Probability of treatment jumps at cutoff
        """)
        
        # RD requirements
        st.markdown("#### ⚖️ RD Requirements")
        
        rd_requirements_df = pd.DataFrame({
            'Requirement': ['Sharp Cutoff', 'No Manipulation', 'Continuity', 'Local Randomization'],
            'Description': [
                'Clear rule determines treatment assignment',
                'Units cannot precisely control running variable',
                'Other factors change smoothly at cutoff',
                'Units near cutoff are similar'
            ],
            'Example': [
                'Test score > 70 gets scholarship',
                'Students cannot manipulate test scores',
                'Family income varies smoothly around cutoff',
                'Students with scores 69 vs 71 are similar'
            ],
            'Threat': [
                'Fuzzy implementation',
                'Gaming the system',
                'Other policies at same cutoff',
                'Systematic differences'
            ]
        })
        
        st.dataframe(rd_requirements_df, use_container_width=True)
        
        # Interactive RD demonstration
        st.markdown("#### 🎮 Interactive RD Simulation:")
        
        rd_effect = st.slider("True RD Effect:", 0.0, 5.0, 2.0, 0.1)
        noise_level_rd = st.slider("Noise Level:", 0.1, 2.0, 1.0, 0.1)
        manipulation = st.checkbox("Include Manipulation (Violation)", value=False)
        
        # Generate RD data
        np.random.seed(42)
        n_rd = 1000
        cutoff = 50
        
        # Running variable (forcing variable)
        if manipulation:
            # Some gaming - excess mass just above cutoff
            running_var = np.concatenate([
                np.random.normal(45, 10, 400),
                np.random.normal(55, 10, 400),
                np.random.normal(51, 1, 200)  # Manipulation
            ])
        else:
            running_var = np.random.normal(50, 10, n_rd)
        
        # Treatment assignment
        treatment = (running_var >= cutoff).astype(int)
        
        # Outcome with discontinuity
        outcome_rd = (
            2 + 0.05 * running_var +  # Smooth relationship
            rd_effect * treatment +    # Discontinuous jump
            np.random.normal(0, noise_level_rd, len(running_var))
        )
        
        # Create DataFrame
        rd_df = pd.DataFrame({
            'Running_Variable': running_var,
            'Treatment': treatment,
            'Outcome': outcome_rd
        })
        
        # Visualization
        fig_rd = px.scatter(rd_df, x='Running_Variable', y='Outcome', color='Treatment',
                           title="Regression Discontinuity Design")
        
        # Add cutoff line
        fig_rd.add_vline(x=cutoff, line_dash="dash", line_color="red",
                        annotation_text="Cutoff")
        
        # Add regression lines
        below_cutoff = rd_df[rd_df['Running_Variable'] < cutoff]
        above_cutoff = rd_df[rd_df['Running_Variable'] >= cutoff]
        
        if len(below_cutoff) > 1:
            slope_below, intercept_below = np.polyfit(below_cutoff['Running_Variable'], below_cutoff['Outcome'], 1)
            x_below = np.linspace(below_cutoff['Running_Variable'].min(), cutoff, 100)
            y_below = intercept_below + slope_below * x_below
            fig_rd.add_trace(go.Scatter(x=x_below, y=y_below, mode='lines', name='Trend Below', line=dict(color='blue')))
        
        if len(above_cutoff) > 1:
            slope_above, intercept_above = np.polyfit(above_cutoff['Running_Variable'], above_cutoff['Outcome'], 1)
            x_above = np.linspace(cutoff, above_cutoff['Running_Variable'].max(), 100)
            y_above = intercept_above + slope_above * x_above
            fig_rd.add_trace(go.Scatter(x=x_above, y=y_above, mode='lines', name='Trend Above', line=dict(color='red')))
        
        st.plotly_chart(fig_rd, use_container_width=True)
        
        # Estimate RD effect
        bandwidth = 5  # Simple bandwidth
        local_data = rd_df[abs(rd_df['Running_Variable'] - cutoff) <= bandwidth]
        
        if len(local_data) > 10:
            below_mean = local_data[local_data['Running_Variable'] < cutoff]['Outcome'].mean()
            above_mean = local_data[local_data['Running_Variable'] >= cutoff]['Outcome'].mean()
            rd_estimate = above_mean - below_mean
            
            st.metric("RD Estimate", f"{rd_estimate:.3f}")
            st.metric("Estimation Error", f"{rd_estimate - rd_effect:.3f}")
        
        # Density test
        if manipulation:
            st.error("⚠️ **McCrary Test Violation**: Excess mass detected at cutoff - suggests manipulation")
        else:
            st.success("✅ **McCrary Test**: No excess mass at cutoff")
        
        # Famous RD studies
        st.markdown("#### 📚 Famous RD Studies")
        
        rd_studies_df = pd.DataFrame({
            'Study': ['Thistlethwaite & Campbell (1960)', 'Angrist & Lavy (1999)', 'Lee (2008)'],
            'Context': ['Merit scholarships', 'Class size', 'Electoral advantage'],
            'Running_Variable': ['Test scores', 'Enrollment', 'Vote margin'],
            'Cutoff': ['Scholarship threshold', "Maimonides' rule", '50% vote share'],
            'Treatment': ['Scholarship receipt', 'Smaller class', 'Electoral victory'],
            'Finding': ['Positive effect on college', 'Improved test scores', 'Incumbency advantage']
        })
        
        st.dataframe(rd_studies_df, use_container_width=True)
    
    elif advanced_topic == "Time Series Econometrics":
        st.markdown("""
        ### 📈 Time Series Econometrics: Modeling Data Over Time
        
        Special considerations when data has temporal structure and potential non-stationarity.
        """)
        
        st.markdown("""
        #### 🕰️ Key Concepts
        
        **Stationarity:** Statistical properties don't change over time
        - Mean, variance, and covariances are constant
        - Most time series methods assume stationarity
        - Non-stationary series need to be transformed
        
        **Unit Roots:** Series has a stochastic trend
        - Random walk behavior
        - Shocks have permanent effects
        - First differences are often stationary
        """)
        
        # ARIMA models
        st.markdown("#### 📊 ARIMA Models")
        
        st.markdown("""
        **ARIMA(p,d,q):** AutoRegressive Integrated Moving Average
        - **AR(p):** Y_t depends on p lagged values
        - **I(d):** Series is differenced d times to achieve stationarity  
        - **MA(q):** Error term depends on q lagged errors
        
        **Model:** (1 - φ₁L - φ₂L² - ... - φₚLᵖ)(1-L)ᵈYₜ = (1 + θ₁L + θ₂L² + ... + θₑLᵠ)εₜ
        """)
        
        # Interactive time series demonstration
        st.markdown("#### 🎮 Interactive Time Series Demo:")
        
        ts_type = st.selectbox(
            "Choose time series type:",
            ["White Noise", "Random Walk", "AR(1)", "Trend + Cycle", "ARIMA(1,1,1)"]
        )
        
        ts_length = st.slider("Time Series Length:", 50, 500, 200, 50)
        
        np.random.seed(42)
        
        if ts_type == "White Noise":
            ts_data = np.random.normal(0, 1, ts_length)
            description = "Stationary white noise - no predictable patterns"
            
        elif ts_type == "Random Walk":
            ts_data = np.cumsum(np.random.normal(0, 1, ts_length))
            description = "Non-stationary random walk - has unit root"
            
        elif ts_type == "AR(1)":
            phi = 0.7
            ts_data = np.zeros(ts_length)
            ts_data[0] = np.random.normal(0, 1)
            for t in range(1, ts_length):
                ts_data[t] = phi * ts_data[t-1] + np.random.normal(0, 1)
            description = f"AR(1) with φ={phi} - stationary with persistence"
            
        elif ts_type == "Trend + Cycle":
            trend = 0.02 * np.arange(ts_length)
            cycle = 2 * np.sin(2 * np.pi * np.arange(ts_length) / 20)
            noise = np.random.normal(0, 0.5, ts_length)
            ts_data = trend + cycle + noise
            description = "Deterministic trend with cyclical component"
            
        else:  # ARIMA(1,1,1)
            # Generate ARIMA(1,1,1)
            phi = 0.3
            theta = 0.5
            ts_diff = np.zeros(ts_length-1)
            errors = np.random.normal(0, 1, ts_length)
            
            ts_diff[0] = errors[0]
            for t in range(1, ts_length-1):
                ts_diff[t] = phi * ts_diff[t-1] + errors[t] + theta * errors[t-1]
            
            ts_data = np.cumsum(np.concatenate([[0], ts_diff]))
            description = "ARIMA(1,1,1) - integrated process"
        
        # Plot time series
        fig_ts = px.line(x=range(ts_length), y=ts_data, title=f"Time Series: {ts_type}")
        fig_ts.update_layout(xaxis_title="Time", yaxis_title="Value")
        st.plotly_chart(fig_ts, use_container_width=True)
        
        st.info(f"📊 **Description**: {description}")
        
        # Basic statistics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Mean", f"{np.mean(ts_data):.3f}")
        with col2:
            st.metric("Standard Deviation", f"{np.std(ts_data):.3f}")
        with col3:
            # Simple stationarity indicator
            first_half_mean = np.mean(ts_data[:ts_length//2])
            second_half_mean = np.mean(ts_data[ts_length//2:])
            mean_change = abs(second_half_mean - first_half_mean)
            st.metric("Mean Change", f"{mean_change:.3f}")
        
        # Cointegration
        st.markdown("#### 🔗 Cointegration")
        
        st.markdown("""
        **Concept:** Two non-stationary series that move together in the long run
        
        **Economic Interpretation:** 
        - Stock prices and dividends
        - Exchange rates and prices (PPP)
        - Income and consumption
        
        **Engle-Granger Test:**
        1. Test each series for unit root
        2. Regress one on the other
        3. Test residuals for unit root
        4. If residuals are stationary → cointegrated
        
        **Vector Error Correction Model (VECM):**
        - Combines short-run dynamics with long-run equilibrium
        - Error correction mechanism pulls series back to equilibrium
        """)
        
        # VAR models
        st.markdown("#### 🔄 Vector Autoregression (VAR)")
        
        st.markdown("""
        **Multivariate Extension:** Multiple time series influence each other
        
        **VAR(p) Model:**
        Y₁ₜ = c₁ + φ₁₁Y₁ₜ₋₁ + φ₁₂Y₂ₜ₋₁ + ... + ε₁ₜ
        Y₂ₜ = c₂ + φ₂₁Y₁ₜ₋₁ + φ₂₂Y₂ₜ₋₁ + ... + ε₂ₜ
        
        **Applications:**
        - Macroeconomic forecasting
        - Policy analysis
        - Impulse response analysis
        - Granger causality testing
        """)
    
    elif advanced_topic == "Limited Dependent Variables":
        st.markdown("""
        ### 🎯 Limited Dependent Variables: Beyond Linear Models
        
        When the dependent variable is not continuous, we need specialized models.
        """)
        
        # Types of limited dependent variables
        st.markdown("#### 📋 Types of Limited Dependent Variables")
        
        types_df = pd.DataFrame({
            'Type': ['Binary', 'Ordered', 'Multinomial', 'Count', 'Censored', 'Truncated'],
            'Examples': [
                'Pass/Fail, Yes/No',
                'Likert scales, Rankings',
                'Brand choice, Occupation',
                'Number of accidents, Patents',
                'Income (bottom-coded at 0)',
                'High earners only sample'
            ],
            'Model': [
                'Logit, Probit',
                'Ordered Logit/Probit',
                'Multinomial Logit',
                'Poisson, Negative Binomial',
                'Tobit',
                'Truncated Regression'
            ],
            'Key_Feature': [
                'S-shaped probability curve',
                'Parallel regression assumption',
                'Independence of irrelevant alternatives',
                'Variance = mean assumption',
                'Clustered at limit values',
                'Missing below/above threshold'
            ]
        })
        
        st.dataframe(types_df, use_container_width=True)
        
        # Binary choice models
        st.markdown("#### 🔄 Binary Choice Models")
        
        st.markdown("""
        **Problem with Linear Probability Model:**
        - Can predict probabilities outside [0,1]
        - Heteroskedastic errors
        - Non-linear relationship assumed linear
        
        **Solution:** Use link function to transform linear predictor
        """)
        
        # Compare logit and probit
        st.markdown("#### ⚖️ Logit vs Probit")
        
        comparison_binary_df = pd.DataFrame({
            'Aspect': ['Link Function', 'CDF', 'Interpretation', 'Tail Behavior', 'Computation'],
            'Logit': [
                'Log odds: ln(p/(1-p))',
                'Logistic',
                'Odds ratios',
                'Heavier tails',
                'Easier'
            ],
            'Probit': [
                'Inverse normal: Φ⁻¹(p)',
                'Standard normal',
                'Normal quantiles',
                'Lighter tails',
                'Harder'
            ]
        })
        
        st.dataframe(comparison_binary_df, use_container_width=True)
        
        # Interactive binary choice demonstration
        st.markdown("#### 🎮 Interactive Binary Choice Demo:")
        
        model_type = st.selectbox("Choose model:", ["Linear Probability", "Logit", "Probit"])
        sample_size_binary = st.slider("Sample Size:", 100, 1000, 500, 100)
        
        # Generate data
        np.random.seed(42)
        x = np.random.normal(0, 1, sample_size_binary)
        linear_predictor = -0.5 + 1.5 * x
        
        if model_type == "Linear Probability":
            # Linear probability model (can go outside [0,1])
            prob = linear_predictor
            y = (prob + np.random.normal(0, 0.3, sample_size_binary) > 0).astype(int)
        elif model_type == "Logit":
            # Logistic function
            prob = 1 / (1 + np.exp(-linear_predictor))
            y = np.random.binomial(1, prob)
        else:  # Probit
            # Normal CDF
            prob = stats.norm.cdf(linear_predictor)
            y = np.random.binomial(1, prob)
        
        # Create visualization
        binary_df = pd.DataFrame({'X': x, 'Y': y, 'Probability': prob})
        
        # Sort for smooth line
        binary_df_sorted = binary_df.sort_values('X')
        
        fig_binary = px.scatter(binary_df, x='X', y='Y', title=f"Binary Choice Model: {model_type}")
        
        # Add fitted probability line
        fig_binary.add_trace(go.Scatter(
            x=binary_df_sorted['X'], 
            y=binary_df_sorted['Probability'],
            mode='lines',
            name='Fitted Probabilities',
            line=dict(color='red', width=3)
        ))
        
        st.plotly_chart(fig_binary, use_container_width=True)
        
        # Model interpretation
        if model_type == "Logit":
            st.markdown("""
            **Logit Model Interpretation:**
            - Coefficient = change in log-odds for 1-unit increase in X
            - Odds ratio = exp(coefficient)
            - Marginal effect varies by X value (steepest at probability = 0.5)
            """)
        elif model_type == "Probit":
            st.markdown("""
            **Probit Model Interpretation:**
            - Coefficient = change in z-score for 1-unit increase in X
            - Marginal effect = coefficient × φ(Xβ) where φ is standard normal PDF
            - Similar to logit but based on normal distribution
            """)
        else:
            st.markdown("""
            **Linear Probability Model:**
            - Coefficient = change in probability for 1-unit increase in X
            - Easy to interpret but can predict impossible probabilities
            - Problems with heteroskedasticity
            """)
        
        # Count models
        st.markdown("#### 🔢 Count Data Models")
        
        st.markdown("""
        **Poisson Regression:**
        - E[Y|X] = Var[Y|X] = exp(Xβ)
        - Assumes equidispersion (mean = variance)
        - Good for rare events
        
        **Negative Binomial:**
        - Allows overdispersion (variance > mean)
        - More flexible than Poisson
        - Better for data with many zeros
        
        **Zero-Inflated Models:**
        - Two-part process: zero vs non-zero, then count
        - Useful when excess zeros exist
        - Example: Number of doctor visits (many people have zero)
        """)
        
        # Selection models
        st.markdown("#### 🎯 Sample Selection Models")
        
        st.markdown("""
        **Heckman Selection Model:**
        
        **Selection Equation:** Observed if Z_i γ + u_i > 0
        **Outcome Equation:** Y_i = X_i β + ε_i (observed only if selected)
        
        **Applications:**
        - Wage equations (only observe wages for workers)
        - Program evaluation (only observe outcomes for participants)
        - Consumer choice (only observe purchases, not consideration sets)
        
        **Identification:** Need exclusion restriction - variable affecting selection but not outcome
        """)

# Complete the educational materials implementation




# Enhanced main function with landing page
def enhanced_main():
    """Enhanced main function with professional landing page"""
    
    # Initialize session state
    initialize_session_state()
    
    # Show landing page or main app
    if st.session_state.show_landing:
        create_landing_page()
    else:
        main_app_content()

def main_app_content():
    """Main application content"""
    
    # Get language preference
    language = get_language()
    
    # App header with professional styling
    st.markdown("""
    <div style="background: linear-gradient(90deg, #667eea 0%, #764ba2 100%); 
                padding: 1rem; border-radius: 10px; margin-bottom: 2rem;">
        <h1 style="color: white; text-align: center; margin: 0;">
            📊 Econometrics Learning Laboratory
        </h1>
        <p style="color: rgba(255,255,255,0.9); text-align: center; margin: 0.5rem 0 0 0;">
            Created by <strong>HAMDI Boulanouar</strong> | Professional Statistical Analysis Platform
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Add "Return to Landing" button
    if st.sidebar.button("🏠 Return to Landing Page", key="return_landing"):
        st.session_state.show_landing = True
        st.experimental_rerun()
    
    # Enhanced sidebar
    enhanced_sidebar_with_history()
    
    # Navigation
    st.sidebar.title("📋 Navigation")
    pages = [
        "Data Upload & Preview",
        "Regression Analysis", 
        "Model Diagnostics",
        "Advanced Diagnostics",
        "Model Comparison",
        "Time Series Analysis",
        "Predictions",
        "Export Results",
        "Educational Materials"
    ]
    
    selected_page = st.sidebar.selectbox("Choose Analysis Step:", pages)
    
    # Route to appropriate page
    if selected_page == "Data Upload & Preview":
        enhanced_data_upload_page(language)
    elif selected_page == "Regression Analysis":
        regression_analysis_page(language)
    elif selected_page == "Model Diagnostics":
        diagnostics_page(language)
    elif selected_page == "Advanced Diagnostics":
        advanced_diagnostics_page()
    elif selected_page == "Model Comparison":
        model_comparison_page()
    elif selected_page == "Time Series Analysis":
        time_series_features()
    elif selected_page == "Predictions":
        predictions_page(language)
    elif selected_page == "Export Results":
        export_results_feature()
    elif selected_page == "Educational Materials":
        educational_materials_page(language)

# Initialize the app
if __name__ == "__main__":
    try:
        # Use enhanced main with landing page
        enhanced_main()
        
    except Exception as e:
        st.error("❌ Application Error")
        st.exception(e)
        
        # Offer recovery options
        if st.button("🔄 Try to Recover Session"):
            initialize_session_state()
            st.experimental_rerun()
    
    # Professional footer
    if not st.session_state.get('show_landing', True):
        st.markdown("---")
        st.markdown(
            """
            <div style='text-align: center; color: #666; padding: 20px;
                        background: linear-gradient(45deg, #f0f2f6, #e1e8ed);
                        border-radius: 10px; margin-top: 2rem;'>
                <p><strong>📊 Econometrics Learning Laboratory</strong></p>
                <p>Professional Statistical Analysis Platform</p>
                <p>Created with ❤️ by <strong>HAMDI Boulanouar</strong></p>
                <p><em>🎓 Making Advanced Econometrics Accessible to Everyone</em></p>
                <div style='margin-top: 1rem;'>
                    <span style='background: #667eea; color: white; padding: 0.3rem 0.8rem; 
                                 border-radius: 15px; margin: 0.2rem; font-size: 0.8rem;'>
                        📊 Professional Analysis
                    </span>
                    <span style='background: #4ECDC4; color: white; padding: 0.3rem 0.8rem; 
                                 border-radius: 15px; margin: 0.2rem; font-size: 0.8rem;'>
                        🎓 Educational Focus
                    </span>
                    <span style='background: #FF6B6B; color: white; padding: 0.3rem 0.8rem; 
                                 border-radius: 15px; margin: 0.2rem; font-size: 0.8rem;'>
                        🔬 Advanced Features
                    </span>
                </div>
            </div>
            """, 
            unsafe_allow_html=True
        )

